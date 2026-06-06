#!/usr/bin/env python3
import os
import sys
import json
import time
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import pymysql
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "deploy" / "backend"))
os.environ.setdefault("CACHE_DIR", str(ROOT / "output" / "script_cache"))

START_TIME = "2026-05-01 00:00:00"
END_TIME = "2026-06-02 23:59:59"
CHUNK_SIZE = 500
MAX_EXCEL_ROWS = 1_048_000


def load_env():
    env_path = ROOT / "deploy" / ".env"
    for line in env_path.read_text().splitlines():
        if not line.strip() or line.strip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


load_env()
from main import fetch_recipe_process_export_rows  # noqa: E402


def source_conn(database="btyc"):
    return pymysql.connect(
        host=os.environ["SOURCE_DB_HOST"],
        port=int(os.environ.get("SOURCE_DB_PORT", "3306")),
        user=os.environ["SOURCE_DB_USER"],
        password=os.environ["SOURCE_DB_PASS"],
        database=database,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=8,
        read_timeout=180,
        write_timeout=60,
    )


def query_all(sql, args=(), database="btyc"):
    with source_conn(database) as conn, conn.cursor() as cur:
        cur.execute(sql, args)
        return cur.fetchall()


def normalize_value(value):
    if isinstance(value, (datetime,)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def make_writer(wb):
    state = {}

    def write_sheet(base_name, rows, columns):
        sheet_index = 1
        ws = wb.create_sheet(base_name[:31])
        ws.append([title for _, title in columns])
        written = 1
        total = 0
        for row in rows:
            if written >= MAX_EXCEL_ROWS:
                sheet_index += 1
                ws = wb.create_sheet(f"{base_name[:27]}_{sheet_index}"[:31])
                ws.append([title for _, title in columns])
                written = 1
            ws.append([normalize_value(row.get(key)) for key, _ in columns])
            written += 1
            total += 1
        state[base_name] = state.get(base_name, 0) + total
        print(f"  - {base_name}: {total} 行")

    return write_sheet, state


def batched(items, size):
    for index in range(0, len(items), size):
        yield items[index:index + size]


def merge_min(current, value):
    if value is None:
        return current
    return value if current is None or value < current else current


def merge_max(current, value):
    if value is None:
        return current
    return value if current is None or value > current else current


def fetch_recipe_meta(recipe_ids):
    result = {}
    ids = [int(x) for x in recipe_ids if x]
    for chunk in batched(ids, 800):
        placeholders = ",".join(["%s"] * len(chunk))
        rows = query_all(
            f"SELECT id, name, group_name, type FROM manage_backend.main_recipe WHERE id IN ({placeholders})",
            tuple(chunk),
            database="manage_backend",
        )
        for row in rows:
            result[int(row["id"])] = row
    return result


def fetch_device_meta(sns):
    result = {}
    clean_sns = [x for x in sns if x]
    for chunk in batched(clean_sns, 800):
        placeholders = ",".join(["%s"] * len(chunk))
        rows = query_all(
            f"""
            SELECT r.machinecode AS sn,
                   r.name AS device_name,
                   r.robot_type,
                   r.latest_update_package,
                   r.company_id,
                   COALESCE(c.common_name, c.company_name, '未知客户') AS customer_name,
                   COALESCE(c.geo_pname, c.area_code, '') AS province,
                   COALESCE(c.geo_cityname, '') AS city
            FROM btyc.sop_robot r
            LEFT JOIN btyc.ums_company c ON r.company_id = c.id
            WHERE r.machinecode IN ({placeholders})
            """,
            tuple(chunk),
            database="btyc",
        )
        for row in rows:
            result[row["sn"]] = row
    return result


def fetch_all_robot_sns():
    rows = query_all(
        """
        SELECT DISTINCT machinecode AS sn
        FROM btyc.sop_robot
        WHERE machinecode IS NOT NULL AND machinecode != ''
        ORDER BY machinecode
        """,
        database="btyc",
    )
    return [row["sn"] for row in rows if row.get("sn")]


def fetch_usage_by_sn_chunks():
    usage = {}
    sns = fetch_all_robot_sns()
    print(f"设备档案SN候选数: {len(sns)}")
    base_sql = """
        SELECT recipe_id,
               sn,
               MAX(recipe_name) AS log_recipe_name,
               COUNT(*) AS production_count,
               SUM(CASE WHEN whether = 2 THEN 1 ELSE 0 END) AS success_count,
               SUM(CASE WHEN whether = 0 THEN 1 ELSE 0 END) AS cancel_count,
               SUM(CASE WHEN whether NOT IN (0, 2) OR whether IS NULL THEN 1 ELSE 0 END) AS other_status_count,
               MIN(create_time) AS first_time,
               MAX(create_time) AS last_time,
               SUM(CAST(time AS SIGNED)) AS total_duration_seconds
        FROM btyc.sop_machinelog FORCE INDEX (idx_sn_time)
        WHERE sn IN ({placeholders})
          AND create_time >= %s AND create_time <= %s
          AND recipe_id IS NOT NULL AND recipe_id != 0
        GROUP BY recipe_id, sn
    """
    for index, chunk in enumerate(batched(sns, 250), start=1):
        placeholders = ",".join(["%s"] * len(chunk))
        sql = base_sql.format(placeholders=placeholders)
        print(f"  SN候选批次 {index}/{(len(sns) + 249) // 250}，数量 {len(chunk)} ...")
        rows = query_all(sql, tuple(chunk) + (START_TIME, END_TIME), database="btyc")
        for row in rows:
            key = (int(row["recipe_id"]), row["sn"])
            item = usage.setdefault(key, {
                "recipe_id": int(row["recipe_id"]),
                "sn": row["sn"],
                "log_recipe_name": row.get("log_recipe_name"),
                "production_count": 0,
                "success_count": 0,
                "cancel_count": 0,
                "other_status_count": 0,
                "first_time": None,
                "last_time": None,
                "total_duration_seconds": 0,
            })
            item["log_recipe_name"] = item.get("log_recipe_name") or row.get("log_recipe_name")
            item["production_count"] += int(row.get("production_count") or 0)
            item["success_count"] += int(row.get("success_count") or 0)
            item["cancel_count"] += int(row.get("cancel_count") or 0)
            item["other_status_count"] += int(row.get("other_status_count") or 0)
            item["total_duration_seconds"] += int(row.get("total_duration_seconds") or 0)
            item["first_time"] = merge_min(item["first_time"], row.get("first_time"))
            item["last_time"] = merge_max(item["last_time"], row.get("last_time"))
        print(f"    批次聚合行 {len(rows)}，累计菜谱-SN {len(usage)}")
    return list(usage.values())


def build_rows_from_usage(recipe_sn_base_rows):
    recipe_ids = sorted({row["recipe_id"] for row in recipe_sn_base_rows})
    sns = sorted({row["sn"] for row in recipe_sn_base_rows if row.get("sn")})
    print("补充菜谱元数据...")
    recipe_meta = fetch_recipe_meta(recipe_ids)
    print("补充设备/客户元数据...")
    device_meta = fetch_device_meta(sns)

    recipe_summary = {}
    sn_summary = {}
    recipe_sn_rows = []
    recipe_customers = defaultdict(set)
    recipe_sns = defaultdict(set)

    for row in recipe_sn_base_rows:
        rid = row["recipe_id"]
        sn = row.get("sn")
        meta = recipe_meta.get(rid) or {}
        device = device_meta.get(sn) or {}
        recipe_name = meta.get("name") or row.get("log_recipe_name") or f"recipe_{rid}"
        category = meta.get("group_name") or "未分类"
        count = int(row.get("production_count") or 0)
        total_duration = int(row.get("total_duration_seconds") or 0)
        avg_duration = round(total_duration / count, 1) if count else None

        recipe_sns[rid].add(sn)
        if device.get("company_id"):
            recipe_customers[rid].add(device.get("company_id"))

        recipe_sn_rows.append({
            **row,
            "recipe_name": recipe_name,
            "category": category,
            "recipe_type": meta.get("type"),
            "device_name": device.get("device_name"),
            "robot_type": device.get("robot_type"),
            "latest_update_package": device.get("latest_update_package"),
            "customer_name": device.get("customer_name"),
            "province": device.get("province"),
            "city": device.get("city"),
            "avg_duration_seconds": avg_duration,
        })

        recipe = recipe_summary.setdefault(rid, {
            "recipe_id": rid,
            "recipe_name": recipe_name,
            "category": category,
            "recipe_type": meta.get("type"),
            "production_count": 0,
            "success_count": 0,
            "cancel_count": 0,
            "other_status_count": 0,
            "first_time": None,
            "last_time": None,
            "total_duration_seconds": 0,
        })
        recipe["production_count"] += count
        recipe["success_count"] += int(row.get("success_count") or 0)
        recipe["cancel_count"] += int(row.get("cancel_count") or 0)
        recipe["other_status_count"] += int(row.get("other_status_count") or 0)
        recipe["total_duration_seconds"] += total_duration
        recipe["first_time"] = merge_min(recipe["first_time"], row.get("first_time"))
        recipe["last_time"] = merge_max(recipe["last_time"], row.get("last_time"))

        sn_item = sn_summary.setdefault(sn, {
            "sn": sn,
            "device_name": device.get("device_name"),
            "robot_type": device.get("robot_type"),
            "latest_update_package": device.get("latest_update_package"),
            "customer_name": device.get("customer_name"),
            "province": device.get("province"),
            "city": device.get("city"),
            "production_count": 0,
            "recipe_ids": set(),
            "success_count": 0,
            "first_time": None,
            "last_time": None,
        })
        sn_item["production_count"] += count
        sn_item["recipe_ids"].add(rid)
        sn_item["success_count"] += int(row.get("success_count") or 0)
        sn_item["first_time"] = merge_min(sn_item["first_time"], row.get("first_time"))
        sn_item["last_time"] = merge_max(sn_item["last_time"], row.get("last_time"))

    recipe_rows = []
    for rid, row in recipe_summary.items():
        count = max(1, int(row["production_count"]))
        recipe_rows.append({
            **row,
            "sn_count": len(recipe_sns[rid]),
            "customer_count": len(recipe_customers[rid]),
            "avg_duration_seconds": round(int(row["total_duration_seconds"]) / count, 1),
        })
    for row in sn_summary.values():
        row["recipe_count"] = len(row.pop("recipe_ids"))

    return (
        sorted(recipe_rows, key=lambda x: (x["production_count"], x.get("last_time") or ""), reverse=True),
        sorted(recipe_sn_rows, key=lambda x: (x["production_count"], x["recipe_id"], x.get("sn") or ""), reverse=True),
        sorted(sn_summary.values(), key=lambda x: x["production_count"], reverse=True),
    )


def main():
    out_dir = ROOT / "output"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"produced_recipes_full_process_20260501_20260602_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    started = time.time()
    print("按设备SN+create_time索引分片聚合菜谱-SN使用次数...")
    recipe_sn_base_rows = fetch_usage_by_sn_chunks()
    print(f"菜谱-SN基础行数: {len(recipe_sn_base_rows)}")

    recipe_rows, recipe_sn_rows, sn_rows = build_rows_from_usage(recipe_sn_base_rows)
    print(f"菜谱数: {len(recipe_rows)}")
    print(f"菜谱-SN行数: {len(recipe_sn_rows)}")
    print(f"SN数: {len(sn_rows)}")

    wb = Workbook(write_only=True)
    write_sheet, sheet_counts = make_writer(wb)

    write_sheet("口径说明", [
        {"field": "时间范围", "value": f"{START_TIME} 至 {END_TIME}"},
        {"field": "生产记录来源", "value": "btyc.sop_machinelog"},
        {"field": "菜谱定义来源", "value": "manage_backend.main_recipe + manage_backend.recipe_detail"},
        {"field": "食材字典来源", "value": "btyc.base_ingredients"},
        {"field": "纳入口径", "value": "只纳入 recipe_id 非空且不等于 0 的生产记录，确保能关联菜谱过程"},
        {"field": "导出生成时间", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ], [("field", "字段"), ("value", "说明")])

    write_sheet("菜谱汇总", recipe_rows, [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"), ("recipe_type", "菜谱类型"),
        ("production_count", "生产次数"), ("success_count", "成功次数"), ("cancel_count", "取消次数"), ("other_status_count", "其他状态次数"),
        ("sn_count", "使用设备SN数"), ("customer_count", "覆盖客户数"), ("first_time", "首次生产"), ("last_time", "末次生产"),
        ("total_duration_seconds", "累计耗时秒"), ("avg_duration_seconds", "平均耗时秒"),
    ])
    write_sheet("菜谱_SN使用次数", recipe_sn_rows, [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("sn", "设备SN"), ("device_name", "设备名称"), ("robot_type", "设备型号"), ("latest_update_package", "升级包/版本"),
        ("customer_name", "客户名称"), ("province", "省份/区域"), ("city", "城市"),
        ("production_count", "该菜谱在该SN生产次数"), ("success_count", "成功次数"), ("cancel_count", "取消次数"),
        ("first_time", "首次生产"), ("last_time", "末次生产"), ("avg_duration_seconds", "平均耗时秒"),
    ])
    write_sheet("SN汇总", sn_rows, [
        ("sn", "设备SN"), ("device_name", "设备名称"), ("robot_type", "设备型号"), ("latest_update_package", "升级包/版本"),
        ("customer_name", "客户名称"), ("province", "省份/区域"), ("city", "城市"),
        ("production_count", "生产次数"), ("recipe_count", "菜谱数"), ("success_count", "成功次数"),
        ("first_time", "首次生产"), ("last_time", "末次生产"),
    ])

    process_acc = {
        "cook_steps": [],
        "wash_steps": [],
        "moisten_steps": [],
        "ingredients": [],
        "prep_notes": [],
        "serve_notes": [],
        "temperature_curve": [],
        "raw_process": [],
    }
    print("分批读取菜谱执行动作和配料...")
    for idx, chunk in enumerate(batched(recipe_rows, CHUNK_SIZE), start=1):
        process_rows = fetch_recipe_process_export_rows(chunk)
        for key in process_acc:
            process_acc[key].extend(process_rows.get(key, []))
        print(f"  批次 {idx}: 累计烹饪步骤 {len(process_acc['cook_steps'])}，配料 {len(process_acc['ingredients'])}")

    step_columns = [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("execution_count", "时间范围内生产次数"), ("device_count", "设备数"), ("customer_count", "客户数"),
        ("section", "步骤段落"), ("step_index", "步骤序号"), ("time", "时间点秒"), ("cook_time", "菜谱烹饪时间"),
        ("type", "步骤类型原值"), ("type_label", "步骤类型"), ("automatic", "自动原值"), ("automatic_label", "自动/手动"),
        ("power", "功率"), ("speed", "速度"), ("stir", "搅拌"), ("stir_mode", "搅拌模式"), ("mode", "模式"),
        ("position", "锅位"), ("movepot", "翻锅/移锅"), ("direction", "方向"), ("type_operation", "操作类型"),
        ("thedof_time", "录菜总时长"), ("ingredients_time", "投料时间"),
        ("initial_temperature", "初始温度"), ("initial_temperature_array", "初始温度数组"),
        ("commands", "原始指令/投料"), ("execution_content", "机器执行内容"), ("raw_json", "步骤原始JSON"),
    ]
    write_sheet("烹饪执行步骤", process_acc["cook_steps"], step_columns)
    write_sheet("洗锅步骤", process_acc["wash_steps"], step_columns)
    write_sheet("润锅步骤", process_acc["moisten_steps"], step_columns)
    write_sheet("配料明细", process_acc["ingredients"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("execution_count", "时间范围内生产次数"), ("device_count", "设备数"), ("customer_count", "客户数"),
        ("ingredient_index", "配料序号"), ("cooking_step_id", "关联烹饪步骤ID"), ("ingredient_id", "食材ID"),
        ("ingredient_name", "食材名称"), ("dosage", "用量"), ("unit", "单位"), ("preprocess", "预处理"),
        ("feeding_mode", "投料模式"), ("insideand", "insideand"), ("position", "位置"), ("error_dosage", "误差"),
        ("automatic", "自动投料标识"), ("ingredient_type", "食材类型"), ("category_1", "一级分类"), ("category_2", "二级分类"),
        ("raw_json", "配料原始JSON"),
    ])
    write_sheet("备菜须知", process_acc["prep_notes"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("note_index", "序号"), ("content", "内容"), ("raw_json", "原始JSON"),
    ])
    write_sheet("出菜须知", process_acc["serve_notes"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("note_index", "序号"), ("content", "内容"), ("raw_json", "原始JSON"),
    ])
    write_sheet("温度曲线", process_acc["temperature_curve"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"),
        ("point_index", "点位序号"), ("time", "时间点"), ("temperature", "温度"), ("raw_json", "原始JSON"),
    ])
    write_sheet("菜谱原始过程", process_acc["raw_process"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("category", "菜谱分类"), ("recipe_type", "菜谱类型"),
        ("execution_count", "时间范围内生产次数"), ("device_count", "设备数"), ("customer_count", "客户数"),
        ("cook_time", "详情烹饪时间"), ("cook_steps_count", "烹饪步骤数"), ("wash_steps_count", "洗锅步骤数"),
        ("moisten_steps_count", "润锅步骤数"), ("ingredient_count", "配料项数"),
        ("description", "描述"), ("steps_describe", "步骤描述"), ("ingredients_total_dosage", "配料总量"),
        ("ingredient_note_json", "备菜须知JSON"), ("serve_note_json", "出菜须知JSON"), ("temperature_curve_json", "温度曲线JSON"),
        ("initial_temperature", "初始温度"), ("initial_temperature_array", "初始温度数组"),
        ("cook_steps_json", "烹饪步骤JSON"), ("wash_steps_json", "洗锅步骤JSON"), ("moisten_steps_json", "润锅步骤JSON"),
        ("cooking_ingredient_json", "配料JSON"), ("detail_missing", "详情是否缺失"),
    ])

    print("保存Excel...")
    wb.save(out_path)
    print(json.dumps({
        "path": str(out_path),
        "recipe_count": len(recipe_rows),
        "recipe_sn_rows": len(recipe_sn_rows),
        "sn_count": len(sn_rows),
        "sheet_counts": sheet_counts,
        "seconds": round(time.time() - started, 1),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
