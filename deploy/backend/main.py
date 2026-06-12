import os
import pymysql
import json
import time
import base64
import hashlib
import hmac
import sqlite3
import threading
import re
import zipfile
import urllib.request
import urllib.error
import statistics
import gc
import uuid
from bisect import bisect_left, bisect_right
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from fastapi import FastAPI, HTTPException, Header, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = FastAPI(title="Zhiku Device Dashboard API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASS', ''),
    'database': os.getenv('DB_NAME', 'zhiku_db'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

SOURCE_DB_CONFIG = {
    'host': os.getenv('SOURCE_DB_HOST', ''),
    'port': int(os.getenv('SOURCE_DB_PORT', '3306')),
    'user': os.getenv('SOURCE_DB_USER', ''),
    'password': os.getenv('SOURCE_DB_PASS', ''),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    'connect_timeout': 8,
    'read_timeout': 30,
    'write_timeout': 30,
}

ADMIN_TOKEN = os.getenv('ADMIN_TOKEN', 'change_me_admin_token')
MAX_DEVICE_LOGS = int(os.getenv('MAX_DEVICE_LOGS', '5000'))
INLINE_LOG_LIMIT = int(os.getenv('INLINE_LOG_LIMIT', '1000'))
CACHE_TTL_SECONDS = int(os.getenv('CACHE_TTL_SECONDS', '300'))
DEVICE_LOOKUP_CACHE_TTL_SECONDS = int(os.getenv('DEVICE_LOOKUP_CACHE_TTL_SECONDS', '600'))
VERSION_STATS_CACHE_TTL_SECONDS = int(os.getenv('VERSION_STATS_CACHE_TTL_SECONDS', '600'))
RECIPE_SEARCH_CACHE_TTL_SECONDS = int(os.getenv('RECIPE_SEARCH_CACHE_TTL_SECONDS', '21600'))
MAX_RECIPE_SEARCH_DAYS = int(os.getenv('MAX_RECIPE_SEARCH_DAYS', '180'))
MAX_RECIPE_SEARCH_LIMIT = int(os.getenv('MAX_RECIPE_SEARCH_LIMIT', '300'))
CACHE_DIR = Path(os.getenv('CACHE_DIR', '/app/cache'))
OIL_THERMAL_DATA_DIR = Path(os.getenv('OIL_THERMAL_DATA_DIR', '/app/data'))
FAULT_CODE_DATA_PATH = OIL_THERMAL_DATA_DIR / 'fault_codes_latest.json'
MAX_LOG_ANALYSIS_DOWNLOAD_MB = int(os.getenv('MAX_LOG_ANALYSIS_DOWNLOAD_MB', '80'))
AUTH_SECRET = os.getenv('AUTH_SECRET', ADMIN_TOKEN)
SESSION_TTL_SECONDS = int(os.getenv('SESSION_TTL_SECONDS', str(7 * 24 * 3600)))
REPORT_CACHE = {}
DEVICE_LOOKUP_CACHE = {}
VERSION_STATS_CACHE = {}
LOG_ANALYSIS_VERSION = 2
LOG_PACKAGE_DIAGNOSTICS_VERSION = 3
CACHE_DIR.mkdir(parents=True, exist_ok=True)
LOG_EVIDENCE_FILE_CACHE_DIR = CACHE_DIR / 'log_evidence_files'
LOG_EVIDENCE_FILE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
AUDIT_DB_PATH = CACHE_DIR / 'zhiku_audit.sqlite3'
AUDIT_LOCK = threading.Lock()
ANALYTICS_DB_READY = False
ANALYTICS_DB_LOCK = threading.Lock()
AUTO_PARSE_WORKER_STARTED = False
AUTO_PARSE_WORKER_LOCK = threading.Lock()
AUTO_PARSE_ENABLED = os.getenv('AUTO_PARSE_ENABLED', '1') == '1'
AUTO_PARSE_INTERVAL_SECONDS = int(os.getenv('AUTO_PARSE_INTERVAL_SECONDS', '90'))
AUTO_PARSE_MAX_PACKAGES_PER_CYCLE = int(os.getenv('AUTO_PARSE_MAX_PACKAGES_PER_CYCLE', '2'))
STRUCTURED_DAY_DETAIL_LIMIT = int(os.getenv('STRUCTURED_DAY_DETAIL_LIMIT', '20'))
AUTO_PARSE_STALE_MINUTES = int(os.getenv('AUTO_PARSE_STALE_MINUTES', '12'))
LOG_EVENT_SCAN_VERSION = 3
LOG_EVENT_SCAN_WORKERS = {}
LOG_EVENT_SCAN_WORKERS_LOCK = threading.Lock()

LOG_KIND_DEFS = {
    'oildrum_board': {
        'label': '猪油桶板日志',
        'purpose': '油温、油管温度、投料电机、油桶加热 PWM、加热目标和功率限制。',
        'key_patterns': ['dev status', 'pump hall start', 'pump hall finish', 'read power', 'oil heat enable', 'pipe heat enable', 'read sw v'],
    },
    'temperature': {
        'label': '温度采样日志',
        'purpose': '主控温度采样，用于辅助看整机温度变化。',
        'key_patterns': ['temperature triplet'],
    },
    'main_board': {
        'label': '主控板日志',
        'purpose': '主控动作、加油/加水/翻锅/温控等过程日志。',
        'key_patterns': ['add oil', 'add water', 'lean', 'tempctr', 'fail'],
    },
    'mcu_debug': {
        'label': 'MCU 调试日志',
        'purpose': 'MCU 通讯、温控、错误码查询、超时等底层调试信息。',
        'key_patterns': ['send to mcu', 'heaterGetErrCode', 'timeout', 'error'],
    },
    'android_app': {
        'label': '安卓应用日志',
        'purpose': 'App 指令、资源动作、网络、数据采集和业务流程。',
        'key_patterns': ['sendMsg', 'readResult', 'SAUCE_STARCH', 'SAUCE_NEW', 'DataCollectManager'],
    },
    'other': {
        'label': '其他文件',
        'purpose': '暂未建立专用解析器的日志或附件。',
        'key_patterns': [],
    },
}

OILDRUM_FIELD_SPECS = [
    {
        'category': '状态数据',
        'pattern': 'dev status',
        'fields': [
            '时间',
            '油温度 * 10',
            '油管温度 * 10',
            '预留',
            '电机速度 RPM',
            '油桶加热 PWM 占空比',
        ],
        'example': '[26.05.10 17:36:22] dev status 506 607 0 468 2.80',
        'meaning': '示例表示油温 50.6℃、油管温度 60.7℃、投料电机 468RPM、PWM 占空比 2.8%。',
    },
    {
        'category': '电机控制开始',
        'pattern': 'pump hall start',
        'fields': ['时间', '霍尔数量，负数投油，正数反抽', '控制电机占空比'],
        'example': '[26.05.10 17:39:51] pump hall start -2476 100',
        'meaning': '示例表示开始投油，目标霍尔数 2476，电机占空比 100%。',
    },
    {
        'category': '电机控制完成',
        'pattern': 'pump hall finish',
        'fields': ['时间', '实际霍尔数量', '本次电机控制实际时间 ms', '霍尔缺相检测，1 异常，0 正常'],
        'example': '[26.05.10 17:40:05] pump hall finish 2477,time 14738,state 1',
        'meaning': '示例表示电机工作 14.738 秒，霍尔缺相检测异常。',
    },
    {
        'category': '功率限制',
        'pattern': 'read power',
        'fields': ['时间', '功率限制状态，1 有限制，0 无限制', '预留', '预留'],
        'example': '[26.05.10 17:40:36] read power 1 0 0',
        'meaning': '示例表示当前存在加热功率限制。',
    },
    {
        'category': '油桶加热目标',
        'pattern': 'oil heat enable',
        'fields': ['时间', '加热状态，1 使能，0 不使能', '目标温度'],
        'example': '[26.04.08 18:19:35] oil heat enable 1 70',
        'meaning': '示例表示使能猪油桶加热，目标温度 70℃。',
    },
    {
        'category': '油管加热目标',
        'pattern': 'pipe heat enable',
        'fields': ['时间', '加热状态，1 使能，0 不使能', '目标温度'],
        'example': '[26.04.08 18:19:37] pipe heat enable 1 70',
        'meaning': '示例表示使能油管加热，目标温度 70℃。',
    },
    {
        'category': '软件版本',
        'pattern': 'read sw v',
        'fields': ['时间', '版本号高位', '版本号中位', '版本号低位'],
        'example': '[26.04.08 18:26:57] read sw v 1 0 11',
        'meaning': '示例表示猪油桶板软件版本 1.0.11。',
    },
]

def get_conn():
    return pymysql.connect(**DB_CONFIG)

def get_source_conn(database=None):
    if not SOURCE_DB_CONFIG['host']:
        raise HTTPException(status_code=500, detail="Source database is not configured")
    config = dict(SOURCE_DB_CONFIG)
    if database:
        config['database'] = database
    return pymysql.connect(**config)

class DeviceReport(BaseModel):
    info: dict
    stats: dict
    logs: list
    intervals: list
    recipes: list
    faults: list
    maintenance: list

class LoginRequest(BaseModel):
    username: str
    password: str

class ProfileUpdate(BaseModel):
    display_name: str = ''

class RecipeTopJobRequest(BaseModel):
    start_date: str = ''
    end_date: str = ''
    top_n: int = 500
    sort_by: str = 'cooking_count'
    recipe_keyword: str = ''
    customer_keyword: str = ''
    sn: str = ''
    region: str = ''
    category: str = ''
    resource_type: str = ''
    stat_object: str = 'all'
    refresh: bool = False

def parse_users():
    raw = os.getenv('ZHIKU_USERS', f"admin:{ADMIN_TOKEN}")
    users = {}
    for item in raw.split(','):
        if ':' not in item:
            continue
        parts = item.split(':')
        username = parts[0].strip()
        password = parts[1].strip() if len(parts) > 1 else ''
        role = parts[2].strip() if len(parts) > 2 else ('admin' if username == 'admin' else 'user')
        username = username.strip()
        if username:
            users[username] = {'password': password, 'role': role}
    return users

ZHIKU_USERS = parse_users()

def audit_conn():
    conn = sqlite3.connect(AUDIT_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_audit_db():
    with AUDIT_LOCK, audit_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_profiles (
                username TEXT PRIMARY KEY,
                display_name TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'user',
                created_at INTEGER NOT NULL,
                updated_at INTEGER NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                action TEXT NOT NULL,
                sn TEXT,
                detail TEXT,
                ip TEXT,
                user_agent TEXT,
                created_at INTEGER NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_user_time ON audit_events(username, created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_action_time ON audit_events(action, created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_sn_time ON audit_events(sn, created_at)")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS analytics_query_jobs (
                job_id TEXT PRIMARY KEY,
                query_type TEXT NOT NULL,
                params_json TEXT NOT NULL,
                params_hash TEXT NOT NULL,
                status TEXT NOT NULL,
                stage TEXT NOT NULL DEFAULT '',
                progress INTEGER NOT NULL DEFAULT 0,
                created_by TEXT NOT NULL,
                created_at INTEGER NOT NULL,
                started_at INTEGER,
                finished_at INTEGER,
                cache_expires_at INTEGER,
                result_path TEXT,
                xlsx_path TEXT,
                error_message TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_analytics_hash ON analytics_query_jobs(query_type, params_hash, status, cache_expires_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_analytics_created ON analytics_query_jobs(created_at)")
        now = int(time.time())
        for username, meta in ZHIKU_USERS.items():
            conn.execute(
                "INSERT OR IGNORE INTO user_profiles(username, role, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (username, meta.get('role', 'user'), now, now),
            )
            conn.execute(
                "UPDATE user_profiles SET role = ?, updated_at = CASE WHEN role != ? THEN ? ELSE updated_at END WHERE username = ?",
                (meta.get('role', 'user'), meta.get('role', 'user'), now, username),
            )
        conn.commit()

init_audit_db()

def sign_token(username, expires_at):
    payload = f"{username}:{int(expires_at)}"
    signature = hmac.new(AUTH_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(f"{payload}:{signature}".encode()).decode()

def verify_token(token):
    try:
        decoded = base64.urlsafe_b64decode(token.encode()).decode()
        username, expires_at, signature = decoded.rsplit(':', 2)
        payload = f"{username}:{expires_at}"
        expected = hmac.new(AUTH_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(expires_at) < int(time.time()):
            return None
        if username not in ZHIKU_USERS:
            return None
        return username
    except Exception:
        return None

def require_auth(authorization: str = None, token: str = None):
    raw = token
    if authorization and authorization.lower().startswith('bearer '):
        raw = authorization.split(' ', 1)[1].strip()
    username = verify_token(raw or '')
    if not username:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return username

def is_admin(username):
    return ZHIKU_USERS.get(username, {}).get('role') == 'admin'

def require_admin(authorization: str = None, token: str = None):
    username = require_auth(authorization=authorization, token=token)
    if not is_admin(username):
        raise HTTPException(status_code=403, detail="Admin only")
    return username

def row_to_dict(row):
    return dict(row) if row else None

def get_profile(username):
    with AUDIT_LOCK, audit_conn() as conn:
        row = conn.execute(
            "SELECT username, display_name, role, created_at, updated_at FROM user_profiles WHERE username = ?",
            (username,),
        ).fetchone()
    profile = row_to_dict(row) or {
        'username': username,
        'display_name': '',
        'role': ZHIKU_USERS.get(username, {}).get('role', 'user'),
        'created_at': int(time.time()),
        'updated_at': int(time.time()),
    }
    profile['profile_required'] = not bool(profile.get('display_name'))
    return profile

def update_profile(username, display_name):
    clean_name = (display_name or '').strip()[:40]
    now = int(time.time())
    with AUDIT_LOCK, audit_conn() as conn:
        conn.execute(
            "INSERT INTO user_profiles(username, display_name, role, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(username) DO UPDATE SET display_name = excluded.display_name, updated_at = excluded.updated_at",
            (username, clean_name, ZHIKU_USERS.get(username, {}).get('role', 'user'), now, now),
        )
        conn.commit()
    return get_profile(username)

def request_ip(request: Request = None):
    if not request or not request.client:
        return ''
    forwarded = request.headers.get('x-forwarded-for')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.client.host or ''

def log_event(username, action, request: Request = None, sn=None, detail=None):
    payload = json.dumps(detail or {}, ensure_ascii=False, default=str)
    user_agent = request.headers.get('user-agent', '')[:240] if request else ''
    now = int(time.time())
    with AUDIT_LOCK, audit_conn() as conn:
        conn.execute(
            "INSERT INTO audit_events(username, action, sn, detail, ip, user_agent, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (username, action, sn, payload, request_ip(request), user_agent, now),
        )
        conn.commit()

def audit_summary():
    now = int(time.time())
    day_ago = now - 86400
    week_ago = now - 7 * 86400
    with AUDIT_LOCK, audit_conn() as conn:
        users = [dict(row) for row in conn.execute("""
            SELECT p.username, p.display_name, p.role,
                   COALESCE(SUM(CASE WHEN e.action = 'login' THEN 1 ELSE 0 END), 0) AS login_count,
                   COALESCE(SUM(CASE WHEN e.action IN ('search', 'search_refresh') THEN 1 ELSE 0 END), 0) AS search_count,
                   COALESCE(SUM(CASE WHEN e.action IN ('recipe_search', 'recipe_search_refresh') THEN 1 ELSE 0 END), 0) AS recipe_search_count,
                   COALESCE(SUM(CASE WHEN e.action = 'export' THEN 1 ELSE 0 END), 0) AS export_count,
                   COALESCE(SUM(CASE WHEN e.action = 'recipe_search_export' THEN 1 ELSE 0 END), 0) AS recipe_export_count,
                   COALESCE(SUM(CASE WHEN e.action = 'log_download' THEN 1 ELSE 0 END), 0) AS log_download_count,
                   COALESCE(SUM(CASE WHEN e.action IN ('log_analysis_view', 'admin_log_analysis_view') THEN 1 ELSE 0 END), 0) AS log_analysis_count,
                   COALESCE(SUM(CASE WHEN e.created_at >= ? THEN 1 ELSE 0 END), 0) AS events_24h,
                   MAX(e.created_at) AS last_seen
            FROM user_profiles p
            LEFT JOIN audit_events e ON e.username = p.username
            GROUP BY p.username, p.display_name, p.role
            ORDER BY events_24h DESC, last_seen DESC
        """, (day_ago,))]
        recent_events = [dict(row) for row in conn.execute("""
            SELECT id, username, action, sn, detail, ip, user_agent, created_at
            FROM audit_events
            ORDER BY created_at DESC
            LIMIT 120
        """)]
        top_devices = [dict(row) for row in conn.execute("""
            SELECT sn, COUNT(*) AS hits, COUNT(DISTINCT username) AS user_count, MAX(created_at) AS last_seen
            FROM audit_events
            WHERE sn IS NOT NULL AND sn != ''
            GROUP BY sn
            ORDER BY hits DESC
            LIMIT 30
        """)]
        ip_summary = [dict(row) for row in conn.execute("""
            SELECT username, COUNT(DISTINCT ip) AS ip_count
            FROM audit_events
            WHERE created_at >= ? AND ip IS NOT NULL AND ip != ''
            GROUP BY username
            HAVING ip_count >= 4
            ORDER BY ip_count DESC
        """, (week_ago,))]
        recipe_keyword_events = [dict(row) for row in conn.execute("""
            SELECT detail, COUNT(*) AS hits, MAX(created_at) AS last_seen
            FROM audit_events
            WHERE action IN ('recipe_search', 'recipe_search_refresh', 'recipe_search_export')
            GROUP BY detail
            ORDER BY hits DESC
            LIMIT 80
        """)]

    alerts = []
    for user in users:
        if user['log_download_count'] >= 20:
            alerts.append({'level': 'high', 'username': user['username'], 'message': '日志下载次数较高，请确认是否为排障需要'})
        if user['events_24h'] >= 80:
            alerts.append({'level': 'medium', 'username': user['username'], 'message': '24 小时内操作频次较高'})
        if not user.get('display_name') and user['role'] != 'admin' and (user['login_count'] or user['search_count']):
            alerts.append({'level': 'low', 'username': user['username'], 'message': '已使用但尚未填写真实姓名'})
    for row in ip_summary:
        alerts.append({'level': 'medium', 'username': row['username'], 'message': f"近 7 天出现 {row['ip_count']} 个不同 IP"})

    for event in recent_events:
        try:
            event['detail'] = json.loads(event.get('detail') or '{}')
        except json.JSONDecodeError:
            event['detail'] = {}
    recipe_keywords = []
    for row in recipe_keyword_events:
        try:
            detail = json.loads(row.get('detail') or '{}')
        except json.JSONDecodeError:
            detail = {}
        label_parts = [detail.get('scope') or 'recipe']
        if detail.get('keyword'):
            label_parts.append(detail.get('keyword'))
        if detail.get('recipe_keyword'):
            label_parts.append(detail.get('recipe_keyword'))
        recipe_keywords.append({
            'label': ' / '.join(label_parts),
            'scope': detail.get('scope'),
            'keyword': detail.get('keyword'),
            'recipe_keyword': detail.get('recipe_keyword'),
            'hits': row.get('hits'),
            'last_seen': row.get('last_seen'),
            'cache_hit': detail.get('cache_hit'),
            'total_logs': detail.get('total_logs'),
        })

    return {
        'users': users,
        'recent_events': recent_events,
        'top_devices': top_devices,
        'recipe_keywords': recipe_keywords[:30],
        'alerts': alerts[:50],
        'totals': {
            'user_count': len(users),
            'events_24h': sum(int(u.get('events_24h') or 0) for u in users),
            'search_count': sum(int(u.get('search_count') or 0) for u in users),
            'recipe_search_count': sum(int(u.get('recipe_search_count') or 0) for u in users),
            'export_count': sum(int(u.get('export_count') or 0) for u in users),
            'recipe_export_count': sum(int(u.get('recipe_export_count') or 0) for u in users),
            'log_download_count': sum(int(u.get('log_download_count') or 0) for u in users),
            'log_analysis_count': sum(int(u.get('log_analysis_count') or 0) for u in users),
        },
    }

def seconds_label(seconds):
    total = int(seconds or 0)
    hours = total // 3600
    minutes = (total % 3600) // 60
    if hours:
        return f"{hours}小时{minutes}分"
    return f"{minutes}分"

def emergency_device_summary(sns):
    result = []
    for input_sn in sns:
        real_sn, info = resolve_sn(input_sn)
        customer = get_company_info(info.get('company_id'))
        daily_rows = fetch_all(
            "SELECT DATE(create_time) AS day, COUNT(*) AS production_count, "
            "COUNT(DISTINCT recipe_id) AS recipe_count, MIN(create_time) AS first_time, "
            "MAX(create_time) AS last_time, "
            "TIMESTAMPDIFF(SECOND, MIN(create_time), MAX(create_time)) AS span_seconds, "
            "SUM(CAST(time AS SIGNED)) AS cook_seconds "
            "FROM sop_machinelog WHERE sn = %s "
            "GROUP BY day ORDER BY day",
            (real_sn,),
            source=True,
            database='btyc',
        )
        total_span_seconds = 0
        max_span_seconds = 0
        for row in daily_rows:
            row['span_seconds'] = int(row.get('span_seconds') or 0)
            row['cook_seconds'] = int(row.get('cook_seconds') or 0)
            row['span_label'] = seconds_label(row['span_seconds'])
            row['cook_label'] = seconds_label(row['cook_seconds'])
            total_span_seconds += row['span_seconds']
            max_span_seconds = max(max_span_seconds, row['span_seconds'])
        result.append({
            'sn': real_sn,
            'input_sn': input_sn,
            'customer': customer,
            'info': {
                'company_id': info.get('company_id'),
                'robot_type': info.get('robot_type'),
                'hardware_version': info.get('hardware_version'),
                'latest_update_package': info.get('latest_update_package'),
                'last_online_time': info.get('last_online_time'),
            },
            'usage_days': len(daily_rows),
            'total_span_seconds': total_span_seconds,
            'total_span_label': seconds_label(total_span_seconds),
            'max_daily_span_seconds': max_span_seconds,
            'max_daily_span_label': seconds_label(max_span_seconds),
            'total_production_count': sum(int(row.get('production_count') or 0) for row in daily_rows),
            'daily': daily_rows,
        })
    return result

@app.post("/api/login")
def login(payload: LoginRequest, request: Request):
    user_meta = ZHIKU_USERS.get(payload.username)
    expected = user_meta.get('password') if user_meta else None
    if expected is None or not hmac.compare_digest(expected, payload.password):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    expires_at = int(time.time()) + SESSION_TTL_SECONDS
    log_event(payload.username, 'login', request)
    return {
        "token": sign_token(payload.username, expires_at),
        "username": payload.username,
        "expires_at": expires_at,
        "role": user_meta.get('role', 'user'),
        "profile": get_profile(payload.username),
    }

def fetch_all(query, args=None, source=False, database=None):
    conn = get_source_conn(database) if source else get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(query, args)
            return cur.fetchall()
    finally:
        conn.close()

def fetch_one(query, args=None, source=False, database=None):
    rows = fetch_all(query, args, source=source, database=database)
    return rows[0] if rows else None

def execute_local(query, args=None):
    ensure_analytics_db()
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(query, args)
        conn.commit()
    finally:
        conn.close()

def executemany_local(query, rows):
    if not rows:
        return
    ensure_analytics_db()
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.executemany(query, rows)
        conn.commit()
    finally:
        conn.close()

def init_analytics_db():
    global ANALYTICS_DB_READY
    with ANALYTICS_DB_LOCK:
        if ANALYTICS_DB_READY:
            return True
        conn = get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS watched_devices (
                        sn VARCHAR(64) PRIMARY KEY,
                        status VARCHAR(32) NOT NULL DEFAULT 'active',
                        priority INT NOT NULL DEFAULT 50,
                        first_seen_at DATETIME NOT NULL,
                        last_seen_at DATETIME NOT NULL,
                        last_sync_at DATETIME NULL,
                        last_parse_at DATETIME NULL,
                        last_error TEXT NULL,
                        created_by VARCHAR(64) NULL,
                        updated_at DATETIME NOT NULL,
                        INDEX idx_watched_status_priority(status, priority, last_seen_at)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS device_log_packages (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        file_name VARCHAR(255) NULL,
                        file_size_mb DECIMAL(12,3) NULL,
                        file_size_label VARCHAR(32) NULL,
                        remote_url_hash CHAR(64) NULL,
                        log_time_hint DATETIME NULL,
                        remote_create_time DATETIME NULL,
                        remote_update_time DATETIME NULL,
                        cos_deleted TINYINT NOT NULL DEFAULT 0,
                        download_status VARCHAR(32) NOT NULL DEFAULT 'remote_available',
                        parse_status VARCHAR(32) NOT NULL DEFAULT 'not_started',
                        storage_status VARCHAR(32) NOT NULL DEFAULT 'not_stored',
                        ui_status VARCHAR(32) NOT NULL DEFAULT '可下载',
                        parse_version INT NOT NULL DEFAULT 0,
                        log_start_time DATETIME NULL,
                        log_end_time DATETIME NULL,
                        cook_count INT NOT NULL DEFAULT 0,
                        sample_count INT NOT NULL DEFAULT 0,
                        error_message TEXT NULL,
                        parse_attempts INT NOT NULL DEFAULT 0,
                        last_attempt_at DATETIME NULL,
                        parsed_at DATETIME NULL,
                        stored_at DATETIME NULL,
                        created_at DATETIME NOT NULL,
                        updated_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_log_package(sn, source_file_id),
                        INDEX idx_log_sn_time(sn, remote_create_time),
                        INDEX idx_log_status(parse_status, storage_status),
                        INDEX idx_log_coverage(sn, log_start_time, log_end_time)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cook_jobs (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        machine_log_id BIGINT NULL,
                        recipe_id BIGINT NULL,
                        recipe_name VARCHAR(255) NULL,
                        cook_start_time DATETIME NULL,
                        cook_end_time DATETIME NULL,
                        duration_seconds INT NULL,
                        max_pot_temp DECIMAL(8,2) NULL,
                        avg_pot_temp DECIMAL(8,2) NULL,
                        sample_count INT NOT NULL DEFAULT 0,
                        step_count INT NOT NULL DEFAULT 0,
                        android_action_count INT NOT NULL DEFAULT 0,
                        payload_json LONGTEXT NULL,
                        parse_version INT NOT NULL,
                        created_at DATETIME NOT NULL,
                        updated_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_cook_job(sn, source_file_id, machine_log_id),
                        INDEX idx_cook_sn_time(sn, cook_start_time),
                        INDEX idx_cook_recipe(recipe_id, recipe_name),
                        INDEX idx_cook_temp(max_pot_temp)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cook_temperature_samples (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        cook_job_id BIGINT NOT NULL,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        sample_time DATETIME NULL,
                        offset_seconds INT NULL,
                        source VARCHAR(64) NULL,
                        pot_temp DECIMAL(8,2) NULL,
                        filtered_temp DECIMAL(8,2) NULL,
                        infrared_temp DECIMAL(8,2) NULL,
                        output_temp DECIMAL(8,2) NULL,
                        core_temp DECIMAL(8,2) NULL,
                        coil_temp DECIMAL(8,2) NULL,
                        raw_value VARCHAR(120) NULL,
                        is_peak TINYINT NOT NULL DEFAULT 0,
                        created_at DATETIME NOT NULL,
                        INDEX idx_temp_cook(cook_job_id, offset_seconds),
                        INDEX idx_temp_sn_time(sn, sample_time),
                        INDEX idx_temp_peak(sn, is_peak, pot_temp)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cook_action_events (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        cook_job_id BIGINT NOT NULL,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        event_time DATETIME NULL,
                        offset_seconds INT NULL,
                        event_type VARCHAR(64) NULL,
                        event_label VARCHAR(120) NULL,
                        raw_log_excerpt TEXT NULL,
                        command_power_kw DECIMAL(10,3) NULL,
                        actual_power_kw DECIMAL(10,3) NULL,
                        core_temp DECIMAL(8,2) NULL,
                        coil_temp DECIMAL(8,2) NULL,
                        output_temp DECIMAL(8,2) NULL,
                        created_at DATETIME NOT NULL,
                        INDEX idx_action_cook(cook_job_id, offset_seconds),
                        INDEX idx_action_sn_time(sn, event_time),
                        INDEX idx_action_type(event_type, event_label)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cook_power_events (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        cook_job_id BIGINT NOT NULL,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        event_time DATETIME NULL,
                        offset_seconds INT NULL,
                        command_power_kw DECIMAL(10,3) NULL,
                        actual_power_kw DECIMAL(10,3) NULL,
                        command_power_w DECIMAL(12,2) NULL,
                        actual_power_w DECIMAL(12,2) NULL,
                        raw_value VARCHAR(120) NULL,
                        created_at DATETIME NOT NULL,
                        INDEX idx_power_cook(cook_job_id, offset_seconds),
                        INDEX idx_power_sn_time(sn, event_time)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS device_log_event_scans (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        scan_type VARCHAR(64) NOT NULL,
                        scan_status VARCHAR(32) NOT NULL DEFAULT 'queued',
                        scan_version INT NOT NULL DEFAULT 0,
                        event_count INT NOT NULL DEFAULT 0,
                        error_message TEXT NULL,
                        attempts INT NOT NULL DEFAULT 0,
                        queued_at DATETIME NULL,
                        started_at DATETIME NULL,
                        finished_at DATETIME NULL,
                        updated_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_device_log_event_scan(sn, source_file_id, scan_type),
                        INDEX idx_event_scan_queue(scan_type, scan_status, queued_at),
                        INDEX idx_event_scan_sn(sn, scan_type, scan_status)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS device_log_events (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NOT NULL,
                        event_category VARCHAR(64) NOT NULL,
                        event_type VARCHAR(64) NOT NULL,
                        event_label VARCHAR(120) NOT NULL,
                        event_time DATETIME NULL,
                        matched_keyword VARCHAR(120) NULL,
                        source_log_name VARCHAR(255) NULL,
                        line_no INT NULL,
                        line_hash CHAR(64) NULL,
                        raw_log_excerpt TEXT NULL,
                        evidence_level VARCHAR(16) NOT NULL DEFAULT '中',
                        event_hash CHAR(64) NOT NULL,
                        scan_version INT NOT NULL,
                        created_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_device_log_event(event_hash),
                        INDEX idx_device_event_sn_time(sn, event_time),
                        INDEX idx_device_event_category(event_category, event_time),
                        INDEX idx_device_event_package(sn, source_file_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                for stmt in [
                    "ALTER TABLE device_log_events ADD COLUMN line_no INT NULL AFTER source_log_name",
                    "ALTER TABLE device_log_events ADD COLUMN line_hash CHAR(64) NULL AFTER line_no",
                    "ALTER TABLE device_log_events ADD COLUMN evidence_level VARCHAR(16) NOT NULL DEFAULT '中' AFTER raw_log_excerpt",
                    "ALTER TABLE device_log_events ADD INDEX idx_device_event_line_hash(line_hash)",
                ]:
                    try:
                        cur.execute(stmt)
                    except Exception:
                        pass
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS ingredient_thermal_properties (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        source_ingredient_id VARCHAR(128) NULL,
                        canonical_name VARCHAR(255) NOT NULL,
                        aliases_json TEXT NULL,
                        category VARCHAR(64) NOT NULL DEFAULT '未分类',
                        source_category_1 VARCHAR(128) NULL,
                        source_category_2 VARCHAR(128) NULL,
                        ingredient_type VARCHAR(32) NULL,
                        automatic VARCHAR(32) NULL,
                        specific_heat_kj_kg_c DECIMAL(8,3) NULL,
                        water_fraction DECIMAL(5,3) NULL,
                        oil_fraction DECIMAL(5,3) NULL,
                        boiling_c DECIMAL(8,2) NULL,
                        smoke_point_c DECIMAL(8,2) NULL,
                        flash_point_c DECIMAL(8,2) NULL,
                        autoignition_c DECIMAL(8,2) NULL,
                        hazard_class VARCHAR(64) NOT NULL DEFAULT '待归类',
                        confidence VARCHAR(32) NOT NULL DEFAULT '低',
                        source_note VARCHAR(255) NOT NULL DEFAULT '规则推断',
                        recipe_usage_count INT NOT NULL DEFAULT 0,
                        recipe_count INT NOT NULL DEFAULT 0,
                        total_amount_g DECIMAL(18,3) NOT NULL DEFAULT 0,
                        total_amount_ml DECIMAL(18,3) NOT NULL DEFAULT 0,
                        last_seen_recipe_id BIGINT NULL,
                        created_at DATETIME NOT NULL,
                        updated_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_ingredient_source(source_ingredient_id),
                        INDEX idx_ingredient_name(canonical_name),
                        INDEX idx_ingredient_category(category, hazard_class),
                        INDEX idx_ingredient_usage(recipe_usage_count)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                for stmt in [
                    "ALTER TABLE ingredient_thermal_properties MODIFY specific_heat_kj_kg_c DECIMAL(8,3) NULL",
                    "ALTER TABLE ingredient_thermal_properties MODIFY water_fraction DECIMAL(5,3) NULL",
                    "ALTER TABLE ingredient_thermal_properties MODIFY oil_fraction DECIMAL(5,3) NULL",
                ]:
                    try:
                        cur.execute(stmt)
                    except Exception:
                        pass
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS ingredient_thermal_sync_runs (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        sync_type VARCHAR(64) NOT NULL,
                        status VARCHAR(32) NOT NULL,
                        base_rows INT NOT NULL DEFAULT 0,
                        recipe_rows INT NOT NULL DEFAULT 0,
                        ingredient_rows INT NOT NULL DEFAULT 0,
                        error_message TEXT NULL,
                        created_by VARCHAR(64) NULL,
                        started_at DATETIME NOT NULL,
                        finished_at DATETIME NULL,
                        INDEX idx_thermal_sync_time(started_at)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS safety_scan_runs (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        scan_type VARCHAR(64) NOT NULL DEFAULT 'full',
                        status VARCHAR(32) NOT NULL DEFAULT 'RUNNING',
                        total_jobs INT NOT NULL DEFAULT 0,
                        high_temp_jobs INT NOT NULL DEFAULT 0,
                        delay_risk_jobs INT NOT NULL DEFAULT 0,
                        sensor_gap_jobs INT NOT NULL DEFAULT 0,
                        total_alerts INT NOT NULL DEFAULT 0,
                        error_message TEXT NULL,
                        started_at DATETIME NOT NULL,
                        finished_at DATETIME NULL,
                        INDEX idx_safety_scan_time(started_at)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS safety_scan_alerts (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        scan_run_id BIGINT NULL,
                        sn VARCHAR(64) NOT NULL,
                        source_file_id BIGINT NULL,
                        cook_job_id BIGINT NULL,
                        recipe_name VARCHAR(255) NULL,
                        cook_start_time DATETIME NULL,
                        cook_end_time DATETIME NULL,
                        duration_seconds INT NULL,
                        rule_key VARCHAR(64) NOT NULL,
                        rule_label VARCHAR(120) NOT NULL,
                        risk_level VARCHAR(16) NOT NULL DEFAULT 'medium',
                        severity_score INT NOT NULL DEFAULT 0,
                        max_pot_temp DECIMAL(8,2) NULL,
                        avg_pot_temp DECIMAL(8,2) NULL,
                        actual_energy_kwh DECIMAL(10,4) NULL,
                        oil_to_food_interval INT NULL,
                        detail_json TEXT NULL,
                        dismissed TINYINT NOT NULL DEFAULT 0,
                        dismissed_by VARCHAR(64) NULL,
                        dismissed_at DATETIME NULL,
                        created_at DATETIME NOT NULL,
                        INDEX idx_alert_sn(sn, created_at),
                        INDEX idx_alert_rule(rule_key, risk_level),
                        INDEX idx_alert_level(risk_level, dismissed, created_at),
                        INDEX idx_alert_cook(cook_job_id),
                        INDEX idx_alert_scan(scan_run_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS safety_daily_stats (
                        id BIGINT AUTO_INCREMENT PRIMARY KEY,
                        stat_date DATE NOT NULL,
                        sn VARCHAR(64) NOT NULL,
                        total_jobs INT NOT NULL DEFAULT 0,
                        high_temp_300c INT NOT NULL DEFAULT 0,
                        high_temp_330c INT NOT NULL DEFAULT 0,
                        oil_delay_60s INT NOT NULL DEFAULT 0,
                        sensor_gap INT NOT NULL DEFAULT 0,
                        max_temp_reached DECIMAL(8,2) NULL,
                        avg_temp_across_jobs DECIMAL(8,2) NULL,
                        total_energy_kwh DECIMAL(12,4) NULL,
                        created_at DATETIME NOT NULL,
                        UNIQUE KEY uniq_stat_date_sn(stat_date, sn),
                        INDEX idx_stat_date(stat_date),
                        INDEX idx_stat_sn(sn)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
            conn.commit()
            ANALYTICS_DB_READY = True
            return True
        finally:
            conn.close()

def ensure_analytics_db():
    global ANALYTICS_DB_READY
    if ANALYTICS_DB_READY:
        return True
    try:
        return init_analytics_db()
    except Exception as exc:
        ANALYTICS_DB_READY = False
        print(f"analytics db init failed: {exc}")
        return False

def candidate_sns(sn):
    raw = sn.strip()
    candidates = [raw]
    if raw and not raw.startswith('0'):
        candidates.append(f'0{raw}')
    return list(dict.fromkeys(candidates))

def resolve_sn(sn):
    candidates = candidate_sns(sn)
    placeholders = ','.join(['%s'] * len(candidates))
    robot = fetch_one(
        f"SELECT * FROM sop_robot WHERE machinecode IN ({placeholders}) LIMIT 1",
        tuple(candidates),
        source=True,
        database='btyc',
    )
    if robot:
        return robot['machinecode'], robot

    robot = fetch_one(
        "SELECT * FROM sop_robot WHERE machinecode LIKE %s LIMIT 1",
        (f'%{sn.strip()}',),
        source=True,
        database='btyc',
    )
    if robot:
        return robot['machinecode'], robot

    raise HTTPException(status_code=404, detail="Device not found")

def format_gap(delta):
    total_seconds = int(delta.total_seconds())
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    if days:
        return f"{days}天 {hours}小时 {minutes}分"
    if hours:
        return f"{hours}小时 {minutes}分"
    return f"{minutes}分"

def classify_production_behavior(log, duration_stats):
    duration = int(log.get('duration_seconds') or 0)
    recipe_id = log.get('recipe_id')
    stats = duration_stats.get(recipe_id) or {}
    expected = stats.get('avg_duration_seconds')
    expected_duration = float(expected) if expected is not None else None
    duration_ratio = round(duration / expected_duration, 2) if expected_duration else None

    whether = log.get('whether')
    manual = log.get('manual')
    whether_label = {0: '取消', 1: '失败', 2: '成功'}.get(whether, '未知')
    manual_label = {0: '手动', 1: '自动'}.get(manual, '未知')
    tags = []

    if whether == 0:
        tags.append('取消')
    elif whether == 1:
        tags.append('失败')
    if manual == 0:
        tags.append('手动控制')

    if expected_duration and expected_duration >= 60 and duration > 0:
        if duration <= max(expected_duration * 0.5, 60) and expected_duration >= 120:
            tags.append('执行过短')
        elif duration >= expected_duration * 1.8:
            tags.append('执行过长')
    elif duration and duration <= 60:
        tags.append('短时执行')

    if not tags:
        tags.append('正常执行')

    return {
        'whether_label': whether_label,
        'manual_label': manual_label,
        'expected_duration_seconds': round(expected_duration, 1) if expected_duration else None,
        'duration_ratio': duration_ratio,
        'behavior_tags': tags,
        'is_behavior_exception': any(tag in tags for tag in ['取消', '失败', '手动控制', '执行过短', '执行过长', '短时执行']),
    }

def parse_json_array(value):
    if not value:
        return []
    if isinstance(value, list):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, json.JSONDecodeError):
        return []

def parse_json_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return value

def command_text(step):
    return str(step.get('commands') or step.get('command') or '').strip()

def recipe_resource_summary(recipe, detail, usage):
    cook_steps = parse_json_array(detail.get('cook_steps') if detail else None)
    wash_steps = parse_json_array(detail.get('wash_steps') if detail else None)
    moisten_steps = parse_json_array(detail.get('moisten_steps') if detail else None)
    cooking_ingredients = parse_json_array(detail.get('cooking_ingredient') if detail else None)

    resource_flags = {
        'heating': False,
        'stirring': False,
        'pot_position': False,
        'auto_feeding': False,
        'manual_feeding': False,
        'oil': False,
        'lard': False,
        'starch_slurry': False,
        'water': False,
        'seasoning': False,
        'moisten_pot': bool(moisten_steps),
        'wash': bool(wash_steps),
        'move_pot': False,
    }
    resources = set()
    steps = []

    for step in cook_steps:
        text = command_text(step)
        step_type = str(step.get('type', ''))
        automatic = str(step.get('automatic', ''))
        power = str(step.get('power') or '')
        speed = str(step.get('speed') or step.get('stirSpeed') or '')
        position = str(step.get('position') or '')

        if step_type == '3' or (power and power not in {'0', '0.0'}):
            resource_flags['heating'] = True
            resources.add('加热/功率')
        if speed and speed not in {'0', '0.0'}:
            resource_flags['stirring'] = True
            resources.add('搅拌')
        if position:
            resource_flags['pot_position'] = True
            resources.add('锅位')
        if str(step.get('movepot', '0')) not in {'', '0', '0.0'}:
            resource_flags['move_pot'] = True
            resources.add('翻锅/移锅')
        if step_type == '2' or automatic == '1' or '自动投入' in text:
            resource_flags['auto_feeding'] = True
            resources.add('自动投料')
        if step_type == '1' and automatic != '1':
            resource_flags['manual_feeding'] = True
            resources.add('人工/预制投料')
        if any(k in text for k in ['水淀粉', '水溶き片栗粉', '片栗粉', '生粉']):
            resource_flags['starch_slurry'] = True
            resource_flags['water'] = True
            resources.add('水淀粉')
        if any(k in text for k in ['猪油', 'lard', 'Lard']):
            resource_flags['lard'] = True
            resource_flags['oil'] = True
            resources.add('油桶注油')
            resources.add('猪油桶/猪油')
        elif any(k in text for k in ['油', 'オイル', 'oil', 'Oil']):
            resource_flags['oil'] = True
            resources.add('油桶注油')
        if any(k in text for k in ['水', '高汤', 'スープ']):
            resource_flags['water'] = True
            resources.add('水/汤汁')
        if any(k in text for k in ['盐', '精盐', '鸡精', '味精', '生抽', '老抽', '蚝油', '豆瓣', '酱', '粉']):
            resource_flags['seasoning'] = True
            resources.add('调料')

        steps.append({
            'time': step.get('time'),
            'type': step.get('type'),
            'automatic': step.get('automatic'),
            'power': step.get('power'),
            'speed': step.get('speed') or step.get('stirSpeed'),
            'position': step.get('position'),
            'movepot': step.get('movepot'),
            'commands': text,
        })

    if moisten_steps:
        resources.add('润锅')
    if wash_steps:
        resources.add('洗锅')

    category = str(recipe.get('group_name') or recipe.get('type') or '未分类')

    return {
        'id': recipe.get('id'),
        'name': recipe.get('name'),
        'category': category,
        'group_name': recipe.get('group_name'),
        'recipe_type': recipe.get('type'),
        'execution_count': usage.get(recipe.get('id'), {}).get('cnt', 0),
        'total_duration_seconds': usage.get(recipe.get('id'), {}).get('total_duration_seconds', 0),
        'first_time': usage.get(recipe.get('id'), {}).get('first_time'),
        'last_time': usage.get(recipe.get('id'), {}).get('last_time'),
        'cook_time': detail.get('cook_time') if detail else None,
        'resource_flags': resource_flags,
        'resources': sorted(resources),
        'steps': steps[:80],
        'wash_steps_count': len(wash_steps),
        'moisten_steps_count': len(moisten_steps),
        'ingredient_count': len(cooking_ingredients),
        'has_lard': resource_flags['lard'],
    }

def build_recipe_category_summary(recipe_archive):
    categories = {}
    for recipe in recipe_archive:
        category = recipe.get('category') or '未分类'
        if category not in categories:
            categories[category] = {
                'category': category,
                'recipe_count': 0,
                'execution_count': 0,
                'lard_recipe_count': 0,
                'resources': set(),
            }
        item = categories[category]
        item['recipe_count'] += 1
        item['execution_count'] += int(recipe.get('execution_count') or 0)
        if recipe.get('has_lard'):
            item['lard_recipe_count'] += 1
        item['resources'].update(recipe.get('resources') or [])

    summary = []
    for item in categories.values():
        summary.append({
            **item,
            'resources': sorted(item['resources']),
        })
    return sorted(summary, key=lambda row: row['execution_count'], reverse=True)

def get_company_info(company_id):
    if not company_id:
        return {}
    company = fetch_one(
        "SELECT id, company_name, common_name, addr, company_addr, area_code, "
        "geo_pname, geo_cityname, geo_adname, geo_name, geo_address, contact_name, contact_phone "
        "FROM ums_company WHERE id = %s LIMIT 1",
        (company_id,),
        source=True,
        database='btyc',
    )
    return company or {}

def get_device_software_info(sn):
    version_record = fetch_one(
        "SELECT apk_version, rom_version, version, create_time, update_time "
        "FROM robot_version_record WHERE sn = %s ORDER BY update_time DESC LIMIT 1",
        (sn,),
        source=True,
        database='btyc',
    ) or {}
    config_info = fetch_one(
        "SELECT app_version, create_time FROM robot_config_info WHERE sn = %s ORDER BY create_time DESC LIMIT 1",
        (sn,),
        source=True,
        database='btyc',
    ) or {}
    return {
        'upper_computer_version': version_record.get('apk_version') or config_info.get('app_version'),
        'apk_version': version_record.get('apk_version'),
        'rom_version': version_record.get('rom_version'),
        'version_record_time': version_record.get('update_time') or version_record.get('create_time'),
        'reported_app_version': config_info.get('app_version'),
        'reported_app_time': config_info.get('create_time'),
    }

def compact_keyword(value):
    return re.sub(r'[\s\(\)（）【】\[\]{}<>《》,，.。·_\-—/\\\\]+', '', value or '').strip()

def device_lookup_cache_key(keyword, limit):
    return hashlib.sha256(json.dumps({
        'keyword': keyword.strip(),
        'limit': limit,
    }, ensure_ascii=False, sort_keys=True).encode('utf-8')).hexdigest()

def search_devices_by_keyword(keyword, limit=50, force_refresh=False):
    clean = (keyword or '').strip()
    if len(clean) < 2:
        raise HTTPException(status_code=400, detail="请输入至少 2 个字符")
    safe_limit = max(1, min(int(limit or 50), 100))
    cache_key = device_lookup_cache_key(clean, safe_limit)
    now = time.time()
    cached = DEVICE_LOOKUP_CACHE.get(cache_key)
    if cached and not force_refresh and cached['expires_at'] > now:
        result = dict(cached['result'])
        result['cache'] = {'hit': True, 'ttl_seconds': int(cached['expires_at'] - now), 'created_at': cached['created_at']}
        return result

    like = f"%{clean}%"
    compact = compact_keyword(clean)
    compact_like = f"%{compact}%" if compact else like
    compact_sql = """
        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
            CONCAT_WS('', r.machinecode, r.name, r.spec, r.robot_type, r.latest_update_package,
                c.company_name, c.common_name, c.addr, c.company_addr, c.geo_pname,
                c.geo_cityname, c.geo_adname, c.geo_name, c.geo_address, c.contact_name),
            ' ', ''), '（', ''), '）', ''), '(', ''), ')', ''), '-', ''), '_', ''), '·', ''), '，', '')
    """
    rows = fetch_all(
        f"""
        SELECT
            r.machinecode AS sn,
            r.name AS device_name,
            r.spec,
            r.robot_type,
            r.hardware_version,
            r.latest_update_package AS version,
            r.last_online_time,
            r.status,
            r.company_id,
            COALESCE(c.common_name, c.company_name, '') AS customer_name,
            c.company_name,
            c.common_name,
            COALESCE(c.geo_cityname, c.geo_pname, c.area_code, '') AS region,
            c.geo_pname,
            c.geo_cityname,
            c.geo_adname,
            c.area_code,
            c.addr,
            c.company_addr
        FROM btyc.sop_robot r
        LEFT JOIN btyc.ums_company c ON r.company_id = c.id
        WHERE r.machinecode IS NOT NULL AND r.machinecode != ''
          AND (
            r.machinecode LIKE %s OR r.name LIKE %s OR r.spec LIKE %s OR
            r.robot_type LIKE %s OR r.latest_update_package LIKE %s OR
            c.company_name LIKE %s OR c.common_name LIKE %s OR c.addr LIKE %s OR
            c.company_addr LIKE %s OR c.geo_pname LIKE %s OR c.geo_cityname LIKE %s OR
            c.geo_adname LIKE %s OR c.geo_name LIKE %s OR c.geo_address LIKE %s OR
            c.contact_name LIKE %s OR {compact_sql} LIKE %s
          )
        ORDER BY
            CASE
                WHEN r.machinecode = %s THEN 0
                WHEN c.common_name LIKE %s OR c.company_name LIKE %s THEN 1
                WHEN c.geo_cityname LIKE %s OR c.geo_pname LIKE %s OR c.geo_adname LIKE %s THEN 2
                WHEN r.name LIKE %s OR r.spec LIKE %s THEN 3
                ELSE 4
            END,
            r.last_online_time DESC,
            r.machinecode
        LIMIT %s
        """,
        tuple([like] * 15 + [compact_like, clean, like, like, like, like, like, like, like, safe_limit]),
        source=True,
        database='btyc',
    )
    result = {
        'keyword': clean,
        'compact_keyword': compact,
        'total': len(rows),
        'devices': rows,
        'cache': {'hit': False, 'ttl_seconds': DEVICE_LOOKUP_CACHE_TTL_SECONDS, 'created_at': int(now)},
    }
    DEVICE_LOOKUP_CACHE[cache_key] = {
        'result': result,
        'created_at': int(now),
        'expires_at': now + DEVICE_LOOKUP_CACHE_TTL_SECONDS,
    }
    return result

def version_stats_cache_key(version, keyword, limit):
    return hashlib.sha256(json.dumps({
        'version': (version or '').strip(),
        'keyword': (keyword or '').strip(),
        'limit': int(limit or 500),
    }, ensure_ascii=False, sort_keys=True).encode('utf-8')).hexdigest()

def build_device_version_stats(version='', keyword='', limit=500, force_refresh=False):
    clean_version = (version or '').strip()
    clean_keyword = (keyword or '').strip()
    safe_limit = max(1, min(int(limit or 500), 5000))
    cache_key = version_stats_cache_key(clean_version, clean_keyword, safe_limit)
    now = time.time()
    cached = VERSION_STATS_CACHE.get(cache_key)
    if cached and not force_refresh and cached['expires_at'] > now:
        result = dict(cached['result'])
        result['cache'] = {'hit': True, 'ttl_seconds': int(cached['expires_at'] - now), 'created_at': cached['created_at']}
        return result

    version_expr = "COALESCE(NULLIF(rv.apk_version, ''), NULLIF(rci.app_version, ''), '未知上位机版本')"
    version_time_expr = "COALESCE(rv.update_time, rv.create_time, rci.create_time)"
    version_from_sql = """
        FROM btyc.sop_robot r
        LEFT JOIN btyc.ums_company c ON r.company_id = c.id
        LEFT JOIN (
            SELECT sn, apk_version, rom_version, version, create_time, update_time
            FROM (
                SELECT sn, apk_version, rom_version, version, create_time, update_time,
                       ROW_NUMBER() OVER(PARTITION BY sn ORDER BY update_time DESC, id DESC) AS rn
                FROM btyc.robot_version_record
            ) t WHERE rn = 1
        ) rv ON r.machinecode = rv.sn
        LEFT JOIN (
            SELECT sn, app_version, create_time
            FROM (
                SELECT sn, app_version, create_time,
                       ROW_NUMBER() OVER(PARTITION BY sn ORDER BY create_time DESC, id DESC) AS rn
                FROM btyc.robot_config_info
            ) t WHERE rn = 1
        ) rci ON r.machinecode = rci.sn
    """
    conditions = ["r.machinecode IS NOT NULL", "r.machinecode != ''"]
    args = []
    if clean_version:
        if clean_version == '未知上位机版本':
            conditions.append("(rv.apk_version IS NULL OR rv.apk_version = '') AND (rci.app_version IS NULL OR rci.app_version = '')")
        else:
            conditions.append(f"{version_expr} = %s")
            args.append(clean_version)
    if clean_keyword:
        like = f"%{clean_keyword}%"
        conditions.append(
            "(r.machinecode LIKE %s OR r.name LIKE %s OR r.spec LIKE %s OR r.robot_type LIKE %s OR "
            "r.latest_update_package LIKE %s OR r.hardware_version LIKE %s OR "
            "rv.apk_version LIKE %s OR rv.rom_version LIKE %s OR rci.app_version LIKE %s OR "
            "c.company_name LIKE %s OR c.common_name LIKE %s OR c.geo_pname LIKE %s OR "
            "c.geo_cityname LIKE %s OR c.geo_adname LIKE %s OR c.area_code LIKE %s)"
        )
        args.extend([like] * 15)
    where_sql = " AND ".join(conditions)

    summary_rows = fetch_all(
        f"""
        SELECT
            {version_expr} AS version,
            COUNT(*) AS device_count,
            COUNT(DISTINCT r.company_id) AS customer_count,
            MIN({version_time_expr}) AS first_upgrade_time,
            MAX({version_time_expr}) AS last_upgrade_time,
            MAX(r.last_online_time) AS latest_online_time
        {version_from_sql}
        WHERE {where_sql}
        GROUP BY {version_expr}
        ORDER BY device_count DESC, last_upgrade_time DESC
        LIMIT 200
        """,
        tuple(args),
        source=True,
        database='btyc',
    )
    device_rows = fetch_all(
        f"""
        SELECT
            r.machinecode AS sn,
            r.name AS device_name,
            r.spec,
            r.robot_type,
            r.hardware_version,
            {version_expr} AS version,
            rv.apk_version,
            rv.rom_version,
            rci.app_version AS reported_app_version,
            r.latest_update_package AS update_package,
            {version_time_expr} AS version_update_time,
            r.update_time AS package_update_time,
            r.last_online_time,
            r.status,
            r.network_status,
            r.company_id,
            COALESCE(c.common_name, c.company_name, '') AS customer_name,
            COALESCE(c.geo_cityname, c.geo_pname, c.area_code, '') AS region
        {version_from_sql}
        WHERE {where_sql}
        ORDER BY {version_time_expr} DESC, r.last_online_time DESC, r.machinecode
        LIMIT %s
        """,
        tuple(args + [safe_limit]),
        source=True,
        database='btyc',
    )
    total_row = fetch_one(
        f"""
        SELECT COUNT(*) AS total_devices, COUNT(DISTINCT r.company_id) AS total_customers
        {version_from_sql}
        WHERE {where_sql}
        """,
        tuple(args),
        source=True,
        database='btyc',
    )
    result = {
        'params': {'version': clean_version, 'keyword': clean_keyword, 'limit': safe_limit},
        'summary': {
            'total_devices': int(total_row.get('total_devices') or 0),
            'total_customers': int(total_row.get('total_customers') or 0),
            'version_count': len(summary_rows),
            'returned_devices': len(device_rows),
            'truncated': int(total_row.get('total_devices') or 0) > len(device_rows),
            'upgrade_time_field': 'robot_version_record.update_time / robot_config_info.create_time',
            'version_field': 'robot_version_record.apk_version',
        },
        'versions': summary_rows,
        'devices': device_rows,
        'cache': {'hit': False, 'ttl_seconds': VERSION_STATS_CACHE_TTL_SECONDS, 'created_at': int(now)},
    }
    VERSION_STATS_CACHE[cache_key] = {
        'result': result,
        'created_at': int(now),
        'expires_at': now + VERSION_STATS_CACHE_TTL_SECONDS,
    }
    return result

def get_device_log_files(sn):
    rows = fetch_all(
        "SELECT id, sn, file_length, file_name, pic AS url, type, create_time, update_time, cos_deleted "
        "FROM machine_ftp WHERE sn = %s ORDER BY create_time DESC LIMIT 300",
        (sn,),
        source=True,
        database='btyc',
    )
    for row in rows:
        row['file_size_label'] = format_bytes(row.get('file_length'))
        row['downloadable'] = bool(row.get('url')) and not bool(row.get('cos_deleted'))
        row['log_time_hint'] = infer_log_time_from_filename(row.get('file_name'))
    upsert_log_package_index(sn, rows)
    attach_log_package_statuses(sn, rows)
    return rows

def parse_mb_value(value):
    try:
        n = float(str(value or '').strip() or 0)
        return round(n, 3) if n >= 0 else None
    except Exception:
        return None

def to_mysql_dt(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19].replace('T', ' '), "%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    return None

def analytics_ui_status(download_status, parse_status, storage_status):
    if storage_status == 'stored':
        return '已匹配生产记录'
    if parse_status == 'parsed':
        return '已解析'
    if parse_status == 'parsing':
        return '解析中'
    if parse_status == 'queued':
        return '等待后台解析'
    if parse_status == 'no_production_match':
        return 'DB 未匹配，待检查日志证据'
    if parse_status in ('parse_failed', 'no_temperature_data'):
        return '解析失败'
    if download_status == 'downloaded':
        return '已下载'
    if download_status in ('download_failed', 'remote_deleted'):
        return '不可下载'
    return '可下载'

def upsert_log_package_index(sn, rows):
    if not rows or not ensure_analytics_db():
        return
    now = datetime.now().replace(microsecond=0)
    payload = []
    for row in rows:
        file_id = row.get('id')
        if not file_id:
            continue
        cos_deleted = 1 if row.get('cos_deleted') else 0
        url_hash = hashlib.sha256(str(row.get('url') or '').encode('utf-8')).hexdigest() if row.get('url') else None
        download_status = 'remote_deleted' if cos_deleted else ('remote_available' if row.get('url') else 'download_failed')
        payload.append((
            sn, int(file_id), row.get('file_name'), parse_mb_value(row.get('file_length')), row.get('file_size_label'),
            url_hash, to_mysql_dt(row.get('log_time_hint')), to_mysql_dt(row.get('create_time')),
            to_mysql_dt(row.get('update_time')), cos_deleted, download_status,
            analytics_ui_status(download_status, 'not_started', 'not_stored'), now, now,
        ))
    executemany_local(
        """
        INSERT INTO device_log_packages(
            sn, source_file_id, file_name, file_size_mb, file_size_label, remote_url_hash,
            log_time_hint, remote_create_time, remote_update_time, cos_deleted,
            download_status, ui_status, created_at, updated_at
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            file_name = VALUES(file_name),
            file_size_mb = VALUES(file_size_mb),
            file_size_label = VALUES(file_size_label),
            remote_url_hash = VALUES(remote_url_hash),
            log_time_hint = VALUES(log_time_hint),
            remote_create_time = VALUES(remote_create_time),
            remote_update_time = VALUES(remote_update_time),
            cos_deleted = VALUES(cos_deleted),
            download_status = CASE
                WHEN storage_status = 'stored' THEN download_status
                WHEN parse_status IN ('parsed', 'parsing', 'queued') THEN download_status
                ELSE VALUES(download_status)
            END,
            ui_status = CASE
                WHEN storage_status = 'stored' THEN '已匹配生产记录'
                WHEN parse_status = 'parsed' THEN '已解析'
                WHEN parse_status = 'parsing' THEN '解析中'
                WHEN parse_status = 'queued' THEN '等待后台解析'
                WHEN parse_status = 'no_production_match' THEN 'DB 未匹配，待检查日志证据'
                WHEN parse_status IN ('parse_failed', 'no_temperature_data') THEN '解析失败'
                WHEN VALUES(download_status) = 'remote_deleted' THEN '不可下载'
                ELSE VALUES(ui_status)
            END,
            updated_at = VALUES(updated_at)
        """,
        payload,
    )

def attach_log_package_statuses(sn, rows):
    if not rows or not ensure_analytics_db():
        for row in rows:
            row['analytics_status'] = '可下载' if row.get('downloadable') else '不可下载'
        return
    ids = [int(row['id']) for row in rows if row.get('id')]
    if not ids:
        return
    placeholders = ','.join(['%s'] * len(ids))
    status_rows = fetch_all(
        f"""
        SELECT source_file_id, download_status, parse_status, storage_status, ui_status,
               parse_version, log_start_time, log_end_time, cook_count, sample_count,
               error_message, parse_attempts, last_attempt_at, parsed_at, stored_at
        FROM device_log_packages
        WHERE sn = %s AND source_file_id IN ({placeholders})
        """,
        tuple([sn] + ids),
    )
    status_map = {int(row['source_file_id']): row for row in status_rows}
    for row in rows:
        status = status_map.get(int(row.get('id') or 0), {})
        row['download_status'] = status.get('download_status') or ('remote_available' if row.get('downloadable') else 'remote_deleted')
        row['parse_status'] = status.get('parse_status') or 'not_started'
        row['storage_status'] = status.get('storage_status') or 'not_stored'
        row['analytics_status'] = analytics_ui_status(
            row['download_status'],
            row['parse_status'],
            row['storage_status'],
        )
        row['structured_cook_count'] = int(status.get('cook_count') or 0)
        row['structured_sample_count'] = int(status.get('sample_count') or 0)
        row['structured_error'] = status.get('error_message')
        row['structured_coverage_start'] = status.get('log_start_time')
        row['structured_coverage_end'] = status.get('log_end_time')
        row['structured_parsed_at'] = status.get('parsed_at')
        row['structured_stored_at'] = status.get('stored_at')
        diagnostics = read_cached_log_package_diagnostics(row.get('id')) if row.get('id') else None
        if diagnostics:
            row['diagnostics_status'] = diagnostics.get('diagnosis_message')
            row['diagnostics_code'] = diagnostics.get('diagnosis_code')
            row['internal_session_count'] = len(diagnostics.get('internal_sessions') or [])
            row['diagnostics_log_time_start'] = diagnostics.get('log_time_start')
            row['diagnostics_log_time_end'] = diagnostics.get('log_time_end')
            if status.get('parse_status') == 'no_production_match':
                row['analytics_status'] = diagnostics.get('diagnosis_message') or row['analytics_status']

def format_bytes(value):
    raw_text = str(value or '').strip()
    try:
        size = float(raw_text or 0)
    except (TypeError, ValueError):
        return '-'
    if size <= 0:
        return '-'
    # machine_ftp.file_length is stored as an MB-like decimal string in this source DB
    # for current log packages, e.g. "3.67" means roughly 3.67 MB.
    if '.' in raw_text and size < 1024:
        return f"{size:.1f}MB"
    units = ['B', 'KB', 'MB', 'GB']
    unit = 0
    while size >= 1024 and unit < len(units) - 1:
        size /= 1024
        unit += 1
    if unit == 0:
        return f"{int(size)}{units[unit]}"
    return f"{size:.1f}{units[unit]}"

def infer_log_time_from_filename(file_name):
    text = str(file_name or '')
    match = re.search(r'log_(\d{4})_(\d{2})_(\d{2})-(\d{2})_(\d{2})_(\d{2})', text)
    if not match:
        return ''
    y, mo, d, h, mi, s = match.groups()
    return f"{y}-{mo}-{d} {h}:{mi}:{s}"

def parse_log_ts(line):
    patterns = [
        (re.compile(r'\[(\d{4})-(\d{2})-(\d{2})[ _](\d{2}:\d{2}:\d{2})\]'), '%Y-%m-%d %H:%M:%S'),
        (re.compile(r'\[(\d{2})\.(\d{2})\.(\d{2}) (\d{2}:\d{2}:\d{2})\]'), None),
        (re.compile(r'\[(\d{2})-(\d{2}) (\d{2}:\d{2}:\d{2})\]'), None),
    ]
    for idx, (pattern, fmt) in enumerate(patterns):
        match = pattern.search(line)
        if not match:
            continue
        try:
            if idx == 0:
                return datetime.strptime(f"{match.group(1)}-{match.group(2)}-{match.group(3)} {match.group(4)}", fmt)
            if idx == 1:
                return datetime.strptime(f"20{match.group(1)}-{match.group(2)}-{match.group(3)} {match.group(4)}", "%Y-%m-%d %H:%M:%S")
            return datetime.strptime(f"2026-{match.group(1)}-{match.group(2)} {match.group(3)}", "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None

def log_analysis_cache_path(file_id):
    return CACHE_DIR / 'log_analysis' / f'{int(file_id)}.json'

def read_cached_log_analysis(file_id):
    path = log_analysis_cache_path(file_id)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text())
        if payload.get('analysis_version') != LOG_ANALYSIS_VERSION:
            return None
        return payload
    except Exception:
        path.unlink(missing_ok=True)
        return None

def save_cached_log_analysis(file_id, payload):
    path = log_analysis_cache_path(file_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, default=str))

COOK_TEMPERATURE_CACHE_VERSION = 9
COOK_TEMPERATURE_CACHE_TTL_SECONDS = int(os.getenv('COOK_TEMPERATURE_CACHE_TTL_SECONDS', str(24 * 3600)))

def cook_temperature_cache_path(sn, file_id):
    digest = hashlib.md5(f"{sn}:{int(file_id)}:v{COOK_TEMPERATURE_CACHE_VERSION}".encode('utf-8')).hexdigest()
    return CACHE_DIR / 'cook_temperature' / f'{digest}.json'

def read_cached_cook_temperature(sn, file_id):
    path = cook_temperature_cache_path(sn, file_id)
    if not path.exists():
        return None
    try:
        wrapper = json.loads(path.read_text())
        if wrapper.get('analysis_version') != COOK_TEMPERATURE_CACHE_VERSION:
            return None
        created_at = float(wrapper.get('created_at') or 0)
        now = time.time()
        if created_at <= 0 or now - created_at > COOK_TEMPERATURE_CACHE_TTL_SECONDS:
            return None
        payload = wrapper.get('payload') or {}
        payload['cache'] = {
            'hit': True,
            'disk': True,
            'created_at': created_at,
            'ttl_seconds': int(COOK_TEMPERATURE_CACHE_TTL_SECONDS - (now - created_at)),
            'cache_key': path.stem,
        }
        return payload
    except Exception:
        path.unlink(missing_ok=True)
        return None

def save_cached_cook_temperature(sn, file_id, payload, created_at=None):
    path = cook_temperature_cache_path(sn, file_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    serializable = dict(payload)
    serializable.pop('cache', None)
    created_at = created_at or time.time()
    with path.open('w', encoding='utf-8') as fp:
        json.dump({
            'analysis_version': COOK_TEMPERATURE_CACHE_VERSION,
            'created_at': created_at,
            'ttl_seconds': COOK_TEMPERATURE_CACHE_TTL_SECONDS,
            'payload': serializable,
        }, fp, ensure_ascii=False, default=str)
    return path.stem

def mark_log_package_status(sn, file_id, download_status=None, parse_status=None, storage_status=None, error_message=None, result=None):
    if not ensure_analytics_db():
        return
    now = datetime.now().replace(microsecond=0)
    if result:
        coverage = result.get('coverage') or {}
        download_status = download_status or 'downloaded'
        parse_status = parse_status or 'parsed'
        storage_status = storage_status or 'stored'
        args = (
            download_status,
            parse_status,
            storage_status,
            analytics_ui_status(download_status, parse_status, storage_status),
            COOK_TEMPERATURE_CACHE_VERSION,
            to_mysql_dt(coverage.get('start')),
            to_mysql_dt(coverage.get('end')),
            int(result.get('cook_count') or 0),
            int(coverage.get('sample_count') or 0),
            error_message,
            now if parse_status in ('parsed', 'parse_failed', 'no_production_match', 'no_temperature_data') else None,
            now if storage_status == 'stored' else None,
            now,
            sn,
            int(file_id),
        )
        execute_local(
            """
            UPDATE device_log_packages
            SET download_status=%s, parse_status=%s, storage_status=%s, ui_status=%s,
                parse_version=%s, log_start_time=%s, log_end_time=%s, cook_count=%s,
                sample_count=%s, error_message=%s,
                parsed_at=COALESCE(%s, parsed_at), stored_at=COALESCE(%s, stored_at),
                last_attempt_at=%s, parse_attempts=parse_attempts+1, updated_at=%s
            WHERE sn=%s AND source_file_id=%s
            """,
            args[:-2] + (now,) + args[-2:],
        )
        return
    existing = fetch_one(
        "SELECT download_status, parse_status, storage_status FROM device_log_packages WHERE sn=%s AND source_file_id=%s",
        (sn, int(file_id)),
    ) if ensure_analytics_db() else None
    next_download = download_status or (existing or {}).get('download_status') or 'remote_available'
    next_parse = parse_status or (existing or {}).get('parse_status') or 'not_started'
    next_storage = storage_status or (existing or {}).get('storage_status') or 'not_stored'
    execute_local(
        """
        UPDATE device_log_packages
        SET download_status=%s, parse_status=%s, storage_status=%s, ui_status=%s,
            error_message=%s, last_attempt_at=%s, parse_attempts=parse_attempts+1,
            updated_at=%s
        WHERE sn=%s AND source_file_id=%s
        """,
        (
            next_download, next_parse, next_storage,
            analytics_ui_status(next_download, next_parse, next_storage),
            error_message, now, now, sn, int(file_id),
        ),
    )

def read_cook_temperature_from_db(sn, file_id):
    if not ensure_analytics_db():
        return None
    package = fetch_one(
        """
        SELECT source_file_id, file_name, file_size_mb, remote_create_time, log_start_time, log_end_time,
               cook_count, sample_count, parse_version, parsed_at, stored_at
        FROM device_log_packages
        WHERE sn=%s AND source_file_id=%s AND storage_status='stored'
        """,
        (sn, int(file_id)),
    )
    if not package or int(package.get('parse_version') or 0) != COOK_TEMPERATURE_CACHE_VERSION:
        return None
    job_rows = fetch_all(
        """
        SELECT payload_json FROM cook_jobs
        WHERE sn=%s AND source_file_id=%s
        ORDER BY cook_start_time ASC, id ASC
        """,
        (sn, int(file_id)),
    )
    cooks = []
    for row in job_rows:
        try:
            cooks.append(json.loads(row.get('payload_json') or '{}'))
        except Exception:
            continue
    if not cooks:
        return None
    selected = cooks[-1]
    created_at = time.mktime((package.get('stored_at') or datetime.now()).timetuple()) if package.get('stored_at') else time.time()
    return {
        'sn': sn,
        'file': {
            'id': int(file_id),
            'file_name': package.get('file_name'),
            'create_time': package.get('remote_create_time'),
            'file_length': package.get('file_size_mb'),
        },
        'coverage': {
            'start': package.get('log_start_time'),
            'end': package.get('log_end_time'),
            'sample_count': int(package.get('sample_count') or 0),
            'android_sample_count': None,
            'temperature_log_sample_count': None,
            'android_file_count': None,
            'temperature_file_count': None,
            'android_files': [],
            'temperature_files': [],
            'temperature_unit': '来自本地结构化库：android 日志功率=指令/实际输出功率，温度=机芯/线盘/输出温度；temperature*.log=滤波/红外/输出温度。',
            'newer_production_not_covered': False,
            'latest_production_time': None,
            'source_summary': [],
        },
        'cook_count': len(cooks),
        'cooks': cooks,
        'cook': selected.get('cook') or {},
        'summary': selected.get('summary') or {},
        'steps': selected.get('steps') or [],
        'main_board_actions': selected.get('main_board_actions') or [],
        'series': selected.get('series') or [],
        'cache': {
            'hit': True,
            'database': True,
            'created_at': created_at,
            'ttl_seconds': COOK_TEMPERATURE_CACHE_TTL_SECONDS,
            'cache_key': f'db:{sn}:{file_id}',
        },
    }

def persist_cook_temperature_result(result):
    if not result or not ensure_analytics_db():
        return False
    sn = result.get('sn')
    file_id = int((result.get('file') or {}).get('id') or 0)
    if not sn or not file_id:
        return False
    now = datetime.now().replace(microsecond=0)
    existing = fetch_all("SELECT id FROM cook_jobs WHERE sn=%s AND source_file_id=%s", (sn, file_id))
    existing_ids = [int(row['id']) for row in existing]
    if existing_ids:
        placeholders = ','.join(['%s'] * len(existing_ids))
        execute_local(f"DELETE FROM cook_temperature_samples WHERE cook_job_id IN ({placeholders})", tuple(existing_ids))
        execute_local(f"DELETE FROM cook_power_events WHERE cook_job_id IN ({placeholders})", tuple(existing_ids))
        execute_local(f"DELETE FROM cook_action_events WHERE cook_job_id IN ({placeholders})", tuple(existing_ids))
        execute_local(f"DELETE FROM cook_jobs WHERE id IN ({placeholders})", tuple(existing_ids))

    for item in result.get('cooks') or []:
        cook = item.get('cook') or {}
        summary = item.get('summary') or {}
        execute_local(
            """
            INSERT INTO cook_jobs(
                sn, source_file_id, machine_log_id, recipe_id, recipe_name, cook_start_time, cook_end_time,
                duration_seconds, max_pot_temp, avg_pot_temp, sample_count, step_count,
                android_action_count, payload_json, parse_version, created_at, updated_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                sn, file_id, cook.get('id'), cook.get('recipe_id'), cook.get('recipe_name'),
                to_mysql_dt(cook.get('start_time')), to_mysql_dt(cook.get('end_time_calc') or cook.get('create_time')),
                int(cook.get('duration_seconds') or 0), summary.get('max_temp'), summary.get('avg_temp'),
                int(summary.get('sample_count') or 0), len(item.get('steps') or []),
                len(item.get('android_actions') or []), json.dumps(item, ensure_ascii=False, default=str),
                COOK_TEMPERATURE_CACHE_VERSION, now, now,
            ),
        )
        job = fetch_one(
            "SELECT id FROM cook_jobs WHERE sn=%s AND source_file_id=%s AND machine_log_id <=> %s ORDER BY id DESC LIMIT 1",
            (sn, file_id, cook.get('id')),
        )
        if not job:
            continue
        job_id = int(job['id'])
        samples = []
        peak_temp = summary.get('max_temp')
        for sample in item.get('temperature_samples') or []:
            pot = sample.get('pot_temp')
            samples.append((
                job_id, sn, file_id, to_mysql_dt(sample.get('time')), sample.get('offset_seconds'), sample.get('source'),
                pot, sample.get('filtered_temp'), sample.get('infrared_temp'), sample.get('output_temp'),
                sample.get('core_temp'), sample.get('coil_temp'), sample.get('raw'), 1 if peak_temp is not None and pot == peak_temp else 0, now,
            ))
        executemany_local(
            """
            INSERT INTO cook_temperature_samples(
                cook_job_id, sn, source_file_id, sample_time, offset_seconds, source, pot_temp,
                filtered_temp, infrared_temp, output_temp, core_temp, coil_temp, raw_value, is_peak, created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            samples,
        )
        actions = []
        powers = []
        for action in item.get('android_actions') or []:
            temp = action.get('temperature') or {}
            power = action.get('power') or {}
            actions.append((
                job_id, sn, file_id, to_mysql_dt(action.get('time')), action.get('offset_seconds'),
                action.get('kind'), action.get('label'), action.get('raw'),
                power.get('command_power_kw'), power.get('actual_power_kw'),
                temp.get('core_temp'), temp.get('coil_temp'), temp.get('output_temp'), now,
            ))
            if power:
                powers.append((
                    job_id, sn, file_id, to_mysql_dt(action.get('time')), action.get('offset_seconds'),
                    power.get('command_power_kw'), power.get('actual_power_kw'),
                    power.get('command_power_w'), power.get('actual_power_w'), power.get('raw'), now,
                ))
        executemany_local(
            """
            INSERT INTO cook_action_events(
                cook_job_id, sn, source_file_id, event_time, offset_seconds, event_type, event_label,
                raw_log_excerpt, command_power_kw, actual_power_kw, core_temp, coil_temp, output_temp, created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            actions,
        )
        executemany_local(
            """
            INSERT INTO cook_power_events(
                cook_job_id, sn, source_file_id, event_time, offset_seconds, command_power_kw,
                actual_power_kw, command_power_w, actual_power_w, raw_value, created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            powers,
        )
    mark_log_package_status(sn, file_id, result=result)
    return True

def mark_device_watched(sn, username='system', priority=50):
    if not sn or not ensure_analytics_db():
        return
    now = datetime.now().replace(microsecond=0)
    execute_local(
        """
        INSERT INTO watched_devices(sn, status, priority, first_seen_at, last_seen_at, created_by, updated_at)
        VALUES (%s, 'active', %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            status = CASE WHEN status = 'paused' THEN status ELSE 'active' END,
            priority = LEAST(priority, VALUES(priority)),
            last_seen_at = VALUES(last_seen_at),
            created_by = COALESCE(created_by, VALUES(created_by)),
            updated_at = VALUES(updated_at)
        """,
        (sn, int(priority or 50), now, now, username, now),
    )

def device_structured_summary(sn):
    if not ensure_analytics_db():
        return {'enabled': False}
    package = fetch_one(
        """
        SELECT
            COUNT(*) AS total_packages,
            SUM(storage_status='stored') AS stored_packages,
            SUM(parse_status IN ('queued','parsing')) AS running_packages,
            SUM(parse_status IN ('parse_failed','no_temperature_data','no_production_match')) AS failed_packages,
            MIN(log_start_time) AS coverage_start,
            MAX(log_end_time) AS coverage_end,
            MAX(parsed_at) AS last_parsed_at,
            MAX(stored_at) AS last_stored_at,
            MAX(last_attempt_at) AS last_attempt_at
        FROM device_log_packages
        WHERE sn=%s
        """,
        (sn,),
    ) or {}
    watched = fetch_one("SELECT status, priority, last_seen_at, last_sync_at, last_parse_at, last_error FROM watched_devices WHERE sn=%s", (sn,))
    recent_errors = fetch_all(
        """
        SELECT source_file_id, file_name, ui_status, error_message, last_attempt_at
        FROM device_log_packages
        WHERE sn=%s AND parse_status IN ('parse_failed','no_temperature_data','no_production_match')
        ORDER BY last_attempt_at DESC
        LIMIT 6
        """,
        (sn,),
    )
    total = int(package.get('total_packages') or 0)
    stored = int(package.get('stored_packages') or 0)
    return {
        'enabled': True,
        'watched': bool(watched),
        'watch_status': (watched or {}).get('status'),
        'total_packages': total,
        'stored_packages': stored,
        'running_packages': int(package.get('running_packages') or 0),
        'failed_packages': int(package.get('failed_packages') or 0),
        'coverage_start': package.get('coverage_start'),
        'coverage_end': package.get('coverage_end'),
        'last_parsed_at': package.get('last_parsed_at'),
        'last_stored_at': package.get('last_stored_at'),
        'last_attempt_at': package.get('last_attempt_at'),
        'completion_rate': round(stored * 100 / total, 1) if total else 0,
        'recent_errors': recent_errors,
    }

def structured_cook_temperature_by_day(sn, day, limit=300):
    if not ensure_analytics_db():
        raise HTTPException(status_code=503, detail="本地结构化库尚未就绪")
    try:
        start = datetime.strptime(day, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="日期格式应为 YYYY-MM-DD")
    end = start + timedelta(days=1)
    safe_limit = max(1, min(int(limit or 300), 1000))
    rows = fetch_all(
        """
        SELECT id, source_file_id, machine_log_id, recipe_id, recipe_name, cook_start_time,
               cook_end_time, duration_seconds, max_pot_temp, avg_pot_temp, sample_count,
               step_count, android_action_count
        FROM cook_jobs
        WHERE sn=%s AND cook_start_time >= %s AND cook_start_time < %s
        ORDER BY cook_start_time ASC
        LIMIT %s
        """,
        (sn, start, end, safe_limit),
    )
    include_details = len(rows) <= STRUCTURED_DAY_DETAIL_LIMIT
    payload_map = {}
    if include_details and rows:
        ids = [int(row['id']) for row in rows if row.get('id')]
        placeholders = ','.join(['%s'] * len(ids))
        payload_rows = fetch_all(
            f"SELECT id, payload_json FROM cook_jobs WHERE id IN ({placeholders})",
            tuple(ids),
        )
        payload_map = {int(row['id']): row.get('payload_json') for row in payload_rows}
    cooks = []
    for row in rows:
        payload = None
        if include_details:
            try:
                payload = json.loads(payload_map.get(int(row.get('id') or 0)) or '{}')
            except Exception:
                payload = None
        if payload:
            cooks.append(payload)
        else:
            cooks.append({
                'cook': {
                    'id': row.get('machine_log_id'),
                    'recipe_id': row.get('recipe_id'),
                    'recipe_name': row.get('recipe_name'),
                    'start_time': row.get('cook_start_time'),
                    'end_time_calc': row.get('cook_end_time'),
                    'duration_seconds': row.get('duration_seconds'),
                },
                'summary': {
                    'max_temp': row.get('max_pot_temp'),
                    'avg_temp': row.get('avg_pot_temp'),
                    'sample_count': row.get('sample_count'),
                },
                'steps': [],
                'android_actions': [],
                'main_board_actions': [],
                'temperature_samples': [],
                'detail_loaded': False,
            })
    temp_stats = fetch_one(
        """
        SELECT COUNT(*) AS sample_count, MAX(pot_temp) AS max_temp, AVG(pot_temp) AS avg_temp
        FROM cook_temperature_samples
        WHERE sn=%s AND sample_time >= %s AND sample_time < %s
        """,
        (sn, start, end),
    ) or {}
    package_rows = fetch_all(
        """
        SELECT source_file_id, file_name, ui_status, parse_status, storage_status, log_start_time, log_end_time, error_message
        FROM device_log_packages
        WHERE sn=%s AND (
            (log_start_time >= %s AND log_start_time < %s) OR
            (log_end_time >= %s AND log_end_time < %s)
        )
        ORDER BY COALESCE(log_start_time, remote_create_time) ASC
        """,
        (sn, start, end, start, end),
    )
    primary_file = package_rows[-1] if package_rows else {}
    selected = cooks[-1] if cooks else {}
    return {
        'sn': sn,
        'day': day,
        'source': 'structured_db',
        'file': {
            'id': primary_file.get('source_file_id'),
            'file_name': primary_file.get('file_name') or f'{day} 结构化作业库',
            'ui_status': primary_file.get('ui_status'),
        },
        'cook_count': len(cooks),
        'cooks': cooks,
        'cook': selected.get('cook') or {},
        'summary': selected.get('summary') or {},
        'steps': selected.get('steps') or [],
        'main_board_actions': selected.get('main_board_actions') or [],
        'coverage': {
            'start': start,
            'end': end - timedelta(seconds=1),
            'sample_count': int(temp_stats.get('sample_count') or 0),
            'temperature_unit': '来自本地结构化库，按日期聚合读取。',
            'newer_production_not_covered': False,
            'source_summary': [],
            'packages': package_rows,
        },
        'cache': {
            'hit': True,
            'database': True,
            'created_at': time.time(),
            'ttl_seconds': COOK_TEMPERATURE_CACHE_TTL_SECONDS,
            'cache_key': f'day:{sn}:{day}',
        },
        'summary_day': {
            'max_temp': temp_stats.get('max_temp'),
            'avg_temp': round(float(temp_stats.get('avg_temp') or 0), 1) if temp_stats.get('avg_temp') is not None else None,
            'sample_count': int(temp_stats.get('sample_count') or 0),
            'package_count': len(package_rows),
            'detail_included': include_details,
            'detail_limit': STRUCTURED_DAY_DETAIL_LIMIT,
        },
    }

def queue_recent_log_packages_for_device(sn, limit=3):
    if not ensure_analytics_db():
        return 0
    recover_stale_parsing_packages(sn)
    rows = fetch_all(
        """
        SELECT source_file_id
        FROM device_log_packages
        WHERE sn=%s
          AND download_status='remote_available'
          AND storage_status != 'stored'
          AND parse_status IN ('not_started','parse_failed','no_temperature_data')
          AND file_size_mb > 0
          AND file_size_mb <= 50
        ORDER BY COALESCE(remote_create_time, log_time_hint) DESC
        LIMIT %s
        """,
        (sn, max(1, min(int(limit or 3), 10))),
    )
    now = datetime.now().replace(microsecond=0)
    for row in rows:
        execute_local(
            """
            UPDATE device_log_packages
            SET parse_status='queued', ui_status='排队中', updated_at=%s
            WHERE sn=%s AND source_file_id=%s AND storage_status!='stored'
            """,
            (now, sn, int(row['source_file_id'])),
        )
    return len(rows)

def parse_date_range(start_date, end_date, max_days=370):
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    except Exception:
        raise HTTPException(status_code=400, detail="日期格式应为 YYYY-MM-DD")
    if end <= start:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    if (end - start).days > max_days:
        raise HTTPException(status_code=400, detail=f"单次查询最多支持 {max_days} 天")
    return start, end

def get_device_log_files_in_range(sn, start, end, limit=1000):
    rows = fetch_all(
        "SELECT id, sn, file_length, file_name, pic AS url, type, create_time, update_time, cos_deleted "
        "FROM machine_ftp WHERE sn=%s ORDER BY create_time DESC LIMIT %s",
        (sn, max(1, min(int(limit or 1000), 2000))),
        source=True,
        database='btyc',
    )
    selected = []
    for row in rows:
        row['file_size_label'] = format_bytes(row.get('file_length'))
        row['downloadable'] = bool(row.get('url')) and not bool(row.get('cos_deleted'))
        row['log_time_hint'] = infer_log_time_from_filename(row.get('file_name'))
        package_time = to_mysql_dt(row.get('log_time_hint')) or to_mysql_dt(row.get('create_time'))
        if package_time and start <= package_time < end:
            selected.append(row)
    upsert_log_package_index(sn, selected)
    attach_log_package_statuses(sn, selected)
    return selected

def queue_temperature_calibration_scans(sn, rows, retry_failed=False):
    if not rows or not ensure_analytics_db():
        return 0
    now = datetime.now().replace(microsecond=0)
    queued = 0
    for row in rows:
        if not row.get('downloadable') or parse_mb_value(row.get('file_length')) in (None, 0):
            continue
        file_id = int(row['id'])
        existing = fetch_one(
            "SELECT scan_status, scan_version, attempts FROM device_log_event_scans "
            "WHERE sn=%s AND source_file_id=%s AND scan_type='temperature_calibration'",
            (sn, file_id),
        )
        if existing:
            status = existing.get('scan_status')
            current_version = int(existing.get('scan_version') or 0)
            attempts = int(existing.get('attempts') or 0)
            if status == 'completed' and current_version == LOG_EVENT_SCAN_VERSION:
                continue
            if status in ('queued', 'scanning'):
                continue
            if status == 'failed' and (not retry_failed or attempts >= 4):
                continue
            execute_local(
                """
                UPDATE device_log_event_scans
                SET scan_status='queued', error_message=NULL, queued_at=%s, updated_at=%s
                WHERE sn=%s AND source_file_id=%s AND scan_type='temperature_calibration'
                """,
                (now, now, sn, file_id),
            )
        else:
            execute_local(
                """
                INSERT INTO device_log_event_scans(
                    sn, source_file_id, scan_type, scan_status, scan_version,
                    event_count, queued_at, updated_at
                ) VALUES (%s,%s,'temperature_calibration','queued',0,0,%s,%s)
                """,
                (sn, file_id, now, now),
            )
        queued += 1
    return queued

def temperature_calibration_scan_worker(sn):
    try:
        while True:
            row = fetch_one(
                """
                SELECT source_file_id
                FROM device_log_event_scans
                WHERE sn=%s AND scan_type='temperature_calibration' AND scan_status='queued'
                ORDER BY queued_at ASC, source_file_id ASC
                LIMIT 1
                """,
                (sn,),
            )
            if not row:
                break
            file_id = int(row['source_file_id'])
            now = datetime.now().replace(microsecond=0)
            execute_local(
                """
                UPDATE device_log_event_scans
                SET scan_status='scanning', started_at=%s, attempts=attempts+1,
                    error_message=NULL, updated_at=%s
                WHERE sn=%s AND source_file_id=%s AND scan_type='temperature_calibration'
                """,
                (now, now, sn, file_id),
            )
            try:
                event_count = scan_temperature_calibration_package(sn, file_id)
                finished = datetime.now().replace(microsecond=0)
                execute_local(
                    """
                    UPDATE device_log_event_scans
                    SET scan_status='completed', scan_version=%s, event_count=%s,
                        finished_at=%s, updated_at=%s
                    WHERE sn=%s AND source_file_id=%s AND scan_type='temperature_calibration'
                    """,
                    (LOG_EVENT_SCAN_VERSION, event_count, finished, finished, sn, file_id),
                )
            except Exception as exc:
                finished = datetime.now().replace(microsecond=0)
                execute_local(
                    """
                    UPDATE device_log_event_scans
                    SET scan_status='failed', error_message=%s, finished_at=%s, updated_at=%s
                    WHERE sn=%s AND source_file_id=%s AND scan_type='temperature_calibration'
                    """,
                    (str(exc)[:1000], finished, finished, sn, file_id),
                )
            gc.collect()
    finally:
        with LOG_EVENT_SCAN_WORKERS_LOCK:
            LOG_EVENT_SCAN_WORKERS.pop(sn, None)

def kick_temperature_calibration_scan(sn):
    with LOG_EVENT_SCAN_WORKERS_LOCK:
        running = LOG_EVENT_SCAN_WORKERS.get(sn)
        if running and running.is_alive():
            return False
        thread = threading.Thread(
            target=temperature_calibration_scan_worker,
            args=(sn,),
            name=f"zhiku-calibration-scan-{sn[-6:]}",
            daemon=True,
        )
        LOG_EVENT_SCAN_WORKERS[sn] = thread
        thread.start()
        return True

def temperature_calibration_payload(sn, start, end, rows):
    device_meta = get_log_evidence_device_meta(sn)
    package_ids = [int(row['id']) for row in rows if row.get('id')]
    scans = []
    if package_ids:
        placeholders = ','.join(['%s'] * len(package_ids))
        scans = fetch_all(
            f"""
            SELECT source_file_id, scan_status, event_count, error_message, attempts,
                   queued_at, started_at, finished_at
            FROM device_log_event_scans
            WHERE sn=%s AND scan_type='temperature_calibration'
              AND source_file_id IN ({placeholders})
            ORDER BY COALESCE(finished_at, started_at, queued_at) DESC
            """,
            tuple([sn] + package_ids),
        )
    scan_map = {int(row['source_file_id']): row for row in scans}
    package_map = {int(row['id']): row for row in rows if row.get('id')}
    event_rows = fetch_all(
        """
        SELECT id, source_file_id, event_type, event_label, event_time, matched_keyword,
               source_log_name, line_no, line_hash, raw_log_excerpt, evidence_level, event_hash
        FROM device_log_events
        WHERE sn=%s AND event_category='temperature_calibration'
          AND event_time >= %s AND event_time < %s
        ORDER BY event_time DESC, id DESC
        LIMIT 1000
        """,
        (sn, start, end),
    )
    job_rows = []
    if package_ids:
        placeholders = ','.join(['%s'] * len(package_ids))
        job_rows = fetch_all(
            f"""
            SELECT id, source_file_id, cook_start_time, cook_end_time
            FROM cook_jobs
            WHERE sn=%s AND source_file_id IN ({placeholders})
              AND cook_start_time < %s AND cook_end_time >= %s
            ORDER BY cook_start_time
            """,
            tuple([sn] + package_ids + [end, start]),
        )
    jobs_by_package = defaultdict(list)
    for job in job_rows:
        jobs_by_package[int(job.get('source_file_id') or 0)].append(job)
    deduped = []
    seen = set()
    for event in event_rows:
        key = (
            event.get('event_time'),
            re.sub(r'\s+', ' ', str(event.get('raw_log_excerpt') or '')),
        )
        if key in seen:
            continue
        seen.add(key)
        package = package_map.get(int(event.get('source_file_id') or 0), {})
        event['source_zip'] = package.get('file_name')
        event['source_file'] = event.get('source_log_name')
        event['timestamp'] = event.get('event_time')
        event['date'] = str(event.get('event_time') or '')[:10] or None
        event['raw_line'] = event.get('raw_log_excerpt')
        event['line_hash'] = event.get('line_hash') or event.get('event_hash')
        event['event_type'] = event.get('event_type') or 'temperature_calibration_event'
        event['evidence_level'] = event.get('evidence_level') or ('高' if event.get('line_no') else '中')
        event_time = to_mysql_dt(event.get('event_time'))
        matched_job_id = None
        if event_time:
            for job in jobs_by_package.get(int(event.get('source_file_id') or 0), []):
                job_start = to_mysql_dt(job.get('cook_start_time'))
                job_end = to_mysql_dt(job.get('cook_end_time'))
                if job_start and job_end and job_start - timedelta(minutes=2) <= event_time <= job_end + timedelta(minutes=2):
                    matched_job_id = job.get('id')
                    break
        event['matched_job_id'] = matched_job_id
        event['sn'] = sn
        event['customer'] = device_meta['customer']
        event['region'] = device_meta['region']
        event['package_name'] = package.get('file_name')
        deduped.append(event)
    status_counts = Counter((row.get('scan_status') or 'pending') for row in scans)
    pending_count = max(0, len(package_ids) - len(scans))
    failed_rows = [
        {
            'source_file_id': row.get('source_file_id'),
            'package_name': (package_map.get(int(row.get('source_file_id') or 0)) or {}).get('file_name'),
            'error': row.get('error_message'),
        }
        for row in scans if row.get('scan_status') == 'failed'
    ][:12]
    event_days = {str(row.get('event_time'))[:10] for row in deduped if row.get('event_time')}
    event_type_counts = Counter(row.get('event_type') or 'unknown' for row in deduped)
    return {
        'sn': sn,
        'range': {
            'start_date': start.strftime("%Y-%m-%d"),
            'end_date': (end - timedelta(days=1)).strftime("%Y-%m-%d"),
        },
        'summary': {
            'event_count': len(deduped),
            'completed_event_count': int(event_type_counts.get('temperature_calibration_completed', 0)),
            'failed_event_count': int(event_type_counts.get('temperature_calibration_failed', 0)),
            'parameter_event_count': int(event_type_counts.get('temperature_calibration_parameter', 0)),
            'status_event_count': int(event_type_counts.get('temperature_calibration_status', 0)),
            'event_day_count': len(event_days),
            'package_count': len(package_ids),
            'completed_packages': int(status_counts.get('completed', 0)),
            'queued_packages': int(status_counts.get('queued', 0)),
            'scanning_packages': int(status_counts.get('scanning', 0)),
            'failed_packages': int(status_counts.get('failed', 0)),
            'pending_packages': pending_count,
            'completion_rate': round(status_counts.get('completed', 0) * 100 / len(package_ids), 1) if package_ids else 0,
        },
        'events': deduped,
        'failed_packages': failed_rows,
        'status_explanation': {
            'completed': '已扫描并写入本地事件库，后续查询不再下载该日志包。',
            'queued': '等待后台轻量扫描，只提取校准事件，不依赖菜谱匹配。',
            'scanning': '正在下载并扫描日志包。',
            'failed': '该日志包扫描失败，可重新补齐；其他包不受影响。',
        },
    }

def recover_stale_parsing_packages(sn=None):
    if not ensure_analytics_db():
        return 0
    args = [AUTO_PARSE_STALE_MINUTES, AUTO_PARSE_STALE_MINUTES]
    sn_sql = ''
    if sn:
        sn_sql = 'AND sn=%s'
        args.append(sn)
    now = datetime.now().replace(microsecond=0)
    execute_local(
        f"""
        UPDATE device_log_packages
        SET parse_status='parse_failed',
            ui_status='解析失败',
            error_message=COALESCE(error_message, '解析任务中断或超时，已自动释放；可重新排队或手动解析。'),
            updated_at=%s
        WHERE parse_status='parsing'
          AND storage_status!='stored'
          AND (
              (last_attempt_at IS NOT NULL AND TIMESTAMPDIFF(MINUTE, last_attempt_at, NOW()) >= %s)
              OR TIMESTAMPDIFF(MINUTE, updated_at, NOW()) >= %s
          )
          {sn_sql}
        """,
        tuple([now] + args),
    )
    row = fetch_one(
        f"""
        SELECT COUNT(*) AS c
        FROM device_log_packages
        WHERE parse_status='parse_failed'
          AND error_message LIKE '解析任务中断或超时%%'
          {sn_sql}
        """,
        (sn,) if sn else (),
    )
    return int((row or {}).get('c') or 0)

def download_log_zip(url):
    limit = MAX_LOG_ANALYSIS_DOWNLOAD_MB * 1024 * 1024
    req = urllib.request.Request(url, headers={'User-Agent': 'ZhikuDashboard/1.0'})
    with urllib.request.urlopen(req, timeout=45) as resp:
        data = bytearray()
        while True:
            chunk = resp.read(1024 * 512)
            if not chunk:
                break
            data.extend(chunk)
            if len(data) > limit:
                raise HTTPException(status_code=413, detail=f"Log file exceeds {MAX_LOG_ANALYSIS_DOWNLOAD_MB}MB analysis limit")
    return bytes(data)

def text_from_zip_member(zf, info, max_bytes=8 * 1024 * 1024):
    if info.file_size > max_bytes:
        data = zf.open(info).read(max_bytes)
    else:
        data = zf.read(info)
    return data.decode('utf-8', errors='ignore')

def get_log_evidence_device_meta(sn, robot=None):
    device = robot or {}
    if not device:
        device = fetch_one(
            "SELECT company_id FROM sop_robot WHERE machinecode=%s LIMIT 1",
            (sn,),
            source=True,
            database='btyc',
        ) or {}
    company = get_company_info(device.get('company_id'))
    return {
        'customer': company.get('common_name') or company.get('company_name') or '未知',
        'region': company.get('geo_cityname') or company.get('geo_pname') or company.get('area_code') or '未知',
    }

def resolve_log_package_source(sn, source_file_id=None, source_zip=''):
    if source_file_id:
        row = fetch_one(
            "SELECT id, sn, file_name, file_length, pic AS url, create_time, cos_deleted "
            "FROM machine_ftp WHERE id=%s AND sn=%s LIMIT 1",
            (int(source_file_id), sn),
            source=True,
            database='btyc',
        )
    elif source_zip:
        row = fetch_one(
            "SELECT id, sn, file_name, file_length, pic AS url, create_time, cos_deleted "
            "FROM machine_ftp WHERE sn=%s AND file_name=%s ORDER BY id DESC LIMIT 1",
            (sn, source_zip),
            source=True,
            database='btyc',
        )
    else:
        raise HTTPException(status_code=400, detail="缺少 source_file_id 或 source_zip，无法定位日志包")
    if not row:
        raise HTTPException(status_code=404, detail="未找到来源日志包，请先补齐或重新同步该设备日志")
    if row.get('cos_deleted') or not row.get('url'):
        raise HTTPException(status_code=409, detail="来源日志包不可下载或已被删除，请先补齐日志包")
    return row

def load_log_evidence_file(sn, source_file_id=None, source_zip='', source_file=''):
    package = resolve_log_package_source(sn, source_file_id=source_file_id, source_zip=source_zip)
    cache_key = hashlib.sha256(
        f"{sn}|{int(package['id'])}|{source_file}".encode('utf-8')
    ).hexdigest()
    cache_path = LOG_EVIDENCE_FILE_CACHE_DIR / f"{cache_key}.log"
    cache_name_path = LOG_EVIDENCE_FILE_CACHE_DIR / f"{cache_key}.name"
    if cache_path.exists() and cache_name_path.exists():
        actual_source_file = cache_name_path.read_text(encoding='utf-8').strip() or source_file
        text = cache_path.read_text(encoding='utf-8', errors='ignore')
        return package, actual_source_file, text.splitlines()
    try:
        zip_bytes = download_log_zip(package['url'])
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
    except HTTPException:
        raise
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=422, detail="来源日志包不是有效 ZIP，无法读取原始证据") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"来源日志包下载失败：{str(exc)[:180]}") from exc
    with zf:
        candidates = [
            info for info in zf.infolist()
            if not info.is_dir() and (
                info.filename == source_file
                or Path(info.filename).name == Path(source_file or '').name
            )
        ]
        if not candidates:
            raise HTTPException(status_code=404, detail="日志包中未找到来源日志文件，请重新扫描该日志包")
        info = candidates[0]
        text = text_from_zip_member(zf, info, max_bytes=32 * 1024 * 1024)
        actual_source_file = info.filename
    try:
        cache_path.write_text(text, encoding='utf-8')
        cache_name_path.write_text(actual_source_file, encoding='utf-8')
    except OSError:
        pass
    del zip_bytes
    gc.collect()
    return package, actual_source_file, text.splitlines()

def find_matched_job_id(sn, source_file_id, timestamp):
    event_time = to_mysql_dt(timestamp)
    if not event_time:
        return None
    row = fetch_one(
        """
        SELECT id
        FROM cook_jobs
        WHERE sn=%s AND source_file_id=%s
          AND cook_start_time IS NOT NULL AND cook_end_time IS NOT NULL
          AND %s BETWEEN DATE_SUB(cook_start_time, INTERVAL 2 MINUTE)
                     AND DATE_ADD(cook_end_time, INTERVAL 2 MINUTE)
        ORDER BY ABS(TIMESTAMPDIFF(SECOND, cook_start_time, %s))
        LIMIT 1
        """,
        (sn, int(source_file_id), event_time, event_time),
    )
    return row.get('id') if row else None

def locate_log_target_line(sn, source_file_id, source_file, lines, line_hash='', line_no=None, timestamp='', raw_line=''):
    if line_hash:
        for idx, line in enumerate(lines):
            candidate = build_log_line_hash(sn, source_file_id, source_file, idx + 1, line)
            if hmac.compare_digest(candidate, line_hash):
                return idx
    if line_no and 1 <= int(line_no) <= len(lines):
        return int(line_no) - 1
    if raw_line:
        normalized = re.sub(r'\s+', ' ', str(raw_line).strip())
        for idx, line in enumerate(lines):
            if re.sub(r'\s+', ' ', line.strip()) == normalized:
                return idx
    target_time = to_mysql_dt(timestamp)
    if target_time:
        nearest = None
        for idx, line in enumerate(lines):
            parsed = parse_log_ts(line)
            if not parsed:
                continue
            delta = abs((parsed - target_time).total_seconds())
            if nearest is None or delta < nearest[0]:
                nearest = (delta, idx)
        if nearest and nearest[0] <= 300:
            return nearest[1]
    return None

def classify_log_file(name):
    base = Path(name).name.lower()
    if base.startswith('android') and base.endswith('.log'):
        return 'android_app'
    if base.startswith('debug'):
        return 'mcu_debug'
    if base == 'main_board.log':
        return 'main_board'
    if base == 'oildrum_board.log':
        return 'oildrum_board'
    if base.startswith('temperature') and base.endswith('.log'):
        return 'temperature'
    return 'other'

def log_kind_label(kind):
    return LOG_KIND_DEFS.get(kind, LOG_KIND_DEFS['other'])['label']

def log_specs_payload():
    return {
        'file_kinds': [
            {'kind': kind, **meta}
            for kind, meta in LOG_KIND_DEFS.items()
        ],
        'oildrum_fields': OILDRUM_FIELD_SPECS,
        'fault_code_source': {
            'file': FAULT_CODE_DATA_PATH.name,
            'sheet': '故障码清单V5.0-屏蔽',
            'code_rule': '日志形如“故障码：1_1_86_F”时，按第 1 段 + 第 2 段补两位 + 第 4 段补两位转换，例如 1_1_86_F => 1010F。',
        },
    }

def load_fault_code_records():
    if not FAULT_CODE_DATA_PATH.exists():
        return []
    try:
        payload = json.loads(FAULT_CODE_DATA_PATH.read_text())
        return payload.get('records', [])
    except Exception:
        return []

FAULT_CODE_RECORDS = load_fault_code_records()
FAULT_CODE_BY_ERROR = {str(row.get('error_code', '')).upper(): row for row in FAULT_CODE_RECORDS if row.get('error_code')}
FAULT_CODE_BY_ERR = {str(row.get('err_code', '')).upper(): row for row in FAULT_CODE_RECORDS if row.get('err_code')}

def normalize_fault_code(value):
    return str(value or '').strip().upper().replace(' ', '')

def fault_code_from_tuple(parts):
    if len(parts) < 4:
        return ''
    first = normalize_fault_code(parts[0])
    second = normalize_fault_code(parts[1]).zfill(2)
    fourth = normalize_fault_code(parts[3]).zfill(2)
    return f"{first}{second}{fourth}"

def extract_fault_matches(text, file_name, kind, limit=80):
    counter = Counter()
    examples = {}
    for line in text.splitlines()[:120000]:
        line_upper = line.upper()
        has_fault_context = bool(re.search(r'故障|报警|异常|FAULT|ERR|ERROR|MALFUNCTION|HEATERGETERRCODE|ERRCODE|ERROR\\s*CODE', line_upper))
        for match in re.finditer(r'(?:故障码|FAULT|ERR(?:OR)?\\s*CODE)[:：=\\s]+([0-9A-F]+)_([0-9A-F]+)_([0-9A-F]+)_([0-9A-F]+)', line_upper):
            code = fault_code_from_tuple(match.groups())
            if code in FAULT_CODE_BY_ERROR:
                counter[code] += 1
                examples.setdefault(code, line.strip()[:180])
        if not has_fault_context:
            continue
        for match in re.finditer(r'(?<![A-Z0-9])([1-9][0-9A-F]{4,5})(?![A-Z0-9])', line_upper):
            code = normalize_fault_code(match.group(1))
            if code in FAULT_CODE_BY_ERROR:
                counter[code] += 1
                examples.setdefault(code, line.strip()[:180])
        for match in re.finditer(r'(?<![A-Z0-9])([A-Z]{2,4}\\d{2})(?![A-Z0-9])', line_upper):
            meta = FAULT_CODE_BY_ERR.get(normalize_fault_code(match.group(1)))
            if meta:
                code = meta.get('error_code')
                counter[code] += 1
                examples.setdefault(code, line.strip()[:180])

    rows = []
    for code, count in counter.most_common(limit):
        meta = FAULT_CODE_BY_ERROR.get(code, {})
        rows.append({
            'file_name': file_name,
            'kind': kind,
            'kind_label': log_kind_label(kind),
            'error_code': code,
            'err_code': meta.get('err_code'),
            'module': meta.get('module'),
            'meaning': meta.get('meaning'),
            'err_message': meta.get('err_message'),
            'priority': meta.get('priority'),
            'remark': meta.get('remark'),
            'count': count,
            'example': examples.get(code, ''),
        })
    return rows

def build_log_category_summary(file_summary):
    grouped = {}
    for row in file_summary:
        kind = row.get('kind') or 'other'
        meta = LOG_KIND_DEFS.get(kind, LOG_KIND_DEFS['other'])
        item = grouped.setdefault(kind, {
            'kind': kind,
            'label': meta['label'],
            'purpose': meta['purpose'],
            'key_patterns': meta.get('key_patterns', []),
            'file_count': 0,
            'total_lines': 0,
            'total_size_mb': 0,
            'first_time': None,
            'last_time': None,
        })
        item['file_count'] += 1
        item['total_lines'] += int(row.get('lines') or 0)
        item['total_size_mb'] += float(row.get('size_mb') or 0)
        if row.get('first_time'):
            item['first_time'] = row['first_time'] if not item['first_time'] else min(item['first_time'], row['first_time'])
        if row.get('last_time'):
            item['last_time'] = row['last_time'] if not item['last_time'] else max(item['last_time'], row['last_time'])
    result = []
    order = ['oildrum_board', 'temperature', 'main_board', 'mcu_debug', 'android_app', 'other']
    for kind in order:
        if kind in grouped:
            grouped[kind]['total_size_mb'] = round(grouped[kind]['total_size_mb'], 2)
            result.append(grouped[kind])
    return result

def analyze_oildrum_text(text):
    status_re = re.compile(r'^\[(\d{2})\.(\d{2})\.(\d{2}) (\d{2}:\d{2}:\d{2})\]\s+dev status\s+(-?\d+)\s+(-?\d+)\s+(-?\d+)\s+(-?\d+)(?:\s+([\d.\-]+))?')
    pump_start_re = re.compile(r'pump hall start\s+(-?\d+)\s+(-?[\d.]+)', re.I)
    pump_finish_re = re.compile(r'pump hall finish\s+(-?\d+),\s*time\s+(\d+),\s*state\s+(\d+)', re.I)
    read_power_re = re.compile(r'read power\s+(\d+)\s+(-?\d+)\s+(-?\d+)', re.I)
    heat_re = re.compile(r'(oil|pipe) heat enable\s+(\d+)\s+(-?[\d.]+)', re.I)
    sw_re = re.compile(r'read sw v\s+(\d+)\s+(\d+)\s+(\d+)', re.I)
    samples = []
    events = []
    event_counts = Counter()
    motor_events = []
    heat_events = []
    power_events = []
    sw_versions = []
    for line in text.splitlines():
        clean = line.strip()
        ts = parse_log_ts(clean)
        match = status_re.search(clean)
        if match:
            if not ts:
                continue
            bucket = int(match.group(5)) / 10
            pipe = int(match.group(6)) / 10
            reserved = int(match.group(7))
            motor_rpm = int(match.group(8))
            heater_pwm = float(match.group(9)) if match.group(9) else None
            samples.append({
                't': ts,
                'bucket_temp': bucket,
                'pipe_temp': pipe,
                'oil_temp': bucket,
                'reserved': reserved,
                'motor_rpm': motor_rpm,
                'heater_pwm': heater_pwm,
                'current': heater_pwm,
            })
            continue
        if not ts:
            continue
        event_text = re.sub(r'^\[[^\]]+\]\s*', '', clean)[:180]
        start = pump_start_re.search(clean)
        if start:
            hall_count = int(start.group(1))
            direction = '投油' if hall_count < 0 else '反抽'
            event_counts[f'电机开始-{direction}'] += 1
            item = {
                'time': ts.isoformat(sep=' '),
                'type': 'pump_start',
                'label': f'电机开始{direction}',
                'hall_count': hall_count,
                'motor_pwm': float(start.group(2)),
                'direction': direction,
                'text': event_text,
            }
            motor_events.append(item)
            events.append(item)
            continue
        finish = pump_finish_re.search(clean)
        if finish:
            state = int(finish.group(3))
            event_counts['电机完成-霍尔异常' if state == 1 else '电机完成-正常'] += 1
            item = {
                'time': ts.isoformat(sep=' '),
                'type': 'pump_finish',
                'label': '电机完成',
                'hall_count': int(finish.group(1)),
                'duration_ms': int(finish.group(2)),
                'hall_phase_state': state,
                'hall_phase_label': '异常' if state == 1 else '正常',
                'text': event_text,
            }
            motor_events.append(item)
            events.append(item)
            continue
        power = read_power_re.search(clean)
        if power:
            limited = int(power.group(1)) == 1
            event_counts['功率限制-有' if limited else '功率限制-无'] += 1
            item = {
                'time': ts.isoformat(sep=' '),
                'type': 'read_power',
                'label': '读取功率限制',
                'power_limited': limited,
                'text': event_text,
            }
            power_events.append(item)
            events.append(item)
            continue
        heat = heat_re.search(clean)
        if heat:
            target = '油桶' if heat.group(1).lower() == 'oil' else '油管'
            enabled = int(heat.group(2)) == 1
            target_temp = float(heat.group(3))
            event_counts[f'{target}加热-{"使能" if enabled else "关闭"}'] += 1
            item = {
                'time': ts.isoformat(sep=' '),
                'type': f'{heat.group(1).lower()}_heat_enable',
                'label': f'{target}加热目标',
                'target': target,
                'enabled': enabled,
                'target_temp': target_temp,
                'text': event_text,
            }
            heat_events.append(item)
            events.append(item)
            continue
        sw = sw_re.search(clean)
        if sw:
            version = f"{sw.group(1)}.{sw.group(2)}.{sw.group(3)}"
            event_counts['软件版本读取'] += 1
            item = {
                'time': ts.isoformat(sep=' '),
                'type': 'read_sw_version',
                'label': '软件版本',
                'version': version,
                'text': event_text,
            }
            sw_versions.append(item)
            events.append(item)
    samples.sort(key=lambda row: row['t'])
    if not samples:
        return {
            'sample_count': 0,
            'series': [],
            'daily': [],
            'events': events[:120],
            'event_counts': [{'category': k, 'count': v} for k, v in event_counts.most_common()],
            'motor_events': motor_events[:120],
            'heat_events': heat_events[:120],
            'power_events': power_events[:120],
            'sw_versions': sw_versions[:20],
            'field_specs': OILDRUM_FIELD_SPECS,
        }

    by_10 = defaultdict(lambda: {'bucket': [], 'pipe': [], 'motor_rpm': [], 'heater_pwm': []})
    by_day = defaultdict(lambda: {'bucket': [], 'pipe': [], 'motor_rpm': [], 'heater_pwm': [], 'first': None, 'last': None, 'samples': 0})
    over_100 = 0
    negative_pipe = 0
    for row in samples:
        if row['bucket_temp'] > 100:
            over_100 += 1
        if row['pipe_temp'] < 0:
            negative_pipe += 1
        minute = row['t'].replace(minute=(row['t'].minute // 10) * 10, second=0)
        if 0 <= row['bucket_temp'] <= 160:
            by_10[minute]['bucket'].append(row['bucket_temp'])
        if 0 <= row['pipe_temp'] <= 120:
            by_10[minute]['pipe'].append(row['pipe_temp'])
        if row['motor_rpm'] is not None:
            by_10[minute]['motor_rpm'].append(row['motor_rpm'])
        if row['heater_pwm'] is not None:
            by_10[minute]['heater_pwm'].append(row['heater_pwm'])
        day = row['t'].strftime('%Y-%m-%d')
        item = by_day[day]
        if 0 <= row['bucket_temp'] <= 160:
            item['bucket'].append(row['bucket_temp'])
        if 0 <= row['pipe_temp'] <= 120:
            item['pipe'].append(row['pipe_temp'])
        if row['motor_rpm'] is not None:
            item['motor_rpm'].append(row['motor_rpm'])
        if row['heater_pwm'] is not None:
            item['heater_pwm'].append(row['heater_pwm'])
        item['first'] = row['t'] if item['first'] is None else min(item['first'], row['t'])
        item['last'] = row['t'] if item['last'] is None else max(item['last'], row['t'])
        item['samples'] += 1

    series = []
    for ts, agg in sorted(by_10.items()):
        series.append({
            'time': ts.isoformat(sep=' '),
            'bucket_temp': round(statistics.mean(agg['bucket']), 2) if agg['bucket'] else None,
            'pipe_temp': round(statistics.mean(agg['pipe']), 2) if agg['pipe'] else None,
            'motor_rpm': round(statistics.mean(agg['motor_rpm']), 2) if agg['motor_rpm'] else None,
            'heater_pwm': round(statistics.mean(agg['heater_pwm']), 2) if agg['heater_pwm'] else None,
        })
    daily = []
    for day, agg in sorted(by_day.items()):
        daily.append({
            'day': day,
            'first_time': agg['first'].isoformat(sep=' ') if agg['first'] else None,
            'last_time': agg['last'].isoformat(sep=' ') if agg['last'] else None,
            'bucket_min': round(min(agg['bucket']), 1) if agg['bucket'] else None,
            'bucket_avg': round(statistics.mean(agg['bucket']), 1) if agg['bucket'] else None,
            'bucket_max': round(max(agg['bucket']), 1) if agg['bucket'] else None,
            'pipe_min': round(min(agg['pipe']), 1) if agg['pipe'] else None,
            'pipe_avg': round(statistics.mean(agg['pipe']), 1) if agg['pipe'] else None,
            'pipe_max': round(max(agg['pipe']), 1) if agg['pipe'] else None,
            'motor_rpm_avg': round(statistics.mean(agg['motor_rpm']), 1) if agg['motor_rpm'] else None,
            'heater_pwm_avg': round(statistics.mean(agg['heater_pwm']), 2) if agg['heater_pwm'] else None,
            'samples': agg['samples'],
        })

    values_bucket = [r['bucket_temp'] for r in samples if 0 <= r['bucket_temp'] <= 160]
    values_pipe = [r['pipe_temp'] for r in samples if 0 <= r['pipe_temp'] <= 120]
    values_rpm = [r['motor_rpm'] for r in samples if r.get('motor_rpm') is not None]
    values_pwm = [r['heater_pwm'] for r in samples if r.get('heater_pwm') is not None]
    finish_events = [row for row in motor_events if row.get('type') == 'pump_finish']
    return {
        'sample_count': len(samples),
        'first_time': samples[0]['t'].isoformat(sep=' '),
        'last_time': samples[-1]['t'].isoformat(sep=' '),
        'bucket_min': round(min(values_bucket), 1) if values_bucket else None,
        'bucket_avg': round(statistics.mean(values_bucket), 1) if values_bucket else None,
        'bucket_max': round(max(values_bucket), 1) if values_bucket else None,
        'pipe_min': round(min(values_pipe), 1) if values_pipe else None,
        'pipe_avg': round(statistics.mean(values_pipe), 1) if values_pipe else None,
        'pipe_max': round(max(values_pipe), 1) if values_pipe else None,
        'bucket_over_100_count': over_100,
        'pipe_negative_count': negative_pipe,
        'motor_rpm_max': max(values_rpm) if values_rpm else None,
        'heater_pwm_max': round(max(values_pwm), 2) if values_pwm else None,
        'motor_start_count': len([row for row in motor_events if row.get('type') == 'pump_start']),
        'motor_finish_count': len(finish_events),
        'hall_abnormal_count': len([row for row in finish_events if row.get('hall_phase_state') == 1]),
        'max_motor_duration_ms': max([row.get('duration_ms', 0) for row in finish_events], default=None),
        'series': series[-900:],
        'daily': daily,
        'events': events[:160],
        'event_counts': [{'category': k, 'count': v} for k, v in event_counts.most_common()],
        'motor_events': motor_events[:160],
        'heat_events': heat_events[:160],
        'power_events': power_events[:160],
        'sw_versions': sw_versions[:20],
        'field_specs': OILDRUM_FIELD_SPECS,
    }

def analyze_temperature_text(text):
    temp_re = re.compile(r'^\[(\d{4}-\d{2}-\d{2})_(\d{2}:\d{2}:\d{2})\]\s+(-?\d+)_(-?\d+)_(-?\d+)')
    values = []
    first = None
    last = None
    count = 0
    for line in text.splitlines():
        match = temp_re.search(line.strip())
        if not match:
            continue
        ts = parse_log_ts(line)
        if ts:
            first = first or ts
            last = ts
        triplet = [int(match.group(i)) for i in (3, 4, 5)]
        values.extend(triplet)
        count += 1
    return {
        'sample_count': count,
        'first_time': first.isoformat(sep=' ') if first else None,
        'last_time': last.isoformat(sep=' ') if last else None,
        'min': min(values) if values else None,
        'avg': round(statistics.mean(values), 1) if values else None,
        'max': max(values) if values else None,
    }

def parse_temperature_series(text, max_points=600000):
    temp_re = re.compile(r'^\[(\d{4}-\d{2}-\d{2})_(\d{2}:\d{2}:\d{2})\]\s+(-?\d+)_(-?\d+)_(-?\d+)')
    series = []
    seen_seconds = set()
    for line in text.splitlines():
        match = temp_re.search(line.strip())
        if not match:
            continue
        ts = parse_log_ts(line)
        if not ts:
            continue
        second_key = int(ts.timestamp())
        if second_key in seen_seconds:
            continue
        seen_seconds.add(second_key)
        raw = [int(match.group(i)) for i in (3, 4, 5)]
        series.append({
            'ts': ts,
            'time': ts.isoformat(sep=' '),
            'temp_1': raw[0],
            'temp_2': raw[1],
            'temp_3': raw[2],
            'filtered_temp': raw[0],
            'infrared_temp': raw[1],
            'output_temp': raw[2],
            'pot_temp': raw[2],
            'raw': '_'.join(str(x) for x in raw),
            'source': 'temperature.log 温度字段：滤波温度/红外实际温度/输出温度，单位℃',
        })
        if len(series) >= max_points:
            break
    return series

def is_android_log_name(name):
    base = (name or '').lower()
    return base.startswith('android') and base.endswith('.log')

def is_temperature_log_name(name):
    base = (name or '').lower()
    return base.startswith('temperature') and base.endswith('.log')

def build_log_line_hash(sn, source_file_id, source_file, line_no, raw_line):
    payload = f"{sn}|{int(source_file_id)}|{source_file}|{int(line_no)}|{raw_line}"
    return hashlib.sha256(payload.encode('utf-8')).hexdigest()

TEMPERATURE_CALIBRATION_PATTERNS = [
    re.compile(r'温度.{0,16}(?:校准|校正|矫正|标定|修正)', re.I),
    re.compile(r'(?:校准|校正|矫正|标定|修正).{0,16}温度', re.I),
    re.compile(r'(?:temperature|temp).{0,20}(?:calibrat|correct|adjust)', re.I),
    re.compile(r'(?:calibrat|correct|adjust).{0,20}(?:temperature|temp)', re.I),
]

def classify_temperature_calibration_line(line):
    text = str(line or '').strip()
    matched = None
    for pattern in TEMPERATURE_CALIBRATION_PATTERNS:
        hit = pattern.search(text)
        if hit:
            matched = hit.group(0)
            break
    if not matched:
        return None
    if re.search(r'操作结果[：:]\s*false|失败|异常|取消|fail|error|cancel', text, re.I):
        event_type = 'temperature_calibration_failed'
        event_label = '温度校正失败'
    elif re.search(r'操作结果[：:]\s*true', text, re.I) or re.search(r'["\']opType["\']\s*:\s*["\']TempCalibrate', text, re.I):
        event_type = 'temperature_calibration_completed'
        event_label = '温度校正完成'
    elif re.search(r'最近.{0,8}(?:校准|校正|矫正|标定|修正).{0,8}时间', text, re.I):
        event_type = 'temperature_calibration_status'
        event_label = '校正状态读取'
    elif re.search(r'设置.{0,12}(?:参数|标定温度)|tempCoefficient', text, re.I):
        event_type = 'temperature_calibration_parameter'
        event_label = '校正参数设置'
    elif re.search(r'完成|成功|结束|完毕|finish|complete|success', text, re.I):
        event_type = 'temperature_calibration_completed'
        event_label = '温度校正完成'
    elif re.search(r'开始|启动|进入|start|begin', text, re.I):
        event_type = 'temperature_calibration_started'
        event_label = '温度校正开始'
    else:
        event_type = 'temperature_calibration_event'
        event_label = '温度校正动作'
    return {
        'event_category': 'temperature_calibration',
        'event_type': event_type,
        'event_label': event_label,
        'matched_keyword': matched[:120],
    }

def scan_temperature_calibration_text(text, source_log_name, sn, source_file_id, fallback_time=None):
    events = []
    seen = set()
    for line_no, line in enumerate(str(text or '').splitlines(), start=1):
        classification = classify_temperature_calibration_line(line)
        if not classification:
            continue
        parsed_ts = parse_log_ts(line)
        ts = parsed_ts or fallback_time
        raw = line.strip()[:1200]
        dedupe_key = (ts.isoformat() if ts else '', re.sub(r'\s+', ' ', raw))
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        events.append({
            **classification,
            'event_time': ts,
            'source_log_name': source_log_name,
            'line_no': line_no,
            'line_hash': build_log_line_hash(sn, source_file_id, source_log_name, line_no, line),
            'raw_log_excerpt': raw,
            'evidence_level': '高' if parsed_ts else '中',
        })
    return events

def persist_device_log_events(sn, source_file_id, scan_type, events):
    if not ensure_analytics_db():
        return 0
    now = datetime.now().replace(microsecond=0)
    execute_local(
        "DELETE FROM device_log_events WHERE sn=%s AND source_file_id=%s AND event_category=%s",
        (sn, int(source_file_id), scan_type),
    )
    payload = []
    for event in events or []:
        raw = str(event.get('raw_log_excerpt') or '')
        event_time = to_mysql_dt(event.get('event_time'))
        event_hash = hashlib.sha256(
            f"{sn}|{int(source_file_id)}|{scan_type}|{event_time}|{raw}".encode('utf-8')
        ).hexdigest()
        payload.append((
            sn, int(source_file_id), scan_type, event.get('event_type') or 'event',
            event.get('event_label') or '日志事件', event_time, event.get('matched_keyword'),
            event.get('source_log_name'), event.get('line_no'), event.get('line_hash'), raw,
            event.get('evidence_level') or '中', event_hash, LOG_EVENT_SCAN_VERSION, now,
        ))
    executemany_local(
        """
        INSERT IGNORE INTO device_log_events(
            sn, source_file_id, event_category, event_type, event_label, event_time,
            matched_keyword, source_log_name, line_no, line_hash, raw_log_excerpt,
            evidence_level, event_hash, scan_version, created_at
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        payload,
    )
    return len(payload)

def scan_temperature_calibration_package(sn, source_file_id):
    file_row = fetch_one(
        "SELECT id, sn, file_name, file_length, pic AS url, create_time, cos_deleted "
        "FROM machine_ftp WHERE id=%s AND sn=%s LIMIT 1",
        (int(source_file_id), sn),
        source=True,
        database='btyc',
    )
    if not file_row or not file_row.get('url') or file_row.get('cos_deleted'):
        raise ValueError('日志包不可下载或已被删除')
    zip_bytes = download_log_zip(file_row['url'])
    try:
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
    except zipfile.BadZipFile as exc:
        raise ValueError('下载内容不是有效 ZIP') from exc
    fallback_time = to_mysql_dt(infer_log_time_from_filename(file_row.get('file_name'))) or to_mysql_dt(file_row.get('create_time'))
    events = []
    with zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = Path(info.filename).name
            if not is_android_log_name(base):
                continue
            text = text_from_zip_member(zf, info, max_bytes=16 * 1024 * 1024)
            events.extend(scan_temperature_calibration_text(
                text,
                info.filename,
                sn,
                source_file_id,
                fallback_time=fallback_time,
            ))
    deduped = {}
    for event in events:
        key = (
            event.get('event_time'),
            re.sub(r'\s+', ' ', str(event.get('raw_log_excerpt') or '')),
        )
        deduped[key] = event
    return persist_device_log_events(
        sn,
        source_file_id,
        'temperature_calibration',
        sorted(deduped.values(), key=lambda item: item.get('event_time') or datetime.min),
    )

def parse_android_pot_temperature_series(text, max_points=300000):
    temp_re = re.compile(r'温度:_?(-?\d+)_(-?\d+)_(-?\d+)')
    power_re = re.compile(r'功率:\s*(-?\d+(?:\.\d+)?)_(-?\d+(?:\.\d+)?)')
    series = []
    seen_seconds = set()
    for line in text.splitlines():
        match = temp_re.search(line)
        if not match:
            continue
        ts = parse_log_ts(line)
        if not ts:
            continue
        second_key = int(ts.timestamp())
        if second_key in seen_seconds:
            continue
        seen_seconds.add(second_key)
        core, coil, output = [int(match.group(i)) for i in (1, 2, 3)]
        row = {
            'ts': ts,
            'time': ts.isoformat(sep=' '),
            'pot_temp': output,
            'core_temp': core,
            'coil_temp': coil,
            'output_temp': output,
            'aux_temp_1': core,
            'aux_temp_2': coil,
            'raw': f"{core}_{coil}_{output}",
            'source': 'android_app 温度字段：机芯温度/线盘温度/测量输出温度，单位℃',
        }
        power_match = power_re.search(line)
        if power_match:
            command_w = float(power_match.group(1))
            actual_w = float(power_match.group(2))
            row.update({
                'command_power_w': command_w,
                'actual_power_w': actual_w,
                'command_power_kw': round(command_w / 1000, 2),
                'actual_power_kw': round(actual_w / 1000, 2),
            })
        series.append(row)
        if len(series) >= max_points:
            break
    return series

def temperature_series_cache_path(file_id):
    path = CACHE_DIR / 'temperature_series'
    path.mkdir(parents=True, exist_ok=True)
    return path / f'{int(file_id)}.json'

def cached_temperature_series_from_file(file_row):
    file_id = file_row.get('id')
    if not file_id or not file_row.get('url'):
        return None
    path = temperature_series_cache_path(file_id)
    if path.exists():
        try:
            payload = json.loads(path.read_text())
            if payload.get('version') == 3:
                series = payload.get('series') or []
                for row in series:
                    row['ts'] = datetime.strptime(row['time'], "%Y-%m-%d %H:%M:%S")
                return {
                    'file': payload.get('file') or {},
                    'coverage': payload.get('coverage') or {},
                    'series': series,
                }
        except Exception:
            path.unlink(missing_ok=True)

    if file_row.get('cos_deleted'):
        return None
    try:
        zip_bytes = download_log_zip(file_row['url'])
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
    except Exception:
        return None

    temperature_series = []
    fallback_temperature_series = []
    with zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = Path(info.filename).name
            if is_android_log_name(base):
                temperature_series.extend(parse_android_pot_temperature_series(text_from_zip_member(zf, info, max_bytes=16 * 1024 * 1024)))
            elif is_temperature_log_name(base):
                fallback_temperature_series.extend(parse_temperature_series(text_from_zip_member(zf, info, max_bytes=16 * 1024 * 1024)))
    temperature_series.sort(key=lambda row: row['ts'])
    if not temperature_series and fallback_temperature_series:
        temperature_series = fallback_temperature_series
    if not temperature_series:
        return None

    payload = {
        'version': 3,
        'file': {
            'id': file_row.get('id'),
            'file_name': file_row.get('file_name'),
            'create_time': file_row.get('create_time'),
            'file_length': file_row.get('file_length'),
        },
        'coverage': {
            'start': temperature_series[0]['ts'].isoformat(sep=' '),
            'end': temperature_series[-1]['ts'].isoformat(sep=' '),
            'sample_count': len(temperature_series),
            'temperature_unit': 'android_app 温度字段：机芯/线盘/输出温度；temperature*.log 字段：滤波/红外/输出温度，单位℃',
        },
        'series': [{k: v for k, v in row.items() if k != 'ts'} for row in temperature_series],
    }
    try:
        path.write_text(json.dumps(payload, ensure_ascii=False, default=str))
    except Exception:
        pass
    return {'file': payload['file'], 'coverage': payload['coverage'], 'series': temperature_series}

def find_temperature_series_for_window(sn, window_start, window_end, max_files=12):
    rows = fetch_all(
        """
        SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted
        FROM machine_ftp
        WHERE sn = %s AND cos_deleted = 0
          AND create_time >= %s AND create_time <= %s
        ORDER BY create_time DESC
        LIMIT %s
        """,
        (sn, window_start - timedelta(days=2), window_end + timedelta(days=10), max_files),
        source=True,
        database='btyc',
    )
    if not rows:
        rows = fetch_all(
            """
            SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted
            FROM machine_ftp
            WHERE sn = %s AND cos_deleted = 0
            ORDER BY create_time DESC
            LIMIT %s
            """,
            (sn, max_files),
            source=True,
            database='btyc',
        )
    for row in rows:
        payload = cached_temperature_series_from_file(row)
        if not payload or not payload.get('series'):
            continue
        coverage_start = payload['series'][0]['ts']
        coverage_end = payload['series'][-1]['ts']
        if coverage_start <= window_end and coverage_end >= window_start:
            return payload
    return None

def cooking_temperature_summary_from_series(series, cook_start, cook_end):
    if not series or not cook_start or not cook_end:
        return {'covered': False, 'status': '无温度样本'}
    window_start = cook_start - timedelta(seconds=20)
    window_end = cook_end + timedelta(seconds=20)
    cook_series = [row for row in series if window_start <= row['ts'] <= window_end]
    values = [row.get('pot_temp') for row in cook_series if row.get('pot_temp') is not None]
    if not values:
        values = [
            value
            for row in cook_series
            for value in (row.get('temp_1'), row.get('temp_2'), row.get('temp_3'))
            if value is not None
        ]
    if not values:
        return {'covered': False, 'status': '作业窗口无温度样本', 'sample_count': len(cook_series)}
    return {
        'covered': True,
        'status': '已匹配',
        'sample_count': len(cook_series),
        'min_temp': round(min(values), 1),
        'max_temp': round(max(values), 1),
        'avg_temp': round(statistics.mean(values), 1),
        'start_temp': round(values[0], 1),
        'end_temp': round(values[-1], 1),
    }

def parse_main_board_actions(text, start_time, end_time):
    action_patterns = [
        ('add_oil', '加油/投油', re.compile(r'add oil|stop add oil', re.I)),
        ('add_water', '加水', re.compile(r'add water|stop add water', re.I)),
        ('roll_move', '翻锅/转锅', re.compile(r'roll mov|roll start|roll stop', re.I)),
        ('heat_info', '加热读取', re.compile(r'read heat info', re.I)),
        ('stir', '搅拌/电机', re.compile(r'stir|motor|pwm', re.I)),
        ('pump', '泵/投料', re.compile(r'pump|hall', re.I)),
        ('error', '异常/失败', re.compile(r'fail|error|err|timeout', re.I)),
    ]
    actions = []
    for line in text.splitlines():
        ts = parse_log_ts(line)
        if not ts or ts < start_time or ts > end_time:
            continue
        label = None
        kind = None
        for key, name, pattern in action_patterns:
            if pattern.search(line):
                kind = key
                label = name
                break
        if not label:
            continue
        actions.append({
            'time': ts.isoformat(sep=' '),
            'offset_seconds': int((ts - start_time).total_seconds()),
            'kind': kind,
            'label': label,
            'raw': line.strip()[:320],
        })
        if len(actions) >= 400:
            break
    return actions

def android_action_label(line):
    text = line.strip()
    if re.search(r'功率:\s*-?\d+(?:\.\d+)?_-?\d+(?:\.\d+)?', text, re.I):
        return 'power_sample', '功率/温度采样'
    checks = [
        ('cook_start', '烹饪开始', r'烹饪开始'),
        ('cook_end', '烹饪结束', r'烹饪.*结束'),
        ('power_set', '功率设置', r'功率设置|功率设置为'),
        ('power_sample', '功率/温度采样', r'功率:\s*-?\d+(?:\.\d+)?_.*温度:_'),
        ('temperature', '温控动作', r'开始检测温度|温度阻塞|设置温度上限'),
        ('roll_start', '转锅指令', r'开始转锅'),
        ('roll_result', '转锅结果', r'转锅操作_'),
        ('liquid_start', '液料投放指令', r'开始投液料'),
        ('liquid_result', '液料投放结果', r'投液料_|消耗液料|设置液料当前容量|液料投料记录'),
        ('manual_prompt', '人工投料提示', r'手动投放|CNEngine speak|speakPoll|onSpeechComplete'),
        ('send_msg', '下发指令', r'sendMsg frame='),
        ('frame_result', '指令回执', r'readResult|read line|findResult'),
        ('data_collect', '数据采集', r'DataCollectManager'),
        ('scene', '状态/场景', r'updateRobotScene|add to poll|readResult'),
        ('warning', '异常/失败', r'失败|error|err|timeout|异常'),
    ]
    for kind, label, pattern in checks:
        if re.search(pattern, text, re.I):
            return kind, label
    if any(key in text for key in ['开始', '成功', '设置', '投料', '加热', '温度', '转锅', '语音']):
        return 'android_action', '安卓动作'
    return None, None

def parse_android_action_rows(text, max_actions=120000):
    actions = []
    temp_re = re.compile(r'温度:_?(-?\d+)_(-?\d+)_(-?\d+)')
    power_re = re.compile(r'功率:\s*(-?\d+(?:\.\d+)?)_(-?\d+(?:\.\d+)?)')
    seen_sample_seconds = set()
    for line in text.splitlines():
        ts = parse_log_ts(line)
        if not ts:
            continue
        kind, label = android_action_label(line)
        if not label:
            continue
        if kind == 'power_sample':
            second_key = int(ts.timestamp())
            if second_key in seen_sample_seconds:
                continue
            seen_sample_seconds.add(second_key)
        temperature = None
        temp_match = temp_re.search(line)
        if temp_match:
            core, coil, output = [int(temp_match.group(i)) for i in (1, 2, 3)]
            temperature = {
                'time': ts.isoformat(sep=' '),
                'pot_temp': output,
                'core_temp': core,
                'coil_temp': coil,
                'output_temp': output,
                'aux_temp_1': core,
                'aux_temp_2': coil,
                'raw': f"{core}_{coil}_{output}",
                'delta_seconds': 0,
            }
        power_payload = None
        power_match = power_re.search(line)
        if power_match:
            command_w = float(power_match.group(1))
            actual_w = float(power_match.group(2))
            power_payload = {
                'command_power_w': command_w,
                'actual_power_w': actual_w,
                'command_power_kw': round(command_w / 1000, 2),
                'actual_power_kw': round(actual_w / 1000, 2),
                'raw': f"{power_match.group(1)}_{power_match.group(2)}",
            }
        actions.append({
            'ts': ts,
            'kind': kind,
            'label': label,
            'raw': line.strip()[:260],
            'temperature': temperature,
            'power': power_payload,
        })
        if len(actions) >= max_actions:
            break
    actions.sort(key=lambda row: row['ts'])
    return actions

def time_index(rows):
    return [row['ts'] for row in (rows or []) if row.get('ts')]

def rows_in_time_window(rows, times, start_time, end_time):
    if not rows or not start_time or not end_time:
        return []
    if not times:
        return [row for row in rows if start_time <= row.get('ts') <= end_time]
    left = bisect_left(times, start_time)
    right = bisect_right(times, end_time)
    return rows[left:right]

def parse_android_cook_actions(action_rows, start_time, end_time, max_actions=180, action_times=None):
    actions = []
    last_offset = None
    for row in rows_in_time_window(action_rows or [], action_times, start_time, end_time):
        ts = row.get('ts')
        if not ts:
            continue
        offset = int((ts - start_time).total_seconds())
        action = {
            'time': ts.isoformat(sep=' '),
            'offset_seconds': offset,
            'delta_from_previous': None if last_offset is None else max(0, offset - last_offset),
            'kind': row.get('kind'),
            'label': row.get('label'),
            'raw': row.get('raw'),
            'temperature': row.get('temperature'),
            'power': row.get('power'),
        }
        actions.append(action)
        last_offset = offset
        if len(actions) >= max_actions:
            break
    return actions

def nearest_temperature_sample(series, target_time, max_delta_seconds=8, series_times=None):
    if not series:
        return None
    candidates = series
    if series_times:
        pos = bisect_left(series_times, target_time)
        indexes = [idx for idx in (pos - 1, pos, pos + 1) if 0 <= idx < len(series)]
        candidates = [series[idx] for idx in indexes]
    best = None
    best_delta = None
    for sample in candidates:
        delta = abs((sample['ts'] - target_time).total_seconds())
        if best_delta is None or delta < best_delta:
            best = sample
            best_delta = delta
            if delta == 0:
                break
    if best is None or best_delta is None or best_delta > max_delta_seconds:
        return None
    payload = {k: v for k, v in best.items() if k != 'ts'}
    payload['delta_seconds'] = round(best_delta, 1)
    return payload

def sample_source_summary(series):
    groups = {}
    for row in series or []:
        source = row.get('source') or 'unknown'
        item = groups.setdefault(source, {
            'source': source,
            'sample_count': 0,
            'first_time': None,
            'last_time': None,
        })
        item['sample_count'] += 1
        ts = row.get('ts')
        if ts:
            if not item['first_time'] or ts < item['_first_ts']:
                item['_first_ts'] = ts
                item['first_time'] = ts.isoformat(sep=' ')
            if not item['last_time'] or ts > item['_last_ts']:
                item['_last_ts'] = ts
                item['last_time'] = ts.isoformat(sep=' ')
    result = []
    for item in groups.values():
        item.pop('_first_ts', None)
        item.pop('_last_ts', None)
        result.append(item)
    return sorted(result, key=lambda x: x['sample_count'], reverse=True)

def recipe_step_action_label(step):
    type_map = {1: '人工投料', 2: '自动投料', 3: '机器控制', 4: '等待/时间', 5: '洗锅', 6: '润锅'}
    try:
        type_value = int(step.get('type') or 0)
    except Exception:
        type_value = 0
    commands = str(step.get('commands') or '').strip()
    return commands or type_map.get(type_value, f"步骤{type_value}")

def get_cook_window(cook_log):
    duration = int(cook_log.get('duration_seconds') or 0)
    cook_end = cook_log.get('create_time')
    cook_start = cook_end - timedelta(seconds=duration) if cook_end and duration else cook_log.get('end_time') or cook_end
    return cook_start, cook_end, duration

def compact_temperature_samples(series, cook_start, max_points=180):
    rows = []
    if not series or not cook_start:
        return rows
    stride = max(1, len(series) // max_points)
    picked = list(series[::stride])
    max_row = max(
        (row for row in series if row.get('pot_temp') is not None),
        key=lambda row: row.get('pot_temp'),
        default=None,
    )
    if max_row and all(row is not max_row for row in picked):
        picked.append(max_row)
    picked.sort(key=lambda row: row['ts'])
    for row in picked:
        item = {k: v for k, v in row.items() if k != 'ts'}
        item['offset_seconds'] = int((row['ts'] - cook_start).total_seconds())
        rows.append(item)
    return rows

def integrate_power_window(power_rows, start_time, end_time):
    if not start_time or not end_time or end_time <= start_time:
        return {
            'sample_count': 0,
            'actual_energy_kwh': 0,
            'command_energy_kwh': 0,
            'avg_actual_power_kw': None,
            'avg_command_power_kw': None,
            'max_actual_power_kw': None,
            'max_command_power_kw': None,
        }
    rows = [
        row for row in (power_rows or [])
        if row.get('ts') and row.get('power')
    ]
    rows.sort(key=lambda row: row['ts'])
    window_rows = [row for row in rows if start_time <= row['ts'] <= end_time]
    previous_row = None
    for row in rows:
        if row['ts'] <= start_time:
            previous_row = row
        else:
            break
    events = [row for row in rows if start_time < row['ts'] < end_time]
    timeline = []
    if previous_row:
        timeline.append({'ts': start_time, 'power': previous_row.get('power') or {}})
    elif window_rows:
        first = window_rows[0]
        if first['ts'] < end_time:
            timeline.append({'ts': max(first['ts'], start_time), 'power': first.get('power') or {}})
        events = [row for row in events if row is not first]
    for row in events:
        timeline.append({'ts': row['ts'], 'power': row.get('power') or {}})
    if not timeline:
        return {
            'sample_count': len(window_rows),
            'actual_energy_kwh': 0,
            'command_energy_kwh': 0,
            'avg_actual_power_kw': None,
            'avg_command_power_kw': None,
            'max_actual_power_kw': None,
            'max_command_power_kw': None,
        }
    actual_energy = 0.0
    command_energy = 0.0
    actual_seconds = 0.0
    command_seconds = 0.0
    actual_values = []
    command_values = []
    for row in window_rows:
        power = row.get('power') or {}
        actual = power.get('actual_power_kw')
        command = power.get('command_power_kw')
        if actual is not None:
            actual_values.append(float(actual))
        if command is not None:
            command_values.append(float(command))
    for idx, item in enumerate(timeline):
        seg_start = max(item['ts'], start_time)
        seg_end = min(timeline[idx + 1]['ts'] if idx + 1 < len(timeline) else end_time, end_time)
        dt = (seg_end - seg_start).total_seconds()
        if dt <= 0 or dt > 3600:
            continue
        power = item.get('power') or {}
        if power.get('actual_power_kw') is not None:
            actual_energy += float(power.get('actual_power_kw') or 0) * dt / 3600.0
            actual_seconds += dt
        if power.get('command_power_kw') is not None:
            command_energy += float(power.get('command_power_kw') or 0) * dt / 3600.0
            command_seconds += dt
    return {
        'sample_count': len(window_rows),
        'actual_energy_kwh': round(actual_energy, 5),
        'command_energy_kwh': round(command_energy, 5),
        'avg_actual_power_kw': round(actual_energy * 3600.0 / actual_seconds, 3) if actual_seconds else (round(statistics.mean(actual_values), 3) if actual_values else None),
        'avg_command_power_kw': round(command_energy * 3600.0 / command_seconds, 3) if command_seconds else (round(statistics.mean(command_values), 3) if command_values else None),
        'max_actual_power_kw': round(max(actual_values), 3) if actual_values else None,
        'max_command_power_kw': round(max(command_values), 3) if command_values else None,
    }

def build_power_segments(steps, power_rows, cook_start, duration):
    if not steps or not cook_start:
        return []
    ordered = sorted(steps, key=lambda row: (row.get('offset_seconds') or 0, row.get('step_index') or 0))
    segments = []
    for idx, step in enumerate(ordered):
        start_offset = max(0, int(step.get('offset_seconds') or 0))
        next_offset = int(ordered[idx + 1].get('offset_seconds') or duration or start_offset) if idx + 1 < len(ordered) else int(duration or start_offset)
        end_offset = max(start_offset, next_offset)
        start_time = cook_start + timedelta(seconds=start_offset)
        end_time = cook_start + timedelta(seconds=end_offset)
        metrics = integrate_power_window(power_rows, start_time, end_time)
        command_energy = metrics.get('command_energy_kwh') or 0
        actual_energy = metrics.get('actual_energy_kwh') or 0
        segments.append({
            'step_index': step.get('step_index'),
            'offset_seconds': start_offset,
            'end_offset_seconds': end_offset,
            'duration_seconds': max(0, end_offset - start_offset),
            'target_time': step.get('target_time'),
            'type': step.get('type'),
            'automatic': step.get('automatic'),
            'commands': step.get('commands'),
            'design_power_w': step.get('power'),
            'design_speed': step.get('speed'),
            'design_position': step.get('position'),
            **metrics,
            'follow_rate_percent': round(actual_energy / command_energy * 100, 1) if command_energy else None,
        })
    return segments

def build_single_cook_temperature(
    cook_log,
    temperature_series,
    main_board_text,
    detail_temperature_series=None,
    android_action_rows=None,
    temperature_times=None,
    detail_temperature_times=None,
    android_action_times=None,
):
    cook_start, cook_end, duration = get_cook_window(cook_log)
    if not cook_start or not cook_end:
        return None
    window_start = cook_start - timedelta(seconds=20)
    window_end = cook_end + timedelta(seconds=20)
    cook_series = rows_in_time_window(temperature_series, temperature_times, window_start, window_end)
    detail_series = rows_in_time_window(detail_temperature_series or temperature_series, detail_temperature_times or temperature_times, window_start, window_end)
    values = [row.get('pot_temp') for row in cook_series if row.get('pot_temp') is not None]
    if not values:
        values = [value for row in cook_series for value in (row.get('temp_1'), row.get('temp_2'), row.get('temp_3')) if value is not None]

    steps = []
    if cook_log.get('recipe_id'):
        detail = fetch_one(
            "SELECT recipe_id, cook_time, cook_steps FROM recipe_detail WHERE recipe_id = %s LIMIT 1",
            (cook_log.get('recipe_id'),),
            source=True,
            database='manage_backend',
        )
        raw_steps = parse_json_array(detail.get('cook_steps') if detail else None)
        for index, step in enumerate(raw_steps, start=1):
            offset = int(float(step.get('time') or 0))
            target = cook_start + timedelta(seconds=offset)
            steps.append({
                'step_index': index,
                'offset_seconds': offset,
                'delta_from_previous': None,
                'target_time': target.isoformat(sep=' '),
                'type': step.get('type'),
                'automatic': step.get('automatic'),
                'power': step.get('power'),
                'speed': step.get('speed'),
                'position': step.get('position'),
                'commands': recipe_step_action_label(step),
                'temperature': nearest_temperature_sample(temperature_series, target, series_times=temperature_times),
            })
    steps.sort(key=lambda row: (row.get('offset_seconds') or 0, row.get('step_index') or 0))
    last_offset = None
    for step in steps:
        offset = int(step.get('offset_seconds') or 0)
        step['delta_from_previous'] = None if last_offset is None else max(0, offset - last_offset)
        last_offset = offset

    main_actions = parse_main_board_actions(main_board_text, cook_start, cook_end) if main_board_text else []
    for action in main_actions[:80]:
        ts = datetime.strptime(action['time'], "%Y-%m-%d %H:%M:%S")
        action['temperature'] = nearest_temperature_sample(temperature_series, ts, series_times=temperature_times)
    main_actions.sort(key=lambda row: (row.get('offset_seconds') or 0, row.get('time') or ''))
    android_actions = parse_android_cook_actions(android_action_rows or [], cook_start, cook_end, action_times=android_action_times)

    cook_actions_full = rows_in_time_window(android_action_rows or [], android_action_times, cook_start, cook_end)
    power_samples_full = [row for row in cook_actions_full if row.get('kind') == 'power_sample' and row.get('power')]
    power_samples_full.sort(key=lambda r: r['ts'])
    power_metrics = integrate_power_window(power_samples_full, cook_start, cook_end)
    power_segments = build_power_segments(steps, power_samples_full, cook_start, duration)

    # 降采样功率事件，最多保留 200 个点用于图表绘制
    max_power_pts = 200
    power_samples_compact = []
    if power_samples_full:
        stride = max(1, len(power_samples_full) // max_power_pts)
        picked = list(power_samples_full[::stride])
        if picked[-1] is not power_samples_full[-1]:
            picked.append(power_samples_full[-1])
        for row in picked:
            offset = int((row['ts'] - cook_start).total_seconds())
            power_samples_compact.append({
                'offset_seconds': offset,
                'time': row['ts'].isoformat(sep=' '),
                'command_power_kw': row['power'].get('command_power_kw'),
                'actual_power_kw': row['power'].get('actual_power_kw'),
                'command_power_w': row['power'].get('command_power_w'),
                'actual_power_w': row['power'].get('actual_power_w'),
            })

    return {
        'cook': {
            **cook_log,
            'duration_seconds': duration,
            'start_time': cook_start.isoformat(sep=' ') if cook_start else None,
            'end_time_calc': cook_end.isoformat(sep=' ') if cook_end else None,
        },
        'summary': {
            'sample_count': len(cook_series),
            'min_temp': round(min(values), 1) if values else None,
            'max_temp': round(max(values), 1) if values else None,
            'avg_temp': round(statistics.mean(values), 1) if values else None,
            'start_temp': round(values[0], 1) if values else None,
            'end_temp': round(values[-1], 1) if values else None,
            'source_summary': sample_source_summary(cook_series),
            'actual_energy_kwh': round(power_metrics.get('actual_energy_kwh') or 0, 4),
            'command_energy_kwh': round(power_metrics.get('command_energy_kwh') or 0, 4),
            'avg_actual_power_kw': power_metrics.get('avg_actual_power_kw'),
            'avg_command_power_kw': power_metrics.get('avg_command_power_kw'),
            'max_actual_power_kw': power_metrics.get('max_actual_power_kw'),
            'max_command_power_kw': power_metrics.get('max_command_power_kw'),
            'power_sample_count': power_metrics.get('sample_count') or 0,
        },
        'steps': steps,
        'power_segments': power_segments,
        'android_actions': android_actions,
        'main_board_actions': main_actions[:80],
        'temperature_samples': compact_temperature_samples(detail_series, cook_start),
        'power_samples': power_samples_compact,
        'series': [],
    }

def build_cook_temperature_analysis(sn, file_id=None, force_refresh=False):
    real_sn, _ = resolve_sn(sn)
    if file_id:
        file_row = fetch_one(
            "SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted "
            "FROM machine_ftp WHERE id = %s AND sn = %s LIMIT 1",
            (file_id, real_sn),
            source=True,
            database='btyc',
        )
    else:
        file_row = fetch_one(
            "SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted "
            "FROM machine_ftp WHERE sn = %s AND cos_deleted = 0 ORDER BY create_time DESC LIMIT 1",
            (real_sn,),
            source=True,
            database='btyc',
        )
    if not file_row or not file_row.get('url'):
        raise HTTPException(status_code=404, detail="Log file not found")
    if file_row.get('cos_deleted'):
        raise HTTPException(status_code=410, detail="Log file has been deleted from COS")

    selected_file_id = int(file_row.get('id'))
    if not force_refresh:
        db_cached = read_cook_temperature_from_db(real_sn, selected_file_id)
        if db_cached:
            return db_cached
        cached = read_cached_cook_temperature(real_sn, selected_file_id)
        if cached:
            try:
                persist_cook_temperature_result(cached)
            except Exception as exc:
                print(f"persist cached cook temperature failed: {exc}")
            return cached

    mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='parsing', storage_status='not_stored')
    zip_bytes = download_log_zip(file_row['url'])
    try:
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='parse_failed', storage_status='not_stored', error_message='Downloaded log file is not a valid zip')
        raise HTTPException(status_code=400, detail="Downloaded log file is not a valid zip")

    android_temperature_series = []
    fallback_temperature_series = []
    main_board_text = ''
    android_action_text_parts = []
    parsed_android_files = []
    parsed_temperature_files = []
    max_total_temperature_points = 220000
    with zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = Path(info.filename).name
            if is_android_log_name(base):
                android_text = text_from_zip_member(zf, info, max_bytes=8 * 1024 * 1024)
                parsed_android_files.append(base)
                android_action_text_parts.append(android_text)
                if len(android_temperature_series) < max_total_temperature_points:
                    remaining = max_total_temperature_points - len(android_temperature_series)
                    android_temperature_series.extend(parse_android_pot_temperature_series(
                        android_text,
                        max_points=remaining,
                    ))
            elif is_temperature_log_name(base):
                parsed_temperature_files.append(base)
                fallback_temperature_series.extend(parse_temperature_series(
                    text_from_zip_member(zf, info, max_bytes=8 * 1024 * 1024),
                    max_points=max_total_temperature_points,
                ))
            elif base == 'main_board.log':
                main_board_text = text_from_zip_member(zf, info, max_bytes=10 * 1024 * 1024)
    android_temperature_series.sort(key=lambda row: row['ts'])
    fallback_temperature_series.sort(key=lambda row: row['ts'])
    # Prefer android business log pot temperature for summary/step alignment when present.
    # Use temperature.log as the detail stream when available because it is often denser.
    temperature_series = android_temperature_series or fallback_temperature_series
    detail_temperature_series = fallback_temperature_series or temperature_series
    android_action_rows = parse_android_action_rows('\n'.join(android_action_text_parts))
    temperature_series.sort(key=lambda row: row['ts'])
    detail_temperature_series.sort(key=lambda row: row['ts'])
    android_action_rows.sort(key=lambda row: row['ts'])
    if not temperature_series:
        mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='no_temperature_data', storage_status='not_stored', error_message='No usable temperature samples found')
        raise HTTPException(status_code=404, detail="No usable temperature samples found")
    temperature_times = time_index(temperature_series)
    detail_temperature_times = time_index(detail_temperature_series)
    android_action_times = time_index(android_action_rows)

    coverage_start = temperature_series[0]['ts']
    coverage_end = temperature_series[-1]['ts']
    latest_overall = fetch_one(
        "SELECT id, recipe_id, recipe_name, time AS duration_seconds, create_time, end_time, whether, manual "
        "FROM sop_machinelog WHERE sn = %s ORDER BY create_time DESC LIMIT 1",
        (real_sn,),
        source=True,
        database='btyc',
    )
    cook_logs = fetch_all(
        "SELECT id, recipe_id, recipe_name, time AS duration_seconds, create_time, end_time, whether, manual "
        "FROM sop_machinelog WHERE sn = %s AND recipe_id IS NOT NULL AND recipe_id != 0 "
        "AND create_time BETWEEN %s AND %s ORDER BY create_time ASC LIMIT 140",
        (real_sn, coverage_start, coverage_end),
        source=True,
        database='btyc',
    )
    if not cook_logs:
        no_match_detail = (
            "DB 未匹配到生产记录；这不代表日志内没有作业。"
            "请返回“日志包”点击“诊断”，查看日志内部烹饪、养锅、功率、温度和投料证据。"
            f" 日志覆盖：{coverage_start.isoformat(sep=' ')} ~ {coverage_end.isoformat(sep=' ')}"
        )
        mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='no_production_match', storage_status='not_stored', error_message=no_match_detail)
        raise HTTPException(
            status_code=404,
            detail=no_match_detail,
        )

    cooks = []
    for cook_log in cook_logs:
        item = build_single_cook_temperature(
            cook_log,
            temperature_series,
            main_board_text,
            detail_temperature_series,
            android_action_rows,
            temperature_times,
            detail_temperature_times,
            android_action_times,
        )
        if item:
            cooks.append(item)
    if not cooks:
        mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='parse_failed', storage_status='not_stored', error_message='No usable production window found in selected log coverage')
        raise HTTPException(status_code=404, detail="No usable production window found in selected log coverage")
    selected = cooks[-1]

    newer_uncovered = bool(latest_overall and latest_overall.get('create_time') and latest_overall.get('create_time') > coverage_end)
    result = {
        'sn': real_sn,
        'file': {
            'id': file_row.get('id'),
            'file_name': file_row.get('file_name'),
            'create_time': file_row.get('create_time'),
            'file_length': file_row.get('file_length'),
        },
        'coverage': {
            'start': coverage_start.isoformat(sep=' '),
            'end': coverage_end.isoformat(sep=' '),
            'sample_count': len(temperature_series),
            'android_sample_count': len(android_temperature_series),
            'temperature_log_sample_count': len(fallback_temperature_series),
            'android_file_count': len(parsed_android_files),
            'temperature_file_count': len(parsed_temperature_files),
            'android_files': parsed_android_files[:20],
            'temperature_files': parsed_temperature_files[:20],
            'temperature_unit': '主口径优先取 android 日志：功率=指令/实际输出功率，W 转 kW；温度=机芯/线盘/输出温度。temperature*.log：滤波/红外/输出温度，单位℃，用于更密采样和交叉验证',
            'newer_production_not_covered': newer_uncovered,
            'latest_production_time': latest_overall.get('create_time') if latest_overall else None,
            'source_summary': sample_source_summary(temperature_series),
        },
        'cook_count': len(cooks),
        'cooks': cooks,
        # Legacy fields keep older frontend code working while the page migrates to multi-cook.
        'cook': selected['cook'],
        'summary': selected['summary'],
        'steps': selected['steps'],
        'main_board_actions': selected['main_board_actions'],
        'series': selected['series'],
    }
    cache_key = save_cached_cook_temperature(real_sn, selected_file_id, result)
    try:
        persist_cook_temperature_result(result)
    except Exception as exc:
        mark_log_package_status(real_sn, selected_file_id, download_status='downloaded', parse_status='parsed', storage_status='store_failed', error_message=f'入库失败：{exc}', result=result)
        print(f"persist cook temperature failed: {exc}")
    result['cache'] = {
        'hit': False,
        'created_at': time.time(),
        'ttl_seconds': COOK_TEMPERATURE_CACHE_TTL_SECONDS,
        'cache_key': cache_key,
    }
    return result

def count_log_keywords(text, kind):
    definitions = {
        'mcu_debug': {
            '加热状态查询': 'getHeaterCurStatus',
            '通讯发包/MCU': 'send to mcu',
            '错误字样error': 'error',
            '超时timeout': 'timeout',
            '温控': 'tempctr',
            '加热错误码查询': 'heaterGetErrCode',
        },
        'main_board': {
            '主控加热': 'sta start',
            '加水': 'add water',
            '加油': 'add oil',
            '加蚝油': 'add oyster',
            '翻锅/倾锅': 'lean',
            '转锅速度': 'roll speed',
            '温控': 'tempctr',
            '无FrameID': 'no FrameID',
            '失败fail': 'fail',
        },
        'android_app': {
            '水淀粉/SAUCE_STARCH': 'SAUCE_STARCH',
            '液料/SAUCE_NEW': 'SAUCE_NEW',
            '网络请求': 'network request',
            '发送指令sendMsg': 'sendMsg',
            '读取结果readResult': 'readResult',
            '错误字样error': 'error',
            '数据采集DataCollect': 'DataCollectManager',
        },
    }
    counts = {}
    lower = text.lower()
    for label, needle in definitions.get(kind, {}).items():
        counts[label] = lower.count(needle.lower())
    return {k: v for k, v in counts.items() if v}

def top_log_patterns(text, limit=10):
    counter = Counter()
    for line in text.splitlines()[:80000]:
        msg = re.sub(r'^\[[^\]]+\]\s*', '', line.strip())
        msg = re.sub(r'0x[0-9a-fA-F]+', '0x#', msg)
        msg = re.sub(r'\d+', '#', msg)[:100]
        if msg:
            counter[msg] += 1
    return [{'pattern': pattern, 'count': count} for pattern, count in counter.most_common(limit)]

def build_log_analysis(file_row, zip_bytes):
    try:
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Downloaded log file is not a valid zip")

    file_summary = []
    keyword_counts = []
    top_patterns = []
    fault_matches = []
    oil = None
    temperature = None
    suggestions = []
    with zf:
        infos = [info for info in zf.infolist() if not info.is_dir()]
        for info in infos:
            base = Path(info.filename).name
            kind = classify_log_file(base)
            text = text_from_zip_member(zf, info)
            first = None
            last = None
            line_count = 0
            for line in text.splitlines():
                line_count += 1
                ts = parse_log_ts(line)
                if ts:
                    first = first or ts
                    last = ts
            file_summary.append({
                'file_name': base,
                'kind': kind,
                'kind_label': log_kind_label(kind),
                'size_mb': round(info.file_size / 1024 / 1024, 2),
                'lines': line_count,
                'first_time': first.isoformat(sep=' ') if first else None,
                'last_time': last.isoformat(sep=' ') if last else None,
            })
            fault_matches.extend(extract_fault_matches(text, base, kind))
            if kind == 'other':
                continue
            if kind == 'oildrum_board':
                oil = analyze_oildrum_text(text)
            elif kind == 'temperature':
                temperature = analyze_temperature_text(text)
            counts = count_log_keywords(text, kind)
            for label, count in counts.items():
                keyword_counts.append({'file_name': base, 'kind': kind, 'category': label, 'count': count})
            top_patterns.extend({'file_name': base, 'kind': kind, **row} for row in top_log_patterns(text, limit=6))

    if oil and oil.get('bucket_over_100_count', 0):
        suggestions.append('油桶板日志存在桶温超过 100℃ 的采样，建议重点核对桶温字段定义、加热目标值和是否存在局部过冲。')
    if oil and oil.get('pipe_max') and oil['pipe_max'] >= 60:
        suggestions.append('油路/管路推定温度达到 60℃ 左右，建议与管材长期耐温、泵附近抖动和加热丝接触位置一起评估。')
    if any(row['category'] == '水淀粉/SAUCE_STARCH' for row in keyword_counts):
        suggestions.append('日志中可识别水淀粉资源动作，后续可以把它纳入设备资源筛查，而不只看猪油/油桶。')
    if oil and oil.get('hall_abnormal_count'):
        suggestions.append('猪油桶日志存在霍尔缺相检测异常记录，后续可以结合投油动作、持续时间和电机占空比进一步定位。')
    if fault_matches:
        suggestions.append('日志文本中命中了故障码清单，可先按优先级、模块和命中次数查看是否与现场问题一致。')
    if not suggestions:
        suggestions.append('该日志包已完成结构化解析，建议先从文件摘要、动作计数和温度统计确认哪些字段对排障最有价值。')

    return {
        'file': {
            'id': file_row.get('id'),
            'sn': file_row.get('sn'),
            'file_name': file_row.get('file_name'),
            'file_length': file_row.get('file_length'),
            'create_time': file_row.get('create_time'),
            'update_time': file_row.get('update_time'),
        },
        'analysis_version': LOG_ANALYSIS_VERSION,
        'generated_at': int(time.time()),
        'zip_size_mb': round(len(zip_bytes) / 1024 / 1024, 2),
        'file_summary': sorted(file_summary, key=lambda row: row['file_name']),
        'category_summary': build_log_category_summary(file_summary),
        'keyword_counts': sorted(keyword_counts, key=lambda row: row['count'], reverse=True)[:120],
        'top_patterns': top_patterns[:160],
        'fault_matches': sorted(fault_matches, key=lambda row: (row.get('priority') or '', -row.get('count', 0)))[:160],
        'oil_thermal': oil,
        'temperature': temperature,
        'suggestions': suggestions,
        'log_specs': log_specs_payload(),
        'field_assumptions': {
            'oildrum_bucket_temp': 'oildrum_board.log dev status 第 1 个数 / 10，按《猪油桶相关日志说明》确认为油温度',
            'oildrum_pipe_temp': 'oildrum_board.log dev status 第 2 个数 / 10，按《猪油桶相关日志说明》确认为油管温度',
            'oildrum_motor_rpm': 'oildrum_board.log dev status 第 4 个数为投料电机速度 RPM',
            'oildrum_heater_pwm': 'oildrum_board.log dev status 第 5 个数为油桶加热 PWM 占空比',
        },
    }

def log_package_diagnostics_cache_path(file_id):
    path = CACHE_DIR / 'log_package_diagnostics'
    path.mkdir(parents=True, exist_ok=True)
    return path / f'{int(file_id)}.json'

def read_cached_log_package_diagnostics(file_id):
    path = log_package_diagnostics_cache_path(file_id)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding='utf-8'))
        if payload.get('diagnostics_version') != LOG_PACKAGE_DIAGNOSTICS_VERSION:
            return None
        payload['cache'] = {'hit': True}
        return payload
    except Exception:
        path.unlink(missing_ok=True)
        return None

def save_cached_log_package_diagnostics(file_id, payload):
    path = log_package_diagnostics_cache_path(file_id)
    serializable = dict(payload)
    serializable.pop('cache', None)
    path.write_text(json.dumps(serializable, ensure_ascii=False, default=str), encoding='utf-8')

def expected_log_package_bytes(file_length):
    size_mb = parse_mb_value(file_length)
    if size_mb is None or size_mb <= 0:
        return None
    return int(size_mb * 1024 * 1024)

def package_size_match_status(downloaded_size, expected_size):
    if not expected_size:
        return 'unknown'
    ratio = abs(downloaded_size - expected_size) / expected_size
    return 'matched' if ratio <= 0.15 else 'mismatch'

def parse_internal_log_event(line, source_file, source_kind, line_no):
    raw = str(line or '').strip()
    ts = parse_log_ts(raw)
    if not ts:
        return None
    lower = raw.lower()
    event_type = None
    event_label = None
    session_hint = None
    recipe_id = None
    recipe_name = None
    command_power_w = None
    actual_power_w = None
    temperatures = None

    cook_match = re.search(r'烹饪开始[：:]?\s*(.+)$', raw)
    if cook_match:
        event_type, event_label, session_hint = 'cook_start', '烹饪开始', 'cooking'
        recipe_name = str(cook_match.group(1) or '').strip(' _:-')
        trailing = re.search(r'^(.+?)_(\d{4,})(?:\D*)$', recipe_name)
        if trailing:
            recipe_name = trailing.group(1).strip(' _:-')
            recipe_id = int(trailing.group(2))
    elif re.search(r'快速养锅|养锅开始|protect.?pot', raw, re.I):
        event_type, event_label, session_hint = 'protect_pot_start', '快速养锅', 'protect_pot'
    elif re.search(r'录制菜谱', raw, re.I):
        event_type, event_label, session_hint = 'recipe_recording_start', '录制菜谱', 'recipe_recording'
    elif re.search(r'RecipeMessageActivity\s+onCreate', raw, re.I):
        event_type, event_label = 'recipe_activity', '菜谱消息页面'
    elif re.search(r'NewCookingActivity\s+onCreate|updateRobotScene\s*:\s*scene\s*=\s*COOKING', raw, re.I):
        event_type, event_label, session_hint = 'cooking_scene', '进入烹饪场景', 'cooking'
    elif re.search(r'烹饪.*结束|COOK_FINISH|COOKING_FINISH', raw, re.I):
        event_type, event_label = 'cook_end', '烹饪结束'
    elif re.search(r'开始检测温度', raw):
        event_type, event_label = 'temperature_detection_start', '开始检测温度'
    elif re.search(r'开始检测输出功率', raw):
        event_type, event_label = 'power_detection_start', '开始检测输出功率'
    elif re.search(r'开始投液料', raw):
        event_type, event_label = 'liquid_material_start', '开始投液料'
    elif re.search(r'投液料_|消耗液料|液料投料记录', raw):
        event_type, event_label = 'liquid_material', '液料投放'
    elif re.search(r'开始称重', raw):
        event_type, event_label = 'weighing_start', '开始称重'
    elif re.search(r'倾锅操作|lean', raw, re.I):
        event_type, event_label = 'tilt_pot', '倾锅操作'
    elif re.search(r'转锅操作|开始转锅|roll mov|roll start|roll stop', raw, re.I):
        event_type, event_label = 'rotate_pot', '转锅操作'
    elif 'datacollectmanager' in lower:
        event_type, event_label = 'data_collect', '数据采集'
    elif re.search(r'CNEngine\s+speak', raw, re.I):
        event_type, event_label = 'voice_prompt', '语音/人工操作提示'

    power_sample = re.search(r'功率\s*:\s*(-?\d+(?:\.\d+)?)_(-?\d+(?:\.\d+)?)', raw)
    power_set = re.search(r'功率设置(?:为|_)?\s*[：:]?\s*(-?\d+(?:\.\d+)?)\s*W?', raw, re.I)
    if power_sample:
        command_power_w = float(power_sample.group(1))
        actual_power_w = float(power_sample.group(2))
        event_type = event_type or 'power_sample'
        event_label = event_label or '功率采样'
    elif power_set:
        command_power_w = float(power_set.group(1))
        event_type = event_type or 'power_set'
        event_label = event_label or '功率设置'

    temp_match = re.search(r'温度\s*:_?\s*(-?\d+)_(-?\d+)_(-?\d+)', raw)
    if not temp_match and source_kind == 'temperature':
        temp_match = re.search(r'\]\s*(-?\d+)_(-?\d+)_(-?\d+)', raw)
    if temp_match:
        raw_values = [int(temp_match.group(i)) for i in (1, 2, 3)]
        if source_kind == 'temperature':
            temperatures = {
                'filtered_temp': raw_values[0],
                'infrared_temp': raw_values[1],
                'pot_temp': raw_values[2],
                'raw_triplet': '_'.join(map(str, raw_values)),
            }
        else:
            temperatures = {
                'core_temp': raw_values[0],
                'coil_temp': raw_values[1],
                'pot_temp': raw_values[2],
                'raw_triplet': '_'.join(map(str, raw_values)),
            }
        event_type = event_type or 'temperature_sample'
        event_label = event_label or '温度采样'

    if not event_type:
        return None
    return {
        'ts': ts,
        'time': ts.isoformat(sep=' '),
        'event_type': event_type,
        'event_label': event_label,
        'session_hint': session_hint,
        'recipe_id': recipe_id,
        'recipe_name': recipe_name,
        'command_power_w': command_power_w,
        'actual_power_w': actual_power_w,
        'temperatures': temperatures,
        'source_file': source_file,
        'source_kind': source_kind,
        'line_no': line_no,
        'raw': raw[:1000],
    }

def internal_session_summary(events, index):
    start = events[0]['ts']
    end = events[-1]['ts']
    session_type = next((row.get('session_hint') for row in events if row.get('session_hint')), None) or 'manual_action'
    recipe_id = next((row.get('recipe_id') for row in events if row.get('recipe_id')), None)
    recipe_name = next((row.get('recipe_name') for row in events if row.get('recipe_name')), None)
    power_values = []
    temperature_values = []
    material_count = 0
    evidence = []
    for row in events:
        for value in (row.get('command_power_w'), row.get('actual_power_w')):
            if value is not None:
                power_values.append(float(value))
        temp = row.get('temperatures') or {}
        if temp.get('pot_temp') is not None:
            temperature_values.append(float(temp['pot_temp']))
        if row.get('event_type') in ('liquid_material_start', 'liquid_material', 'weighing_start'):
            material_count += 1
        if len(evidence) < 80:
            evidence.append({
                'time': row.get('time'),
                'event_type': row.get('event_type'),
                'event_label': row.get('event_label'),
                'source_file': row.get('source_file'),
                'line_no': row.get('line_no'),
                'raw': row.get('raw'),
                'command_power_w': row.get('command_power_w'),
                'actual_power_w': row.get('actual_power_w'),
                'temperatures': row.get('temperatures'),
            })
    return {
        'session_id': f'internal-{index}-{int(start.timestamp())}',
        'session_type': session_type if session_type in ('cooking', 'protect_pot', 'recipe_recording', 'manual_action') else 'unknown',
        'start_time': start.isoformat(sep=' '),
        'end_time': end.isoformat(sep=' '),
        'duration_seconds': max(0, int((end - start).total_seconds())),
        'recipe_id': recipe_id,
        'recipe_name': recipe_name,
        'max_power_w': round(max(power_values), 1) if power_values else None,
        'max_pot_temp': round(max(temperature_values), 1) if temperature_values else None,
        'avg_pot_temp': round(statistics.mean(temperature_values), 1) if temperature_values else None,
        'first_pot_temp': round(temperature_values[0], 1) if temperature_values else None,
        'last_pot_temp': round(temperature_values[-1], 1) if temperature_values else None,
        'event_count': len(events),
        'temperature_sample_count': len(temperature_values),
        'power_event_count': sum(1 for row in events if row.get('command_power_w') is not None or row.get('actual_power_w') is not None),
        'material_event_count': material_count,
        'evidence_lines': evidence,
    }

def build_internal_sessions(events):
    deduped = {}
    for row in events or []:
        key = (
            row.get('ts'),
            row.get('event_type'),
            re.sub(r'\s+', ' ', str(row.get('raw') or '')),
        )
        deduped[key] = row
    rows = sorted(deduped.values(), key=lambda row: (row['ts'], row.get('source_file') or '', row.get('line_no') or 0))
    if not rows:
        return []
    sessions = []
    current = []
    current_hint = None
    strong_start_types = {'cook_start', 'protect_pot_start', 'recipe_recording_start', 'cooking_scene'}
    for row in rows:
        gap = (row['ts'] - current[-1]['ts']).total_seconds() if current else 0
        hint = row.get('session_hint')
        begins_new = row.get('event_type') in strong_start_types
        if current and (
            gap > 900
            or (begins_new and hint and current_hint and hint != current_hint and current_hint != 'cooking')
            or (row.get('event_type') == 'cook_start' and current_hint == 'cooking' and gap > 120)
        ):
            sessions.append(internal_session_summary(current, len(sessions) + 1))
            current = []
            current_hint = None
        current.append(row)
        current_hint = current_hint or hint
        if row.get('event_type') == 'cook_end':
            sessions.append(internal_session_summary(current, len(sessions) + 1))
            current = []
            current_hint = None
    if current:
        sessions.append(internal_session_summary(current, len(sessions) + 1))
    return sessions[:120]

def compact_internal_session(session):
    return {key: value for key, value in session.items() if key != 'evidence_lines'}

def build_target_job_match_diagnosis(target_job_time, coverage_start, coverage_end, db_rows, internal_sessions, detected_kinds):
    target = to_mysql_dt(target_job_time)
    if not target:
        return None
    before = []
    after = []
    for session in internal_sessions:
        start = to_mysql_dt(session.get('start_time'))
        end = to_mysql_dt(session.get('end_time')) or start
        if not start:
            continue
        if start <= target <= end:
            before.append((0, session))
            after.append((0, session))
        elif end and end < target:
            before.append(((target - end).total_seconds(), session))
        else:
            after.append(((start - target).total_seconds(), session))
    before.sort(key=lambda item: item[0])
    after.sort(key=lambda item: item[0])
    db_match = None
    for row in db_rows:
        start = to_mysql_dt(row.get('create_time'))
        end = to_mysql_dt(row.get('end_time')) or start
        if start and end and start - timedelta(minutes=2) <= target <= end + timedelta(minutes=2):
            db_match = row
            break
    nearest = (before[:1] + after[:1])
    nearest_delta = min([item[0] for item in nearest], default=None)
    if coverage_start and coverage_end and not (coverage_start <= target <= coverage_end):
        failure_reason = 'log_time_not_cover_job'
        suggested = '更换覆盖目标作业时间的日志包，或先核对设备与日志时钟。'
    elif db_match:
        failure_reason = None
        suggested = '已匹配数据库生产记录，可继续查看温度和功率证据。'
    elif internal_sessions and nearest_delta is not None and nearest_delta <= 300:
        failure_reason = 'internal_session_exists_but_no_db_record'
        suggested = '优先查看日志内作业证据；同时核对数据库漏记、时区或设备时钟漂移。'
    elif internal_sessions:
        failure_reason = 'timezone_or_clock_drift_suspected'
        suggested = '日志内存在作业但时间偏差较大，建议核对设备时区和日志/数据库时钟。'
    elif 'temperature' not in detected_kinds and 'android_app' not in detected_kinds:
        failure_reason = 'no_temperature_signal'
        suggested = '当前包缺少温度信号，请换更完整日志包。'
    else:
        failure_reason = 'no_internal_session'
        suggested = '当前日志覆盖内未重建出有效作业片段，请查看关键文件和原始日志。'
    return {
        'target_job_time': target.isoformat(sep=' '),
        'searched_file_ids': [],
        'db_match_status': 'matched' if db_match else 'not_matched',
        'internal_session_match_status': 'matched' if internal_sessions and nearest_delta is not None and nearest_delta <= 300 else 'not_matched',
        'nearest_internal_session_before': compact_internal_session(before[0][1]) if before else None,
        'nearest_internal_session_after': compact_internal_session(after[0][1]) if after else None,
        'log_time_start': coverage_start.isoformat(sep=' ') if coverage_start else None,
        'log_time_end': coverage_end.isoformat(sep=' ') if coverage_end else None,
        'failure_reason': failure_reason,
        'suggested_next_action': suggested,
    }

def build_log_package_diagnostics(file_row, target_job_time=None):
    result = {
        'diagnostics_version': LOG_PACKAGE_DIAGNOSTICS_VERSION,
        'generated_at': int(time.time()),
        'file_id': int(file_row.get('id')),
        'sn': file_row.get('sn'),
        'file_name': file_row.get('file_name'),
        'file_length': file_row.get('file_length'),
        'create_time': file_row.get('create_time'),
        'update_time': file_row.get('update_time'),
        'cos_deleted': bool(file_row.get('cos_deleted')),
        'has_cos_url': bool(file_row.get('url')),
        'download_status': 'not_started',
        'http_status': None,
        'downloaded_size': None,
        'expected_size': expected_log_package_bytes(file_row.get('file_length')),
        'size_match_status': 'unknown',
        'is_zip': False,
        'zip_open_status': 'not_started',
        'zip_error': None,
        'file_list_count': 0,
        'detected_files': [],
        'detected_log_kinds': [],
        'has_android_log': False,
        'has_temperature_log': False,
        'has_main_board_log': False,
        'has_oildrum_board_log': False,
        'has_debug_log': False,
        'log_time_start': None,
        'log_time_end': None,
        'db_production_match_count': 0,
        'internal_sessions': [],
        'diagnosis_level': 'error',
        'diagnosis_code': None,
        'diagnosis_message': '',
        'suggested_action': '',
        'match_diagnosis': None,
    }
    if not file_row.get('url'):
        result.update(diagnosis_code='cos_url_missing', diagnosis_message='日志索引存在，但没有 COS 下载地址。', suggested_action='联系设备侧重新上传日志包。')
        return result
    if file_row.get('cos_deleted'):
        result.update(download_status='remote_deleted', diagnosis_code='cos_deleted', diagnosis_message='COS 文件已删除。', suggested_action='联系设备侧重新上传，或查找相邻时间日志包。')
        return result
    try:
        zip_bytes = download_log_zip(file_row['url'])
        result['download_status'] = 'downloaded'
        result['http_status'] = 200
        result['downloaded_size'] = len(zip_bytes)
        result['size_match_status'] = package_size_match_status(len(zip_bytes), result['expected_size'])
    except HTTPException as exc:
        code = 'file_too_large' if exc.status_code == 413 else 'download_failed'
        result.update(download_status='failed', http_status=exc.status_code, diagnosis_code=code, diagnosis_message=str(exc.detail), suggested_action='日志包超过保护上限时请先缩小范围；其他下载错误可稍后重试。')
        return result
    except TimeoutError as exc:
        result.update(download_status='failed', diagnosis_code='download_timeout', diagnosis_message=str(exc), suggested_action='稍后重试或检查 COS 网络状态。')
        return result
    except urllib.error.HTTPError as exc:
        result.update(download_status='failed', http_status=exc.code, diagnosis_code='download_failed', diagnosis_message=f'COS HTTP {exc.code}', suggested_action='检查 COS 地址是否过期或文件是否已删除。')
        return result
    except Exception as exc:
        result.update(download_status='failed', diagnosis_code='download_failed', diagnosis_message=str(exc)[:300], suggested_action='稍后重试并检查 COS 地址。')
        return result
    result['is_zip'] = zipfile.is_zipfile(BytesIO(zip_bytes))
    if not result['is_zip']:
        result.update(zip_open_status='failed', diagnosis_code='not_zip', diagnosis_message='下载内容不是 ZIP 文件。', suggested_action='核对 machine_ftp 文件地址和上传内容。')
        return result
    try:
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
        bad_member = zf.testzip()
        if bad_member:
            raise zipfile.BadZipFile(f'CRC error: {bad_member}')
        result['zip_open_status'] = 'opened'
    except zipfile.BadZipFile as exc:
        result.update(zip_open_status='failed', zip_error=str(exc), diagnosis_code='zip_corrupted', diagnosis_message='ZIP 文件损坏或 CRC 校验失败。', suggested_action='重新上传日志包或查看相邻包。')
        return result

    internal_events = []
    first_time = None
    last_time = None
    detected_kinds = set()
    with zf:
        infos = [info for info in zf.infolist() if not info.is_dir()]
        result['file_list_count'] = len(infos)
        for info in infos:
            base = Path(info.filename).name
            kind = classify_log_file(base)
            if kind != 'other':
                detected_kinds.add(kind)
            detected = {
                'file_name': info.filename,
                'base_name': base,
                'kind': kind,
                'kind_label': log_kind_label(kind),
                'size': info.file_size,
                'first_time': None,
                'last_time': None,
                'event_count': 0,
            }
            if kind != 'other':
                try:
                    text = text_from_zip_member(zf, info, max_bytes=32 * 1024 * 1024)
                    file_first = None
                    file_last = None
                    event_count = 0
                    for line_no, line in enumerate(text.splitlines(), start=1):
                        ts = parse_log_ts(line)
                        if ts:
                            file_first = file_first or ts
                            file_last = ts
                            first_time = ts if first_time is None or ts < first_time else first_time
                            last_time = ts if last_time is None or ts > last_time else last_time
                        event = parse_internal_log_event(line, info.filename, kind, line_no)
                        if event:
                            internal_events.append(event)
                            event_count += 1
                    detected.update(
                        first_time=file_first.isoformat(sep=' ') if file_first else None,
                        last_time=file_last.isoformat(sep=' ') if file_last else None,
                        event_count=event_count,
                    )
                except Exception as exc:
                    detected['parse_error'] = str(exc)[:200]
            result['detected_files'].append(detected)

    result['detected_log_kinds'] = sorted(detected_kinds)
    result['has_android_log'] = 'android_app' in detected_kinds
    result['has_temperature_log'] = 'temperature' in detected_kinds
    result['has_main_board_log'] = 'main_board' in detected_kinds
    result['has_oildrum_board_log'] = 'oildrum_board' in detected_kinds
    result['has_debug_log'] = 'mcu_debug' in detected_kinds
    result['log_time_start'] = first_time.isoformat(sep=' ') if first_time else None
    result['log_time_end'] = last_time.isoformat(sep=' ') if last_time else None
    sessions = build_internal_sessions(internal_events)
    result['internal_sessions'] = sessions

    db_rows = []
    if first_time and last_time:
        db_rows = fetch_all(
            "SELECT id, recipe_id, recipe_name, create_time, end_time, time AS duration_seconds "
            "FROM sop_machinelog WHERE sn=%s AND create_time BETWEEN %s AND %s "
            "ORDER BY create_time ASC LIMIT 140",
            (file_row.get('sn'), first_time, last_time),
            source=True,
            database='btyc',
        )
    result['db_production_match_count'] = len(db_rows)
    if target_job_time:
        result['match_diagnosis'] = build_target_job_match_diagnosis(
            target_job_time, first_time, last_time, db_rows, sessions, detected_kinds,
        )
        if result['match_diagnosis']:
            result['match_diagnosis']['searched_file_ids'] = [int(file_row.get('id'))]

    cooking_sessions = [row for row in sessions if row.get('session_type') in ('cooking', 'recipe_recording')]
    manual_sessions = [row for row in sessions if row.get('session_type') in ('protect_pot', 'manual_action')]
    if not detected_kinds:
        result.update(diagnosis_code='key_log_missing', diagnosis_message='ZIP 可打开，但没有识别到关键日志文件。', suggested_action='检查设备日志版本和文件命名规则。')
    elif not first_time:
        result.update(diagnosis_code='log_time_not_found', diagnosis_message='关键日志存在，但无法识别日志时间范围。', suggested_action='查看原始时间戳格式并补充解析规则。')
    elif db_rows:
        result.update(diagnosis_level='ok', diagnosis_code='db_matched', diagnosis_message='已匹配生产记录。', suggested_action='可继续进入温度、功率或日志结论分析。')
    elif cooking_sessions:
        result.update(diagnosis_level='warning', diagnosis_code='internal_work_found', diagnosis_message='DB 未匹配，日志内有作业片段。', suggested_action='查看日志内作业证据，并核对数据库漏记、时区或设备时钟。')
    elif manual_sessions:
        result.update(diagnosis_level='warning', diagnosis_code='internal_manual_found', diagnosis_message='DB 未匹配，日志内仅有养锅/手动动作。', suggested_action='查看内部动作证据，确认是否属于非标准烹饪或人工操作。')
    else:
        result.update(diagnosis_level='warning', diagnosis_code='no_internal_session', diagnosis_message='DB 未匹配，日志内无有效作业。', suggested_action='检查关键日志是否完整，或选择相邻时间日志包。')
    return result

def build_device_report(sn: str):
    real_sn, info = resolve_sn(sn)
    customer = get_company_info(info.get('company_id'))
    software_info = get_device_software_info(real_sn)

    log_summary = fetch_one(
        "SELECT COUNT(*) AS total_logs, MIN(create_time) AS first_time, MAX(create_time) AS last_time "
        "FROM sop_machinelog WHERE sn = %s",
        (real_sn,),
        source=True,
        database='btyc',
    )
    total_logs = int(log_summary['total_logs'] or 0)
    monthly_summary = fetch_all(
        "SELECT DATE_FORMAT(create_time, '%%Y-%%m') AS month, COUNT(*) AS total_logs, "
        "COUNT(DISTINCT recipe_id) AS recipe_count, MIN(create_time) AS first_time, "
        "MAX(create_time) AS last_time, SUM(CAST(time AS SIGNED)) AS total_duration_seconds, "
        "ROUND(AVG(CAST(time AS SIGNED)), 1) AS avg_duration_seconds "
        "FROM sop_machinelog WHERE sn = %s GROUP BY month ORDER BY month",
        (real_sn,),
        source=True,
        database='btyc',
    )
    if total_logs > MAX_DEVICE_LOGS:
        logs = fetch_all(
            "SELECT id, recipe_id, recipe_name, time AS duration_seconds, create_time, end_time, "
            "username, data_time, comment, whether, manual, component, uuid "
            "FROM sop_machinelog WHERE sn = %s ORDER BY create_time DESC LIMIT %s",
            (real_sn, MAX_DEVICE_LOGS),
            source=True,
            database='btyc',
        )
        logs = list(reversed(logs))
    else:
        logs = fetch_all(
            "SELECT id, recipe_id, recipe_name, time AS duration_seconds, create_time, end_time, "
            "username, data_time, comment, whether, manual, component, uuid "
            "FROM sop_machinelog WHERE sn = %s ORDER BY create_time ASC",
            (real_sn,),
            source=True,
            database='btyc',
        )

    duration_stat_rows = fetch_all(
        "SELECT recipe_id, COUNT(*) AS sample_count, "
        "ROUND(AVG(CAST(time AS SIGNED)), 1) AS avg_duration_seconds, "
        "MIN(CAST(time AS SIGNED)) AS min_duration_seconds, "
        "MAX(CAST(time AS SIGNED)) AS max_duration_seconds "
        "FROM sop_machinelog WHERE sn = %s AND recipe_id IS NOT NULL AND recipe_id != 0 "
        "AND time IS NOT NULL GROUP BY recipe_id",
        (real_sn,),
        source=True,
        database='btyc',
    )
    duration_stats = {row['recipe_id']: row for row in duration_stat_rows}
    for log in logs:
        log.update(classify_production_behavior(log, duration_stats))

    intervals = []
    last_time = None
    for log in logs:
        curr = log['create_time']
        gap = None
        if last_time:
            delta = curr - last_time
            gap_hours = delta.total_seconds() / 3600
            gap = {
                'hours': round(gap_hours, 2),
                'days': delta.days,
                'label': format_gap(delta),
                'flag': 'LONG_DELAY' if gap_hours > 24 else 'NORMAL'
            }
        intervals.append({
            'id': log['id'],
            'create_time': curr,
            'recipe_name': log.get('recipe_name'),
            'gap': gap
        })
        last_time = curr

    recipe_usage_rows = fetch_all(
        "SELECT recipe_id, MAX(recipe_name) AS recipe_name, COUNT(*) AS cnt, "
        "SUM(CAST(time AS SIGNED)) AS total_duration_seconds, MIN(create_time) AS first_time, "
        "MAX(create_time) AS last_time FROM sop_machinelog "
        "WHERE sn = %s AND recipe_id IS NOT NULL AND recipe_id != 0 "
        "GROUP BY recipe_id ORDER BY cnt DESC",
        (real_sn,),
        source=True,
        database='btyc',
    )
    recipe_usage = {row['recipe_id']: row for row in recipe_usage_rows}
    recipe_ids = sorted(recipe_usage.keys())
    recipes = []
    recipe_archive = []
    has_lard = False
    if recipe_ids:
        placeholders = ','.join(['%s'] * len(recipe_ids))
        raw_recipes = fetch_all(
            f"SELECT id, name, group_name, type, steps_describe, ingredients_total_dosage FROM main_recipe WHERE id IN ({placeholders})",
            tuple(recipe_ids),
            source=True,
            database='manage_backend',
        )
        raw_details = fetch_all(
            f"SELECT recipe_id, cook_time, cook_steps, wash_steps, moisten_steps, cooking_ingredient "
            f"FROM recipe_detail WHERE recipe_id IN ({placeholders})",
            tuple(recipe_ids),
            source=True,
            database='manage_backend',
        )
        detail_by_recipe_id = {row['recipe_id']: row for row in raw_details}
        for r in raw_recipes:
            archive_item = recipe_resource_summary(r, detail_by_recipe_id.get(r['id']), recipe_usage)
            is_lard = archive_item['has_lard'] or '猪油' in str(r.get('name', '')) or '猪油' in str(r.get('steps_describe', ''))
            if is_lard: has_lard = True
            recipes.append({**r, 'has_lard': is_lard, 'execution_count': recipe_usage.get(r['id'], {}).get('cnt', 0)})
            recipe_archive.append(archive_item)

        found_ids = {r['id'] for r in raw_recipes}
        for missing_id in sorted(set(recipe_ids) - found_ids):
            usage = recipe_usage.get(missing_id, {})
            recipe_archive.append({
                'id': missing_id,
                'name': usage.get('recipe_name') or f'菜谱 {missing_id}',
                'category': '菜谱详情缺失',
                'execution_count': usage.get('cnt', 0),
                'total_duration_seconds': usage.get('total_duration_seconds', 0),
                'first_time': usage.get('first_time'),
                'last_time': usage.get('last_time'),
                'resources': [],
                'resource_flags': {},
                'steps': [],
                'has_lard': False,
            })

    faults = fetch_all(
        "SELECT id, create_time, module, second_level_error_details as details, deal_state "
        "FROM bytc_robot_malfunctions_log WHERE sn = %s ORDER BY create_time DESC LIMIT 200",
        (real_sn,),
        source=True,
        database='btyc',
    )

    maint = fetch_all(
        "SELECT id, create_time, status, mode, duration FROM robot_conservation_pot_log "
        "WHERE machine_code = %s ORDER BY create_time DESC LIMIT 200",
        (real_sn,),
        source=True,
        database='btyc',
    )
    device_logs = get_device_log_files(real_sn)
    structured_summary = device_structured_summary(real_sn)

    recipe_archive = sorted(recipe_archive, key=lambda item: item.get('execution_count', 0), reverse=True)

    return {
        "info": info,
        "software": software_info,
        "customer": customer,
        "stats": {
            "sn": real_sn,
            "total_logs": total_logs,
            "returned_logs": len(logs),
            "first_time": log_summary['first_time'],
            "last_time": log_summary['last_time'],
            "truncated": total_logs > MAX_DEVICE_LOGS,
            "has_lard": has_lard,
        },
        "monthly_summary": monthly_summary,
        "logs": logs if len(logs) <= INLINE_LOG_LIMIT else logs[-INLINE_LOG_LIMIT:],
        "intervals": intervals if len(intervals) <= INLINE_LOG_LIMIT else intervals[-INLINE_LOG_LIMIT:],
        "recipes": recipes,
        "recipe_archive": recipe_archive,
        "recipe_category_summary": build_recipe_category_summary(recipe_archive),
        "faults": faults,
        "maintenance": maint,
        "device_logs": device_logs,
        "structured_summary": structured_summary,
    }

def get_cached_report(sn: str, force_refresh: bool = False):
    normalized = sn.strip()
    cached = REPORT_CACHE.get(normalized)
    now = time.time()
    if not force_refresh and cached and now - cached['created_at'] < CACHE_TTL_SECONDS:
        report = cached['report']
        report['cache'] = {'hit': True, 'created_at': cached['created_at'], 'ttl_seconds': CACHE_TTL_SECONDS}
        return report

    if not force_refresh:
        disk_report = load_disk_report(normalized, now)
        if disk_report:
            real_sn = disk_report['stats']['sn']
            for key in set(candidate_sns(normalized) + [real_sn]):
                REPORT_CACHE[key] = {'created_at': disk_report['cache']['created_at'], 'report': disk_report}
            return disk_report

    report = build_device_report(normalized)
    real_sn = report['stats']['sn']
    for key in set(candidate_sns(normalized) + [real_sn]):
        REPORT_CACHE[key] = {'created_at': now, 'report': report}
    report['cache'] = {'hit': False, 'created_at': now, 'ttl_seconds': CACHE_TTL_SECONDS}
    save_disk_report(real_sn, report, now)
    return report

def cache_path(sn):
    digest = hashlib.sha256(sn.encode()).hexdigest()
    return CACHE_DIR / f"{digest}.json"

def load_disk_report(sn, now):
    for candidate in candidate_sns(sn):
        path = cache_path(candidate)
        if not path.exists():
            continue
        try:
            payload = json.loads(path.read_text())
            created_at = float(payload.get('created_at', 0))
            if now - created_at >= CACHE_TTL_SECONDS:
                continue
            report = payload.get('report')
            if not report:
                continue
            report['cache'] = {'hit': True, 'disk': True, 'created_at': created_at, 'ttl_seconds': CACHE_TTL_SECONDS}
            return report
        except Exception:
            path.unlink(missing_ok=True)
    return None

def save_disk_report(sn, report, created_at):
    serializable = dict(report)
    serializable.pop('cache', None)
    payload = {'created_at': created_at, 'report': serializable}
    path = cache_path(sn)
    path.write_text(json.dumps(payload, ensure_ascii=False, default=str))

def get_device_month_detail(sn: str, month: str):
    if not re.match(r'^\d{4}-\d{2}$', month or ''):
        raise HTTPException(status_code=400, detail="Invalid month")
    real_sn, _ = resolve_sn(sn)
    start = datetime.strptime(f"{month}-01", "%Y-%m-%d")
    end = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)

    summary = fetch_one(
        "SELECT COUNT(*) AS total_logs, COUNT(DISTINCT recipe_id) AS recipe_count, "
        "MIN(create_time) AS first_time, MAX(create_time) AS last_time, "
        "SUM(CAST(time AS SIGNED)) AS total_duration_seconds, ROUND(AVG(CAST(time AS SIGNED)), 1) AS avg_duration_seconds "
        "FROM sop_machinelog WHERE sn = %s AND create_time >= %s AND create_time < %s",
        (real_sn, start, end),
        source=True,
        database='btyc',
    ) or {}
    recipes = fetch_all(
        "SELECT recipe_id, MAX(recipe_name) AS recipe_name, COUNT(*) AS execution_count, "
        "SUM(CAST(time AS SIGNED)) AS total_duration_seconds, ROUND(AVG(CAST(time AS SIGNED)), 1) AS avg_duration_seconds, "
        "MIN(create_time) AS first_time, MAX(create_time) AS last_time "
        "FROM sop_machinelog WHERE sn = %s AND create_time >= %s AND create_time < %s "
        "GROUP BY recipe_id ORDER BY execution_count DESC, last_time DESC LIMIT 300",
        (real_sn, start, end),
        source=True,
        database='btyc',
    )
    logs = fetch_all(
        "SELECT id, recipe_id, recipe_name, time AS duration_seconds, create_time, end_time, username, data_time, comment "
        "FROM sop_machinelog WHERE sn = %s AND create_time >= %s AND create_time < %s "
        "ORDER BY create_time DESC LIMIT %s",
        (real_sn, start, end, MAX_DEVICE_LOGS),
        source=True,
        database='btyc',
    )
    logs = list(reversed(logs))
    total_logs = int(summary.get('total_logs') or 0)
    return {
        "sn": real_sn,
        "month": month,
        "summary": {
            "total_logs": total_logs,
            "returned_logs": len(logs),
            "truncated": total_logs > len(logs),
            "recipe_count": int(summary.get('recipe_count') or 0),
            "first_time": summary.get('first_time'),
            "last_time": summary.get('last_time'),
            "total_duration_seconds": int(summary.get('total_duration_seconds') or 0),
            "avg_duration_seconds": summary.get('avg_duration_seconds'),
        },
        "recipes": recipes,
        "logs": logs,
    }

RECIPE_SEARCH_SCOPES = {
    'customer': '客户维度',
    'device': '设备号维度',
    'active_machine': '当日/区间使用机器',
}
RECIPE_SEARCH_CACHE_VERSION = 1
RECIPE_ANALYSIS_CACHE_VERSION = 4

def parse_recipe_search_date(value, default_date, end=False):
    raw = (value or '').strip()
    if not raw:
        day = default_date
    else:
        try:
            day = datetime.strptime(raw[:10], '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式需要是 YYYY-MM-DD")
    if end:
        return datetime.combine(day, datetime.max.time().replace(microsecond=0))
    return datetime.combine(day, datetime.min.time())

def normalize_recipe_search_params(scope, keyword, recipe_keyword, start_date, end_date, limit):
    clean_scope = (scope or 'customer').strip()
    if clean_scope not in RECIPE_SEARCH_SCOPES:
        raise HTTPException(status_code=400, detail="不支持的检索维度")

    today = datetime.now().date()
    start_dt = parse_recipe_search_date(start_date, today - timedelta(days=7), end=False)
    end_dt = parse_recipe_search_date(end_date, today, end=True)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")

    span_days = (end_dt.date() - start_dt.date()).days + 1
    if span_days > MAX_RECIPE_SEARCH_DAYS:
        raise HTTPException(status_code=400, detail=f"单次检索最多支持 {MAX_RECIPE_SEARCH_DAYS} 天，请缩小时间范围")

    clean_keyword = (keyword or '').strip()
    clean_recipe_keyword = (recipe_keyword or '').strip()
    if clean_scope in {'customer', 'device'} and not clean_keyword:
        raise HTTPException(status_code=400, detail="客户/设备号维度需要填写模糊关键词")
    if clean_scope == 'active_machine' and not clean_keyword and not clean_recipe_keyword and span_days > 7:
        raise HTTPException(status_code=400, detail="空关键词查询使用机器最多支持 7 天，请缩小时间范围或补充关键词")

    capped_limit = max(20, min(int(limit or 200), MAX_RECIPE_SEARCH_LIMIT))
    return {
        'version': RECIPE_SEARCH_CACHE_VERSION,
        'scope': clean_scope,
        'scope_label': RECIPE_SEARCH_SCOPES[clean_scope],
        'keyword': clean_keyword,
        'recipe_keyword': clean_recipe_keyword,
        'start_date': start_dt.date().isoformat(),
        'end_date': end_dt.date().isoformat(),
        'start_time': start_dt,
        'end_time': end_dt,
        'span_days': span_days,
        'limit': capped_limit,
    }

def recipe_search_cache_key(params):
    payload = {
        'version': RECIPE_SEARCH_CACHE_VERSION,
        'scope': params['scope'],
        'keyword': params['keyword'],
        'recipe_keyword': params['recipe_keyword'],
        'start_date': params['start_date'],
        'end_date': params['end_date'],
        'limit': params['limit'],
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

def recipe_search_cache_path(key):
    cache_dir = CACHE_DIR / 'recipe_search'
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f'{key}.json'

def load_cached_recipe_search(params, now):
    key = recipe_search_cache_key(params)
    path = recipe_search_cache_path(key)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text())
        created_at = float(payload.get('created_at', 0))
        if now - created_at >= RECIPE_SEARCH_CACHE_TTL_SECONDS:
            return None
        result = payload.get('result')
        if not result:
            return None
        result['cache'] = {
            'hit': True,
            'disk': True,
            'created_at': created_at,
            'ttl_seconds': RECIPE_SEARCH_CACHE_TTL_SECONDS,
            'cache_key': key,
        }
        return result
    except Exception:
        path.unlink(missing_ok=True)
        return None

def save_cached_recipe_search(params, result, created_at):
    key = recipe_search_cache_key(params)
    serializable = dict(result)
    serializable.pop('cache', None)
    recipe_search_cache_path(key).write_text(json.dumps({
        'created_at': created_at,
        'params': {
            'scope': params['scope'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'limit': params['limit'],
        },
        'result': serializable,
    }, ensure_ascii=False, default=str))
    return key

def recipe_analysis_cache_key(params, recipe_id=None, recipe_name=''):
    payload = {
        'version': RECIPE_ANALYSIS_CACHE_VERSION,
        'type': 'recipe_analysis',
        'scope': params['scope'],
        'keyword': params['keyword'],
        'recipe_keyword': params['recipe_keyword'],
        'start_date': params['start_date'],
        'end_date': params['end_date'],
        'limit': params['limit'],
        'recipe_id': recipe_id,
        'recipe_name': recipe_name,
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

def recipe_analysis_cache_path(key):
    cache_dir = CACHE_DIR / 'recipe_analysis'
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f'{key}.json'

def get_cached_recipe_analysis(params, recipe_id=None, recipe_name='', force_refresh=False):
    key = recipe_analysis_cache_key(params, recipe_id=recipe_id, recipe_name=recipe_name)
    path = recipe_analysis_cache_path(key)
    now = time.time()
    if not force_refresh and path.exists():
        try:
            payload = json.loads(path.read_text())
            created_at = float(payload.get('created_at') or 0)
            if now - created_at < RECIPE_SEARCH_CACHE_TTL_SECONDS and payload.get('result'):
                result = payload['result']
                result['cache'] = {
                    'hit': True,
                    'disk': True,
                    'created_at': created_at,
                    'ttl_seconds': RECIPE_SEARCH_CACHE_TTL_SECONDS,
                    'cache_key': key,
                }
                return result
        except Exception:
            path.unlink(missing_ok=True)
    result = build_recipe_analysis_result(params, recipe_id=recipe_id, recipe_name=recipe_name)
    serializable = dict(result)
    serializable.pop('cache', None)
    path.write_text(json.dumps({'created_at': now, 'result': serializable}, ensure_ascii=False, default=str))
    result['cache'] = {
        'hit': False,
        'created_at': now,
        'ttl_seconds': RECIPE_SEARCH_CACHE_TTL_SECONDS,
        'cache_key': key,
    }
    return result

ANALYTICS_QUERY_TTL_SECONDS = int(os.getenv('ANALYTICS_QUERY_TTL_SECONDS', str(7 * 24 * 3600)))
ANALYTICS_OWNER_BATCH_SIZE = int(os.getenv('ANALYTICS_OWNER_BATCH_SIZE', '100'))
ANALYTICS_MAX_OWNERS = int(os.getenv('ANALYTICS_MAX_OWNERS', '50000'))
ANALYTICS_CACHE_DIR = CACHE_DIR / 'analytics'
ANALYTICS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
ANALYTICS_JOB_THREADS = {}

def parse_analytics_date(value, default_date, end=False):
    raw = (value or '').strip()
    day = default_date if not raw else datetime.strptime(raw[:10], '%Y-%m-%d').date()
    if end:
        return datetime.combine(day, datetime.max.time().replace(microsecond=0))
    return datetime.combine(day, datetime.min.time())

def normalize_recipe_top_params(payload: RecipeTopJobRequest):
    today = datetime.now().date()
    start_dt = parse_analytics_date(payload.start_date, today - timedelta(days=180), end=False)
    end_dt = parse_analytics_date(payload.end_date, today, end=True)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    span_days = (end_dt.date() - start_dt.date()).days + 1
    if span_days > 366:
        raise HTTPException(status_code=400, detail="Top 榜单次最多支持 366 天")
    top_n = max(50, min(int(payload.top_n or 500), 1000))
    sort_by = payload.sort_by if payload.sort_by in {'cooking_count', 'device_count', 'customer_count', 'total_duration_seconds'} else 'cooking_count'
    return {
        'query_type': 'recipe_top',
        'start_date': start_dt.date().isoformat(),
        'end_date': end_dt.date().isoformat(),
        'start_time': start_dt,
        'end_time': end_dt,
        'span_days': span_days,
        'top_n': top_n,
        'sort_by': sort_by,
        'recipe_keyword': (payload.recipe_keyword or '').strip(),
        'customer_keyword': (payload.customer_keyword or '').strip(),
        'sn': (payload.sn or '').strip(),
        'region': (payload.region or '').strip(),
        'category': (payload.category or '').strip(),
        'resource_type': (payload.resource_type or '').strip(),
        'stat_object': (payload.stat_object or 'all').strip(),
    }

def analytics_params_hash(params):
    stable = {k: v for k, v in params.items() if k not in {'start_time', 'end_time'}}
    return hashlib.sha256(json.dumps(stable, sort_keys=True, ensure_ascii=False, default=str).encode()).hexdigest()

def analytics_result_path(job_id):
    return ANALYTICS_CACHE_DIR / f'{job_id}.json'

def analytics_xlsx_path(job_id):
    return ANALYTICS_CACHE_DIR / f'{job_id}.xlsx'

def rowdict(row):
    return dict(row) if row else None

def get_analytics_job(job_id):
    with AUDIT_LOCK, audit_conn() as conn:
        row = conn.execute("SELECT * FROM analytics_query_jobs WHERE job_id = ?", (job_id,)).fetchone()
    job = rowdict(row)
    if job and job.get('params_json'):
        job['params'] = json.loads(job['params_json'])
    return job

def find_reusable_analytics_job(query_type, params_hash, now):
    with AUDIT_LOCK, audit_conn() as conn:
        row = conn.execute(
            "SELECT * FROM analytics_query_jobs WHERE query_type = ? AND params_hash = ? "
            "AND status IN ('PENDING','RUNNING','COMPLETED') "
            "AND (cache_expires_at IS NULL OR cache_expires_at > ?) "
            "ORDER BY created_at DESC LIMIT 1",
            (query_type, params_hash, now),
        ).fetchone()
    return rowdict(row)

def insert_analytics_job(query_type, params_hash, params, username):
    job_id = f"{query_type}_{uuid.uuid4().hex[:16]}"
    now = int(time.time())
    with AUDIT_LOCK, audit_conn() as conn:
        conn.execute(
            "INSERT INTO analytics_query_jobs(job_id, query_type, params_json, params_hash, status, stage, progress, created_by, created_at, cache_expires_at) "
            "VALUES (?, ?, ?, ?, 'PENDING', '排队中', 0, ?, ?, ?)",
            (job_id, query_type, json.dumps(params, ensure_ascii=False, default=str), params_hash, username, now, now + ANALYTICS_QUERY_TTL_SECONDS),
        )
        conn.commit()
    return job_id

def update_analytics_job(job_id, **fields):
    if not fields:
        return
    assignments = ', '.join([f"{key} = ?" for key in fields])
    values = list(fields.values()) + [job_id]
    with AUDIT_LOCK, audit_conn() as conn:
        conn.execute(f"UPDATE analytics_query_jobs SET {assignments} WHERE job_id = ?", values)
        conn.commit()

def start_analytics_thread(job_id, params, username):
    if job_id in ANALYTICS_JOB_THREADS and ANALYTICS_JOB_THREADS[job_id].is_alive():
        return
    thread = threading.Thread(target=run_recipe_top_job, args=(job_id, params, username), daemon=True)
    ANALYTICS_JOB_THREADS[job_id] = thread
    thread.start()

def resolve_recipe_top_owners(params):
    from_sql = "FROM btyc.sop_robot r LEFT JOIN btyc.ums_company c ON r.company_id = c.id"
    conditions = ["r.company_id IS NOT NULL", "r.company_id <> 0"]
    args = []
    if params.get('customer_keyword'):
        like = f"%{params['customer_keyword']}%"
        conditions.append("(c.common_name LIKE %s OR c.company_name LIKE %s OR c.geo_cityname LIKE %s OR c.geo_pname LIKE %s)")
        args.extend([like] * 4)
    if params.get('sn'):
        conditions.append("r.machinecode LIKE %s")
        args.append(f"%{params['sn']}%")
    if params.get('region'):
        like = f"%{params['region']}%"
        conditions.append("(c.geo_cityname LIKE %s OR c.geo_pname LIKE %s OR c.area_code LIKE %s)")
        args.extend([like] * 3)
    rows = fetch_all(
        f"SELECT DISTINCT r.company_id AS owner {from_sql} WHERE {' AND '.join(conditions)} LIMIT %s",
        tuple(args + [ANALYTICS_MAX_OWNERS]),
        source=True,
        database='btyc',
    )
    return [int(row['owner']) for row in rows if row.get('owner')]

def owner_chunks(owners):
    for idx in range(0, len(owners), ANALYTICS_OWNER_BATCH_SIZE):
        yield owners[idx:idx + ANALYTICS_OWNER_BATCH_SIZE]

def recipe_top_owner_sql(params, owners):
    placeholders = ','.join(['%s'] * len(owners))
    conditions = [f"l.owner IN ({placeholders})", "l.mac_time >= %s", "l.mac_time <= %s", "l.recipe_id IS NOT NULL", "l.recipe_id <> 0"]
    args = [*owners, params['start_date'], params['end_date']]
    if params.get('recipe_keyword'):
        like = f"%{params['recipe_keyword']}%"
        conditions.append("(l.recipe_name LIKE %s OR mr.name LIKE %s OR mr.group_name LIKE %s OR CAST(l.recipe_id AS CHAR) LIKE %s)")
        args.extend([like] * 4)
    if params.get('sn'):
        conditions.append("l.sn LIKE %s")
        args.append(f"%{params['sn']}%")
    if params.get('category'):
        conditions.append("mr.group_name LIKE %s")
        args.append(f"%{params['category']}%")
    sql = f"""
        SELECT l.recipe_id,
               COALESCE(MAX(mr.name), MAX(l.recipe_name), CONCAT('菜谱', l.recipe_id)) AS recipe_name,
               COALESCE(MAX(mr.group_name), '未分类') AS category,
               MAX(mr.type) AS recipe_type,
               COUNT(*) AS cooking_count,
               COUNT(DISTINCT l.sn) AS device_count,
               COUNT(DISTINCT l.owner) AS customer_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               SUM(CAST(l.time AS SIGNED)) AS total_duration_seconds,
               ROUND(AVG(CAST(l.time AS SIGNED)), 1) AS avg_duration_seconds
        FROM btyc.sop_machinelog l FORCE INDEX(idx_owner_mac_time)
        LEFT JOIN manage_backend.main_recipe mr ON l.recipe_id = mr.id
        WHERE {' AND '.join(conditions)}
        GROUP BY l.recipe_id
    """
    return sql, tuple(args)

def merge_recipe_top_partials(partial_rows, params, limit):
    merged = {}
    for row in partial_rows:
        rid = int(row['recipe_id'])
        item = merged.setdefault(rid, {
            'recipe_id': rid,
            'recipe_name': row.get('recipe_name'),
            'category': row.get('category') or '未分类',
            'recipe_type': row.get('recipe_type'),
            'cooking_count': 0,
            'device_count': 0,
            'customer_count': 0,
            'first_time': row.get('first_time'),
            'last_time': row.get('last_time'),
            'total_duration_seconds': 0,
        })
        item['recipe_name'] = item.get('recipe_name') or row.get('recipe_name')
        item['category'] = item.get('category') or row.get('category') or '未分类'
        item['cooking_count'] += int(row.get('cooking_count') or 0)
        item['device_count'] += int(row.get('device_count') or 0)
        item['customer_count'] += int(row.get('customer_count') or 0)
        if row.get('first_time') and (not item.get('first_time') or row['first_time'] < item['first_time']):
            item['first_time'] = row['first_time']
        if row.get('last_time') and (not item.get('last_time') or row['last_time'] > item['last_time']):
            item['last_time'] = row['last_time']
        item['total_duration_seconds'] += int(row.get('total_duration_seconds') or 0)
    rows = []
    for item in merged.values():
        count = max(1, int(item.get('cooking_count') or 0))
        item['avg_duration_seconds'] = round(item.get('total_duration_seconds', 0) / count, 1)
        rows.append(item)
    sort_key = params.get('sort_by') or 'cooking_count'
    return sorted(rows, key=lambda row: (row.get(sort_key) or 0, row.get('last_time') or ''), reverse=True)[:limit]

def ingredient_classes(row):
    text = ' '.join(str(row.get(k) or '') for k in ['ingredient_name', 'category_1', 'category_2', 'raw_json'])
    water = any(k in text for k in ['水', '清水', '饮用水', '开水', '冷水', '热水', '高汤', '汤汁', '汤底', '水淀粉'])
    oil = any(k in text for k in ['油', '猪油', '菜籽油', '色拉油', '橄榄油', 'lard', 'oil'])
    liquid = str(row.get('ingredient_type')) == '3' or oil or water or any(k in text for k in ['汁', '酱', '醋', '料酒', '生抽', '老抽', '蚝油', '汤', '液', '乳', '奶'])
    seasoning = any(k in text for k in ['盐', '精盐', '鸡精', '味精', '生抽', '老抽', '蚝油', '醋', '料酒', '酱', '粉'])
    return {'liquid': liquid, 'water': water, 'oil': oil, 'seasoning': seasoning}

def summarize_recipe_top_rows(top_rows, process_rows):
    ingredients = process_rows.get('ingredients', [])
    steps = process_rows.get('cook_steps', [])
    raw_rows = process_rows.get('raw_process', [])
    by_recipe = {int(row['recipe_id']): row for row in top_rows if row.get('recipe_id')}
    summary = {rid: {
        'ingredient_count': 0,
        'cook_step_count': 0,
        'has_detail': False,
        'liquid_feed_count': 0,
        'liquid_amount': 0,
        'water_feed_count': 0,
        'water_amount': 0,
        'oil_feed_count': 0,
        'oil_amount': 0,
        'seasoning_feed_count': 0,
        'seasoning_amount': 0,
    } for rid in by_recipe}
    for row in raw_rows:
        rid = int(row.get('recipe_id') or 0)
        if rid in summary:
            summary[rid]['ingredient_count'] = int(row.get('ingredient_count') or 0)
            summary[rid]['cook_step_count'] = int(row.get('cook_steps_count') or 0)
            summary[rid]['has_detail'] = not bool(row.get('detail_missing'))
    for row in steps:
        rid = int(row.get('recipe_id') or 0)
        if rid in summary:
            summary[rid]['cook_step_count'] = max(summary[rid]['cook_step_count'], int(row.get('step_index') or 0))
    for row in ingredients:
        rid = int(row.get('recipe_id') or 0)
        if rid not in summary:
            continue
        classes = ingredient_classes(row)
        amount = normalized_amount(row.get('dosage'), row.get('unit'))
        if classes['liquid']:
            summary[rid]['liquid_feed_count'] += 1
            if amount is not None:
                summary[rid]['liquid_amount'] += amount
        if classes['water']:
            summary[rid]['water_feed_count'] += 1
            if amount is not None:
                summary[rid]['water_amount'] += amount
        if classes['oil']:
            summary[rid]['oil_feed_count'] += 1
            if amount is not None:
                summary[rid]['oil_amount'] += amount
        if classes['seasoning']:
            summary[rid]['seasoning_feed_count'] += 1
            if amount is not None:
                summary[rid]['seasoning_amount'] += amount
        row.update({
            'is_liquid': classes['liquid'],
            'is_water': classes['water'],
            'is_oil': classes['oil'],
            'is_seasoning': classes['seasoning'],
            'normalized_amount': amount,
        })
    result = []
    for index, row in enumerate(top_rows, start=1):
        rid = int(row.get('recipe_id') or 0)
        stats = summary.get(rid, {})
        result.append({
            'rank': index,
            **row,
            **stats,
            'liquid_amount': round(stats.get('liquid_amount', 0), 2),
            'water_amount': round(stats.get('water_amount', 0), 2),
            'oil_amount': round(stats.get('oil_amount', 0), 2),
            'seasoning_amount': round(stats.get('seasoning_amount', 0), 2),
        })
    return result

def fetch_recipe_top_coverage(params, recipe_ids, owners):
    if not recipe_ids:
        return [], []
    placeholders = ','.join(['%s'] * len(recipe_ids))
    devices = []
    customers = []
    for chunk in owner_chunks(owners):
        owner_placeholders = ','.join(['%s'] * len(chunk))
        base_conditions = f"l.owner IN ({owner_placeholders}) AND l.mac_time >= %s AND l.mac_time <= %s AND l.recipe_id IN ({placeholders})"
        args = [*chunk, params['start_date'], params['end_date'], *recipe_ids]
        devices.extend(fetch_all(
            f"""
            SELECT l.recipe_id, l.sn, MAX(r.name) AS device_name, MAX(r.robot_type) AS robot_type,
                   MAX(r.latest_update_package) AS version,
                   COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
                   COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region,
                   COUNT(*) AS cooking_count, MIN(l.create_time) AS first_time, MAX(l.create_time) AS last_time
            FROM btyc.sop_machinelog l FORCE INDEX(idx_owner_mac_time)
            LEFT JOIN btyc.sop_robot r ON l.sn = r.machinecode
            LEFT JOIN btyc.ums_company c ON l.owner = c.id
            WHERE {base_conditions}
            GROUP BY l.recipe_id, l.sn
            """,
            tuple(args),
            source=True,
            database='btyc',
        ))
        customers.extend(fetch_all(
            f"""
            SELECT l.recipe_id, l.owner AS company_id,
                   COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
                   COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region,
                   COUNT(*) AS cooking_count, COUNT(DISTINCT l.sn) AS device_count,
                   MIN(l.create_time) AS first_time, MAX(l.create_time) AS last_time
            FROM btyc.sop_machinelog l FORCE INDEX(idx_owner_mac_time)
            LEFT JOIN btyc.ums_company c ON l.owner = c.id
            WHERE {base_conditions}
            GROUP BY l.recipe_id, l.owner
            """,
            tuple(args),
            source=True,
            database='btyc',
        ))
    return devices, customers

def build_recipe_top_result(params, job_id=None):
    owners = resolve_recipe_top_owners(params)
    if job_id:
        update_analytics_job(job_id, stage=f'查询生产日志（{len(owners)} 个 owner）', progress=18)
    partial_rows = []
    chunks = list(owner_chunks(owners))
    for index, chunk in enumerate(chunks, start=1):
        sql, args = recipe_top_owner_sql(params, chunk)
        partial_rows.extend(fetch_all(sql, args, source=True, database='btyc'))
        if job_id and (index == len(chunks) or index % 10 == 0):
            progress = 18 + int(32 * index / max(1, len(chunks)))
            update_analytics_job(job_id, stage=f'查询生产日志 {index}/{len(chunks)}', progress=progress)
    if job_id:
        update_analytics_job(job_id, stage='合并 TopN', progress=55)
    fetch_limit = min(max(params['top_n'] * 3, params['top_n']), 3000) if params.get('resource_type') else params['top_n']
    top_rows = merge_recipe_top_partials(partial_rows, params, fetch_limit)
    for row in top_rows:
        row['execution_count'] = row.get('cooking_count')
    if job_id:
        update_analytics_job(job_id, stage='读取菜谱详情', progress=62)
    top_rows = hydrate_recipe_search_resources(top_rows)
    if params.get('resource_type'):
        top_rows = [row for row in top_rows if params['resource_type'] in (row.get('resources') or [])][:params['top_n']]
    else:
        top_rows = top_rows[:params['top_n']]
    process_rows = fetch_recipe_process_export_rows(top_rows)
    if job_id:
        update_analytics_job(job_id, stage='统计投料与步骤', progress=72)
    summary = summarize_recipe_top_rows(top_rows, process_rows)
    recipe_ids = [int(row['recipe_id']) for row in summary if row.get('recipe_id')]
    if job_id:
        update_analytics_job(job_id, stage='读取设备/客户覆盖', progress=78)
    devices, customers = fetch_recipe_top_coverage(params, recipe_ids, owners)
    liquid_rows = [
        row for row in process_rows.get('ingredients', [])
        if row.get('is_liquid') or row.get('is_water') or row.get('is_oil')
    ]
    return {
        'params': {k: v for k, v in params.items() if k not in {'start_time', 'end_time'}},
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'summary': {
            'recipe_count': len(summary),
            'owner_count': len(owners),
            'device_coverage_rows': len(devices),
            'customer_coverage_rows': len(customers),
        },
        'top_recipes': summary,
        'liquid_water_ingredients': liquid_rows,
        'all_ingredients': process_rows.get('ingredients', []),
        'cook_steps': process_rows.get('cook_steps', []),
        'raw_process': process_rows.get('raw_process', []),
        'device_coverage': devices,
        'customer_coverage': customers,
        'rule_notes': [
            '液料优先按 btyc.base_ingredients.ingredinent_type = 3 判断，并用油/汁/酱/醋/料酒/生抽/老抽/蚝油/高汤/汤/液/乳/奶等关键词兜底。',
            '水按食材名、分类或原始字段中包含水、清水、饮用水、开水、冷水、热水、高汤、汤汁、汤底、水淀粉等关键词判断。',
            'g/克直接计入，kg/千克/公斤乘1000，L/升乘1000，ml/毫升保留原数值并保留原单位；不确定单位保留原值和原始JSON。',
        ],
    }

def build_recipe_top_workbook(result, job):
    wb = Workbook()
    wb.remove(wb.active)
    params = result.get('params', {})
    write_sheet(wb, '口径说明', [
        {'field': '查询类型', 'value': '菜谱 Top 榜与投料统计'},
        {'field': '开始日期', 'value': params.get('start_date')},
        {'field': '结束日期', 'value': params.get('end_date')},
        {'field': 'Top N', 'value': params.get('top_n')},
        {'field': '排序指标', 'value': params.get('sort_by')},
        {'field': '生成时间', 'value': result.get('generated_at')},
        {'field': '任务ID', 'value': job.get('job_id')},
        {'field': '缓存过期', 'value': datetime.fromtimestamp(job.get('cache_expires_at')).strftime('%Y-%m-%d %H:%M:%S') if job.get('cache_expires_at') else ''},
        {'field': '液料/水判断规则', 'value': '；'.join(result.get('rule_notes', []))},
    ], [('field', '字段'), ('value', '值')])
    write_sheet(wb, 'Top菜谱汇总', result.get('top_recipes', []), [
        ('rank', '排名'), ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '标准菜名/分类'),
        ('cooking_count', '烹饪次数'), ('device_count', '覆盖设备数'), ('customer_count', '覆盖客户数'),
        ('first_time', '首次烹饪时间'), ('last_time', '末次烹饪时间'),
        ('total_duration_seconds', '累计烹饪时长秒'), ('avg_duration_seconds', '平均烹饪时长秒'),
        ('liquid_feed_count', '液料投料次数'), ('liquid_amount', '液料克数/毫升数'),
        ('water_feed_count', '水投料次数'), ('water_amount', '水克数/毫升数'),
        ('oil_feed_count', '油投料次数'), ('oil_amount', '油克数/毫升数'),
        ('seasoning_feed_count', '调料投料次数'), ('seasoning_amount', '调料克数/毫升数'),
        ('ingredient_count', '配料项数量'), ('cook_step_count', '烹饪步骤数量'), ('has_detail', '是否有原始详情'), ('resources', '资源覆盖')
    ])
    ingredient_cols = [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('ingredient_index', '配料序号'),
        ('cooking_step_id', '关联烹饪步骤ID'), ('ingredient_id', '食材ID'), ('ingredient_name', '食材名称'),
        ('dosage', '原始用量'), ('unit', '原始单位'), ('normalized_amount', '折算克/毫升数'),
        ('is_liquid', '是否液料'), ('is_water', '是否水'), ('is_oil', '是否油'), ('is_seasoning', '是否调料'),
        ('preprocess', '备菜处理方式'), ('feeding_mode', '投料模式'), ('ingredient_type', '食材类型'), ('category_1', '一级分类'), ('category_2', '二级分类'), ('raw_json', '原始JSON')
    ]
    write_sheet(wb, '液料水投料明细', result.get('liquid_water_ingredients', []), ingredient_cols)
    write_sheet(wb, '全部配料明细', result.get('all_ingredients', []), ingredient_cols)
    write_sheet(wb, '全部烹饪步骤', result.get('cook_steps', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('step_index', '步骤序号'),
        ('time', '时间'), ('type_label', '步骤类型'), ('automatic_label', '自动/手动'), ('power', '功率'),
        ('speed', '搅拌速度'), ('position', '锅位'), ('movepot', '翻锅/移锅'), ('commands', '原始指令内容'), ('execution_content', '机器执行内容'), ('raw_json', '原始步骤JSON')
    ])
    write_sheet(wb, '菜谱原始JSON', result.get('raw_process', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('cook_time', '烹饪时长'),
        ('cook_steps_count', '烹饪步骤数'), ('ingredient_count', '配料数'), ('description', '备菜/预制说明'),
        ('cook_steps_json', '烹饪步骤JSON'), ('cooking_ingredient_json', '配料JSON'), ('ingredient_note_json', '备菜须知JSON'), ('serve_note_json', '出菜须知JSON')
    ])
    write_sheet(wb, '设备覆盖明细', result.get('device_coverage', []), [
        ('recipe_id', '菜谱ID'), ('sn', '设备SN'), ('device_name', '设备名'), ('customer_name', '客户'), ('region', '地区'),
        ('robot_type', '设备类型'), ('version', '版本'), ('cooking_count', '烹饪次数'), ('first_time', '首次'), ('last_time', '末次')
    ])
    write_sheet(wb, '客户覆盖明细', result.get('customer_coverage', []), [
        ('recipe_id', '菜谱ID'), ('company_id', '客户ID'), ('customer_name', '客户'), ('region', '地区'),
        ('cooking_count', '烹饪次数'), ('device_count', '设备数'), ('first_time', '首次'), ('last_time', '末次')
    ])
    return wb

def run_recipe_top_job(job_id, params, username):
    started = int(time.time())
    try:
        update_analytics_job(job_id, status='RUNNING', stage='查询生产日志', progress=12, started_at=started)
        result = build_recipe_top_result(params, job_id=job_id)
        update_analytics_job(job_id, stage='生成 Excel', progress=82)
        result_path = analytics_result_path(job_id)
        result_path.write_text(json.dumps(result, ensure_ascii=False, default=str))
        job = get_analytics_job(job_id) or {'job_id': job_id}
        xlsx_path = analytics_xlsx_path(job_id)
        wb = build_recipe_top_workbook(result, job)
        wb.save(xlsx_path)
        update_analytics_job(
            job_id,
            status='COMPLETED',
            stage='完成',
            progress=100,
            finished_at=int(time.time()),
            cache_expires_at=int(time.time()) + ANALYTICS_QUERY_TTL_SECONDS,
            result_path=str(result_path),
            xlsx_path=str(xlsx_path),
            error_message='',
        )
    except Exception as exc:
        update_analytics_job(job_id, status='FAILED', stage='失败', progress=100, finished_at=int(time.time()), error_message=str(exc)[:1000])

def resolve_recipe_search_sns(params):
    keyword = params['keyword']
    like = f"%{keyword}%" if keyword else None
    if params['scope'] in {'customer', 'device'}:
        if params['scope'] == 'customer':
            where_sql = (
                "(c.company_name LIKE %s OR c.common_name LIKE %s OR c.addr LIKE %s OR "
                "c.company_addr LIKE %s OR c.geo_pname LIKE %s OR c.geo_cityname LIKE %s OR "
                "c.geo_adname LIKE %s OR c.geo_name LIKE %s OR c.geo_address LIKE %s OR "
                "c.contact_name LIKE %s OR r.name LIKE %s OR r.spec LIKE %s)"
            )
            args = [like] * 12
        else:
            where_sql = (
                "(r.machinecode LIKE %s OR r.name LIKE %s OR r.spec LIKE %s OR "
                "r.robot_type LIKE %s OR r.latest_update_package LIKE %s)"
            )
            args = [like] * 5
        rows = fetch_all(
            f"""
            SELECT DISTINCT r.machinecode AS sn
            FROM btyc.sop_robot r
            LEFT JOIN btyc.ums_company c ON r.company_id = c.id
            WHERE r.machinecode IS NOT NULL AND r.machinecode != '' AND {where_sql}
            ORDER BY r.machinecode
            LIMIT 2000
            """,
            tuple(args),
            source=True,
            database='btyc',
        )
        return [row['sn'] for row in rows if row.get('sn')]

    conditions = ["d.`烹饪日期` >= %s", "d.`烹饪日期` <= %s"]
    args = [params['start_time'], params['end_time']]
    if like:
        conditions.append(
            "(d.`设备编号` LIKE %s OR d.`门店名称` LIKE %s OR d.`企业名称` LIKE %s OR "
            "r.name LIKE %s OR r.spec LIKE %s OR c.company_name LIKE %s OR c.common_name LIKE %s)"
        )
        args.extend([like] * 7)
    rows = fetch_all(
        f"""
        SELECT DISTINCT d.`设备编号` AS sn
        FROM btyc_statics.robot_cook_day d
        LEFT JOIN btyc.sop_robot r ON d.`设备编号` = r.machinecode
        LEFT JOIN btyc.ums_company c ON r.company_id = c.id
        WHERE d.`设备编号` IS NOT NULL AND d.`设备编号` != '' AND {' AND '.join(conditions)}
        ORDER BY d.`设备编号`
        LIMIT 2000
        """,
        tuple(args),
        source=True,
        database='btyc_statics',
    )
    return [row['sn'] for row in rows if row.get('sn')]

def recipe_search_sql_parts(params):
    detail_join = "LEFT JOIN manage_backend.recipe_detail rd ON l.recipe_id = rd.recipe_id" if params['recipe_keyword'] else ""
    from_sql = f"""
        FROM btyc.sop_machinelog l
        LEFT JOIN btyc.sop_robot r ON l.sn = r.machinecode
        LEFT JOIN btyc.ums_company c ON r.company_id = c.id
        LEFT JOIN manage_backend.main_recipe mr ON l.recipe_id = mr.id
        {detail_join}
    """
    conditions = ["l.create_time >= %s", "l.create_time <= %s"]
    args = [params['start_time'], params['end_time']]

    sn_filter = params.get('sn_filter') or []
    if sn_filter:
        placeholders = ','.join(['%s'] * len(sn_filter))
        conditions.append(f"l.sn IN ({placeholders})")
        args.extend(sn_filter)

    recipe_keyword = params['recipe_keyword']
    if recipe_keyword:
        like = f"%{recipe_keyword}%"
        conditions.append(
            "(l.recipe_name LIKE %s OR mr.name LIKE %s OR mr.group_name LIKE %s OR "
            "CAST(l.recipe_id AS CHAR) LIKE %s OR rd.cook_steps LIKE %s OR "
            "rd.cooking_ingredient LIKE %s OR rd.wash_steps LIKE %s OR rd.moisten_steps LIKE %s)"
        )
        args.extend([like] * 8)

    return from_sql, " AND ".join(conditions), args

def hydrate_recipe_search_resources(recipe_rows):
    recipe_ids = [int(row['recipe_id']) for row in recipe_rows if row.get('recipe_id')]
    if not recipe_ids:
        for row in recipe_rows:
            row['resources'] = []
            row['has_lard'] = False
            row['resource_flags'] = {}
        return recipe_rows

    placeholders = ','.join(['%s'] * len(recipe_ids))
    raw_recipes = fetch_all(
        f"SELECT id, name, group_name, type, steps_describe, ingredients_total_dosage "
        f"FROM main_recipe WHERE id IN ({placeholders})",
        tuple(recipe_ids),
        source=True,
        database='manage_backend',
    )
    raw_details = fetch_all(
        f"SELECT recipe_id, cook_time, cook_steps, wash_steps, moisten_steps, cooking_ingredient "
        f"FROM recipe_detail WHERE recipe_id IN ({placeholders})",
        tuple(recipe_ids),
        source=True,
        database='manage_backend',
    )
    recipe_by_id = {int(row['id']): row for row in raw_recipes}
    detail_by_recipe_id = {int(row['recipe_id']): row for row in raw_details}

    for row in recipe_rows:
        recipe_id = int(row['recipe_id']) if row.get('recipe_id') else None
        recipe = recipe_by_id.get(recipe_id)
        if recipe:
            usage = {
                'cnt': row.get('execution_count'),
                'total_duration_seconds': row.get('total_duration_seconds'),
                'first_time': row.get('first_time'),
                'last_time': row.get('last_time'),
            }
            archive = recipe_resource_summary(recipe, detail_by_recipe_id.get(recipe_id), {recipe_id: usage})
            row['resources'] = archive.get('resources', [])
            row['has_lard'] = bool(archive.get('has_lard'))
            row['resource_flags'] = archive.get('resource_flags', {})
            row['category'] = row.get('category') or archive.get('category') or '未分类'
            row['recipe_name'] = row.get('recipe_name') or archive.get('name')
        else:
            row['resources'] = []
            row['has_lard'] = False
            row['resource_flags'] = {}
            row['category'] = row.get('category') or '菜谱详情缺失'
    return recipe_rows

def json_cell(value):
    if value is None:
        return ''
    return json.dumps(value, ensure_ascii=False, default=str)

def fetch_ingredient_name_map(ingredient_ids):
    ids = sorted({str(item).strip() for item in ingredient_ids if str(item or '').strip()})
    if not ids:
        return {}
    placeholders = ','.join(['%s'] * len(ids))
    rows = fetch_all(
        f"SELECT ingredinent_id, ingredients_name, ingredinent_type, categories_1, categories_2, automatic, lang "
        f"FROM base_ingredients WHERE ingredinent_id IN ({placeholders}) "
        f"ORDER BY CASE WHEN lang = 'cn' THEN 0 WHEN lang = 'zh' THEN 1 WHEN lang = '' THEN 2 ELSE 3 END, lang",
        tuple(ids),
        source=True,
        database='btyc',
    )
    result = {}
    for row in rows:
        key = str(row['ingredinent_id'])
        if key not in result:
            result[key] = row
    return result

def first_present(data, *keys):
    for key in keys:
        value = data.get(key)
        if value not in (None, ''):
            return value
    return None

def number_value(value):
    if value in (None, ''):
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None

def normalized_amount(value, unit):
    raw = number_value(value)
    if raw is None:
        return None
    clean_unit = str(unit or '').strip().lower()
    if clean_unit in {'kg', '千克', '公斤'}:
        return raw * 1000
    if clean_unit in {'l', '升'}:
        return raw * 1000
    if clean_unit in {'g', '克', 'ml', '毫升'}:
        return raw
    return None

THERMAL_RULES = [
    {'category': '自定义/待确认', 'keywords': ['自定义', '未知', '测试', '默认', '其他', '手动捞出'], 'specific_heat': None, 'water_fraction': None, 'oil_fraction': None, 'hazard_class': '待人工归类', 'confidence': '低'},
    {'category': '水/汤汁', 'keywords': ['水', '清水', '饮用水', '开水', '冷水', '热水', '冰水', '高汤', '汤', '汤汁', '汤底', 'water', '水淀粉', '淀粉水', '生粉水'], 'specific_heat': 4.0, 'water_fraction': 0.9, 'oil_fraction': 0.0, 'boiling_c': 100, 'hazard_class': '低风险含水液体', 'confidence': '中'},
    {'category': '液体调料', 'keywords': ['酱油', '醬油', 'soy sauce', '生抽', '老抽', '醋', '米醋', '陈醋', '香醋', '料酒', '黄酒', '米酒', '白酒', '啤酒', '蚝油', '蠔油', '耗油', '素蠔油', '鱼露', '豉油', '东古', '美极', '一品鲜', '酱汁', '调味汁', '鲜味汁', '辣鲜露', '蒸鱼豉油', 'white sauce', 'squid sauce', 'sweet and sour sauce', '다크소스', '汁', '液'], 'specific_heat': 3.7, 'water_fraction': 0.75, 'oil_fraction': 0.02, 'boiling_c': 100, 'hazard_class': '含水液体调料', 'confidence': '低'},
    {'category': '酱料/发酵调料', 'keywords': ['豆瓣酱', '辣酱', '甜面酱', '黄豆酱', '柱侯酱', '海鲜酱', '芝麻酱', '沙茶酱', '酱料', '酱包', '豆豉', '豆䜴', '黑豆豉', 'black bean', 'black beans', '磨豉酱', '剁椒', '泡椒', '鲜椒酱', '脆椒酱', '淳牌酱', '干锅酱', '酱椒酱', 'pasta sauce', '小炒料', '鲜上鲜', '复合', '湖洋', '底料'], 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.08, 'hazard_class': '复合调料/需复核', 'confidence': '低'},
    {'category': '油脂', 'keywords': ['oil', 'aliejus', '猪油', '牛油', '鸡油', '鸭油', '黄油', 'butter', '色拉油', '沙拉油', '菜籽油', '花生油', '大豆油', '玉米油', '调和油', '混合油', '植物油', '食用油', '香油', '麻油', '芝麻油', '葱油', '花椒油', '辣椒油', '油辣子', '红油', '明油', '料油', '熟油', '油脂', '油桶'], 'exclude_keywords': ['酱油', '醬油', 'soy sauce', '蚝油', '蠔油', '耗油', '豉油'], 'specific_heat': 2.0, 'water_fraction': 0.0, 'oil_fraction': 1.0, 'smoke_point_c': None, 'flash_point_c': None, 'autoignition_c': None, 'hazard_class': '可燃油脂/温度待校准', 'confidence': '低'},
    {'category': '香辛料/干货', 'keywords': ['干辣椒', '辣椒干', '干椒', '辣椒粉', '辣椒面', '刀口辣椒', '胡椒', '胡椒粉', 'black pepper', '花椒', '麻椒', '藤椒', '八角', '八角粉', '桂皮', '香叶', '孜然', '孜然粉', '芝麻', '白芝麻', '黑芝麻', '花生米', '十三香', '五香粉', '咖喱粉', '香料', '丁香', '草果', '白芷', '山奈', '良姜', '陈皮', '香茅', '甘草', '沙姜', '肉蔻'], 'specific_heat': 1.6, 'water_fraction': 0.12, 'oil_fraction': 0.08, 'hazard_class': '可燃干货/燃点待测', 'confidence': '低'},
    {'category': '鲜椒/蔬菜', 'keywords': ['小米辣', '小米椒', '美人椒', '野山椒', '子天椒', '指天椒', '贡椒', '黄贡椒', '软皮椒', '二荆条', '土辣椒', '手锤红辣椒', '塌辣子', '辣椒', '辣椒片', '辣椒圈', '辣椒段', '辣椒碎', '青椒', '红椒', '尖椒', '线椒', '螺丝椒', '杭椒', 'capsicum', 'mix pepper', 'chili sliced', 'pepper', '菜', '葱', '䓤', '蔥', '姜', '薑', '蒜', '마늘', 'garlic', '笋', '土豆', '萝卜', '罗卜', '豆角', '缸豆', '豇豆', '四季豆', '扁豆', '豆芽', 'bean sprouts', '洋葱', '洋蔥', 'onion', 'courgettes', '茄', '瓜', '芹', '韭', 'leek', '蒜苗', '蒜苔', '莲藕', '藕', '茭白', '花菜', '西蓝花', '包菜', 'cabbage', '白菜', '香菜', '西红柿', '番茄', '玉米粒', 'carrot', 'mixed vege', '藠头', '芥兰', '九层塔', '金不换', '果蔬粒', 'pineapple', '菠萝', '芋头', '甜豆', '兰豆', '红苕尖'], 'specific_heat': 3.7, 'water_fraction': 0.82, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '中'},
    {'category': '菌菇类', 'keywords': ['菌', '菇', '蘑菇', '香菇', '平菇', '杏鲍菇', '金针菇', '木耳', '银耳'], 'specific_heat': 3.6, 'water_fraction': 0.85, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '中'},
    {'category': '肉类', 'keywords': ['鸡', 'chicken', '牛', 'beef', '猪', 'pork', '羊', '鸭', '鹅', '肉', '里脊', '排骨', '肥肠', '肥膘', '油渣', '叉烧', '腊肠', '香肠', '火腿', '午餐肉'], 'specific_heat': 3.2, 'water_fraction': 0.62, 'oil_fraction': 0.1, 'hazard_class': '含水可焦化食材', 'confidence': '中'},
    {'category': '水产类', 'keywords': ['鱼', '虾', 'prawn', '蟹', '贝', '蛤', '鱿鱼', 'squid', '墨鱼', '海参', '鲍鱼', '牛蛙'], 'specific_heat': 3.4, 'water_fraction': 0.72, 'oil_fraction': 0.04, 'hazard_class': '含水可焦化食材', 'confidence': '中'},
    {'category': '主食/淀粉食材', 'keywords': ['米饭', 'rice', '杂粮饭', '米粉', '河粉', '面条', 'noodles', 'yellow noodle', 'hokkien noodles', 'fettucine', 'fettuccine', 'pasta', '面片', '面疙瘩', '年糕', '粉皮', '粉丝', '粉条', '米线', '饼', '馍', '面筋'], 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.02, 'hazard_class': '淀粉类/可焦化', 'confidence': '低'},
    {'category': '蛋奶豆制品', 'keywords': ['egg', 'eggs', '蛋', '鸡蛋', '鸭蛋', 'cheese', '奶', 'cream', 'crème', '乳', '豆腐', '豆皮', '腐竹', '香干', '豆干', '千张', '青豆', '杂豆'], 'specific_heat': 3.3, 'water_fraction': 0.68, 'oil_fraction': 0.06, 'hazard_class': '含水可焦化食材', 'confidence': '中'},
    {'category': '粉类/增稠', 'keywords': ['starch powder', 'cheese powder', '淀粉', '生粉', '面粉', '玉米粉', '土豆粉', '红薯粉', '粉料', '裹粉', '浆粉'], 'specific_heat': 1.6, 'water_fraction': 0.12, 'oil_fraction': 0.0, 'hazard_class': '粉类/需复核', 'confidence': '低'},
    {'category': '糖类', 'keywords': ['sugar', '糖', '白糖', '冰糖', '红糖', '砂糖', '二砂', '麦芽糖', '蜂蜜'], 'specific_heat': 1.3, 'water_fraction': 0.02, 'oil_fraction': 0.0, 'hazard_class': '糖类/可焦化', 'confidence': '低'},
    {'category': '盐味精/基础调味', 'keywords': ['salt', 'msg', 'baking soda', '鹽', '盐', '味精', '鸡精', '鸡粉', '味粉', '调味粉', '增味剂'], 'specific_heat': 0.9, 'water_fraction': 0.02, 'oil_fraction': 0.0, 'hazard_class': '无机/低燃烧风险', 'confidence': '低'},
]

SOURCE_CATEGORY_RULES = {
    ('3', '1'): {'category': '鲜椒/蔬菜', 'specific_heat': 3.7, 'water_fraction': 0.82, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('3', '2'): {'category': '菌菇类', 'specific_heat': 3.6, 'water_fraction': 0.85, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('3', '23'): {'category': '鲜椒/蔬菜', 'specific_heat': 3.7, 'water_fraction': 0.82, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('4', '46'): {'category': '油脂', 'specific_heat': 2.0, 'water_fraction': 0.0, 'oil_fraction': 1.0, 'hazard_class': '可燃油脂/温度待校准', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('4', '47'): {'category': '油脂', 'specific_heat': 2.0, 'water_fraction': 0.0, 'oil_fraction': 1.0, 'hazard_class': '可燃油脂/温度待校准', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '33'): {'category': '液体调料', 'specific_heat': 3.7, 'water_fraction': 0.75, 'oil_fraction': 0.02, 'boiling_c': 100, 'hazard_class': '含水液体调料', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '40'): {'category': '酱料/发酵调料', 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.08, 'hazard_class': '复合调料/需复核', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '42'): {'category': '酱料/发酵调料', 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.08, 'hazard_class': '复合调料/需复核', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '43'): {'category': '香辛料/干货', 'specific_heat': 1.6, 'water_fraction': 0.12, 'oil_fraction': 0.08, 'hazard_class': '可燃干货/燃点待测', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '44'): {'category': '液体调料', 'specific_heat': 3.7, 'water_fraction': 0.75, 'oil_fraction': 0.02, 'boiling_c': 100, 'hazard_class': '含水液体调料', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '45'): {'category': '糖类', 'specific_heat': 1.3, 'water_fraction': 0.02, 'oil_fraction': 0.0, 'hazard_class': '糖类/可焦化', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('6', '48'): {'category': '粉类/增稠', 'specific_heat': 1.6, 'water_fraction': 0.12, 'oil_fraction': 0.0, 'hazard_class': '粉类/需复核', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('8', '44'): {'category': '盐味精/基础调味', 'specific_heat': 0.9, 'water_fraction': 0.02, 'oil_fraction': 0.0, 'hazard_class': '无机/低燃烧风险', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('1', '22'): {'category': '主食/淀粉食材', 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.02, 'hazard_class': '淀粉类/可焦化', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('10', '5'): {'category': '香辛料/干货', 'specific_heat': 1.6, 'water_fraction': 0.12, 'oil_fraction': 0.08, 'hazard_class': '可燃干货/燃点待测', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('10', '27'): {'category': '酱料/发酵调料', 'specific_heat': 2.6, 'water_fraction': 0.45, 'oil_fraction': 0.08, 'hazard_class': '复合调料/需复核', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('10', '29'): {'category': '鲜椒/蔬菜', 'specific_heat': 3.7, 'water_fraction': 0.82, 'oil_fraction': 0.0, 'hazard_class': '含水低风险食材', 'confidence': '低', 'source_note': '源库分类编码推断'},
    ('2', '14'): {'category': '蛋奶豆制品', 'specific_heat': 3.3, 'water_fraction': 0.68, 'oil_fraction': 0.06, 'hazard_class': '含水可焦化食材', 'confidence': '低', 'source_note': '源库分类编码推断'},
}

def infer_thermal_property(name='', category_1='', category_2='', ingredient_type=None):
    text = ' '.join(str(x or '') for x in [name, category_1, category_2]).lower()
    if str(ingredient_type or '') == '3':
        text = f"{text} 液"
    for rule in THERMAL_RULES:
        if any(keyword and keyword.lower() in text for keyword in rule.get('exclude_keywords', [])):
            continue
        if any(keyword and keyword.lower() in text for keyword in rule['keywords']):
            result = dict(rule)
            result.pop('keywords', None)
            result.pop('exclude_keywords', None)
            result.setdefault('boiling_c', None)
            result.setdefault('smoke_point_c', None)
            result.setdefault('flash_point_c', None)
            result.setdefault('autoignition_c', None)
            result['source_note'] = '源库分类+名称关键词推断'
            return result
    code_rule = SOURCE_CATEGORY_RULES.get((str(category_1 or ''), str(category_2 or '')))
    if code_rule:
        result = dict(code_rule)
        result.setdefault('boiling_c', None)
        result.setdefault('smoke_point_c', None)
        result.setdefault('flash_point_c', None)
        result.setdefault('autoignition_c', None)
        return result
    return {
        'category': '未分类',
        'specific_heat': None,
        'water_fraction': None,
        'oil_fraction': None,
        'boiling_c': None,
        'smoke_point_c': None,
        'flash_point_c': None,
        'autoignition_c': None,
        'hazard_class': '待归类',
        'confidence': '低',
        'source_note': '默认兜底/待人工校准',
    }

def canonical_ingredient_key(ingredient_id, name):
    raw_id = str(ingredient_id or '').strip()
    if raw_id:
        return raw_id
    return f"name:{str(name or '').strip()}"

def upsert_thermal_ingredient_rows(rows):
    if not rows:
        return 0
    ensure_analytics_db()
    conn = get_conn()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sql = """
        INSERT INTO ingredient_thermal_properties (
            source_ingredient_id, canonical_name, aliases_json, category, source_category_1, source_category_2,
            ingredient_type, automatic, specific_heat_kj_kg_c, water_fraction, oil_fraction, boiling_c,
            smoke_point_c, flash_point_c, autoignition_c, hazard_class, confidence, source_note,
            recipe_usage_count, recipe_count, total_amount_g, total_amount_ml, last_seen_recipe_id, created_at, updated_at
        ) VALUES (
            %(source_ingredient_id)s, %(canonical_name)s, %(aliases_json)s, %(category)s, %(source_category_1)s, %(source_category_2)s,
            %(ingredient_type)s, %(automatic)s, %(specific_heat_kj_kg_c)s, %(water_fraction)s, %(oil_fraction)s, %(boiling_c)s,
            %(smoke_point_c)s, %(flash_point_c)s, %(autoignition_c)s, %(hazard_class)s, %(confidence)s, %(source_note)s,
            %(recipe_usage_count)s, %(recipe_count)s, %(total_amount_g)s, %(total_amount_ml)s, %(last_seen_recipe_id)s, %(created_at)s, %(updated_at)s
        )
        ON DUPLICATE KEY UPDATE
            canonical_name = VALUES(canonical_name),
            aliases_json = VALUES(aliases_json),
            category = VALUES(category),
            source_category_1 = VALUES(source_category_1),
            source_category_2 = VALUES(source_category_2),
            ingredient_type = VALUES(ingredient_type),
            automatic = VALUES(automatic),
            specific_heat_kj_kg_c = VALUES(specific_heat_kj_kg_c),
            water_fraction = VALUES(water_fraction),
            oil_fraction = VALUES(oil_fraction),
            boiling_c = VALUES(boiling_c),
            smoke_point_c = VALUES(smoke_point_c),
            flash_point_c = VALUES(flash_point_c),
            autoignition_c = VALUES(autoignition_c),
            hazard_class = VALUES(hazard_class),
            confidence = VALUES(confidence),
            source_note = VALUES(source_note),
            recipe_usage_count = GREATEST(recipe_usage_count, VALUES(recipe_usage_count)),
            recipe_count = GREATEST(recipe_count, VALUES(recipe_count)),
            total_amount_g = GREATEST(total_amount_g, VALUES(total_amount_g)),
            total_amount_ml = GREATEST(total_amount_ml, VALUES(total_amount_ml)),
            last_seen_recipe_id = COALESCE(VALUES(last_seen_recipe_id), last_seen_recipe_id),
            updated_at = VALUES(updated_at)
    """
    payload = []
    for row in rows:
        item = dict(row)
        for key in ['aliases_json', 'source_category_1', 'source_category_2', 'ingredient_type', 'automatic', 'boiling_c', 'smoke_point_c', 'flash_point_c', 'autoignition_c', 'last_seen_recipe_id']:
            item.setdefault(key, None)
        item.setdefault('recipe_usage_count', 0)
        item.setdefault('recipe_count', 0)
        item.setdefault('total_amount_g', 0)
        item.setdefault('total_amount_ml', 0)
        item['created_at'] = now
        item['updated_at'] = now
        payload.append(item)
    try:
        with conn.cursor() as cur:
            cur.executemany(sql, payload)
        conn.commit()
        return len(payload)
    finally:
        conn.close()

def sync_base_ingredients_to_thermal():
    rows = fetch_all(
        "SELECT ingredinent_id, ingredients_name, ingredinent_type, categories_1, categories_2, automatic, lang "
        "FROM base_ingredients WHERE ingredients_name IS NOT NULL AND ingredients_name != '' "
        "ORDER BY ingredinent_id, CASE WHEN lang = 'cn' THEN 0 WHEN lang = 'zh' THEN 1 WHEN lang = '' THEN 2 ELSE 3 END",
        source=True,
        database='btyc',
    )
    best = {}
    aliases = defaultdict(set)
    for row in rows:
        key = str(row.get('ingredinent_id') or '').strip()
        name = str(row.get('ingredients_name') or '').strip()
        if not key or not name:
            continue
        aliases[key].add(name)
        if key not in best:
            best[key] = row
    payload = []
    for key, row in best.items():
        name = str(row.get('ingredients_name') or '').strip()
        inferred = infer_thermal_property(name, row.get('categories_1'), row.get('categories_2'), row.get('ingredinent_type'))
        payload.append({
            'source_ingredient_id': key,
            'canonical_name': name,
            'aliases_json': json.dumps(sorted(aliases[key] - {name}), ensure_ascii=False),
            'category': inferred['category'],
            'source_category_1': row.get('categories_1'),
            'source_category_2': row.get('categories_2'),
            'ingredient_type': row.get('ingredinent_type'),
            'automatic': row.get('automatic'),
            'specific_heat_kj_kg_c': inferred['specific_heat'],
            'water_fraction': inferred['water_fraction'],
            'oil_fraction': inferred['oil_fraction'],
            'boiling_c': inferred.get('boiling_c'),
            'smoke_point_c': inferred.get('smoke_point_c'),
            'flash_point_c': inferred.get('flash_point_c'),
            'autoignition_c': inferred.get('autoignition_c'),
            'hazard_class': inferred['hazard_class'],
            'confidence': inferred['confidence'],
            'source_note': inferred['source_note'],
        })
    total = 0
    for i in range(0, len(payload), 1000):
        total += upsert_thermal_ingredient_rows(payload[i:i + 1000])
    return {'source_rows': len(rows), 'ingredient_rows': total}

def sync_recipe_ingredients_to_thermal(limit=20000):
    limit = max(100, min(int(limit or 20000), 100000))
    rows = fetch_all(
        "SELECT rd.recipe_id, mr.name AS recipe_name, mr.group_name, rd.cooking_ingredient "
        "FROM recipe_detail rd JOIN main_recipe mr ON mr.id = rd.recipe_id "
        "WHERE rd.cooking_ingredient IS NOT NULL AND rd.cooking_ingredient != '' "
        "ORDER BY rd.recipe_id DESC LIMIT %s",
        (limit,),
        source=True,
        database='manage_backend',
    )
    agg = {}
    recipe_sets = defaultdict(set)
    ingredient_ids = set()
    for row in rows:
        recipe_id = row.get('recipe_id')
        for item in parse_json_array(row.get('cooking_ingredient')):
            if not isinstance(item, dict):
                continue
            ingredient_id = first_present(item, 'ingredientsId', 'Ingredients_id', 'ingredientId', 'id')
            name = first_present(item, 'name', 'materialName', 'food_name', 'ingredient_name')
            key = canonical_ingredient_key(ingredient_id, name)
            if not key or key == 'name:':
                continue
            if ingredient_id:
                ingredient_ids.add(str(ingredient_id))
            entry = agg.setdefault(key, {
                'source_ingredient_id': str(ingredient_id or key),
                'canonical_name': str(name or '').strip() or str(ingredient_id),
                'aliases': set(),
                'recipe_usage_count': 0,
                'total_amount_g': 0.0,
                'total_amount_ml': 0.0,
                'last_seen_recipe_id': recipe_id,
            })
            if name:
                entry['aliases'].add(str(name).strip())
            entry['recipe_usage_count'] += 1
            entry['last_seen_recipe_id'] = recipe_id
            recipe_sets[key].add(recipe_id)
            unit = first_present(item, 'ingredientsUnit', 'Ingredients_unit', 'unit', 'dosageUnit', 'unit_name')
            amount = normalized_amount(first_present(item, 'ingredientsDosage', 'Ingredients_dosage', 'dosage', 'weight', 'num', 'amount'), unit)
            unit_text = str(unit or '').strip().lower()
            if amount:
                if unit_text in {'ml', '毫升', 'l', '升'}:
                    entry['total_amount_ml'] += amount
                else:
                    entry['total_amount_g'] += amount
    meta = fetch_ingredient_name_map(ingredient_ids)
    payload = []
    for key, entry in agg.items():
        m = meta.get(str(entry['source_ingredient_id'])) or {}
        name = str(m.get('ingredients_name') or entry['canonical_name'] or '').strip()
        inferred = infer_thermal_property(name, m.get('categories_1'), m.get('categories_2'), m.get('ingredinent_type'))
        aliases = sorted((entry['aliases'] | {entry['canonical_name']}) - {name})
        payload.append({
            'source_ingredient_id': entry['source_ingredient_id'],
            'canonical_name': name,
            'aliases_json': json.dumps(aliases, ensure_ascii=False),
            'category': inferred['category'],
            'source_category_1': m.get('categories_1'),
            'source_category_2': m.get('categories_2'),
            'ingredient_type': m.get('ingredinent_type'),
            'automatic': m.get('automatic'),
            'specific_heat_kj_kg_c': inferred['specific_heat'],
            'water_fraction': inferred['water_fraction'],
            'oil_fraction': inferred['oil_fraction'],
            'boiling_c': inferred.get('boiling_c'),
            'smoke_point_c': inferred.get('smoke_point_c'),
            'flash_point_c': inferred.get('flash_point_c'),
            'autoignition_c': inferred.get('autoignition_c'),
            'hazard_class': inferred['hazard_class'],
            'confidence': inferred['confidence'],
            'source_note': '菜谱配料聚合+' + inferred['source_note'],
            'recipe_usage_count': entry['recipe_usage_count'],
            'recipe_count': len(recipe_sets[key]),
            'total_amount_g': round(entry['total_amount_g'], 3),
            'total_amount_ml': round(entry['total_amount_ml'], 3),
            'last_seen_recipe_id': entry['last_seen_recipe_id'],
        })
    total = 0
    for i in range(0, len(payload), 1000):
        total += upsert_thermal_ingredient_rows(payload[i:i + 1000])
    return {'recipe_rows': len(rows), 'ingredient_rows': total}

def latest_thermal_sync():
    ensure_analytics_db()
    rows = fetch_all("SELECT * FROM ingredient_thermal_sync_runs ORDER BY started_at DESC LIMIT 1")
    return rows[0] if rows else None

def thermal_knowledge_summary():
    ensure_analytics_db()
    row = fetch_one(
        "SELECT COUNT(*) AS total, COUNT(DISTINCT category) AS categories, "
        "SUM(CASE WHEN category = '油脂' OR oil_fraction >= 0.5 THEN 1 ELSE 0 END) AS oils, "
        "SUM(CASE WHEN hazard_class LIKE '%%可燃%%' OR hazard_class LIKE '%%油脂%%' OR hazard_class LIKE '%%挥发%%' THEN 1 ELSE 0 END) AS risky, "
        "SUM(recipe_usage_count) AS usage_count FROM ingredient_thermal_properties"
    ) or {}
    cats = fetch_all("SELECT category, COUNT(*) AS count FROM ingredient_thermal_properties GROUP BY category ORDER BY count DESC, category")
    hazards = fetch_all("SELECT hazard_class, COUNT(*) AS count FROM ingredient_thermal_properties GROUP BY hazard_class ORDER BY count DESC, hazard_class")
    return {
        'total': int(row.get('total') or 0),
        'categories': int(row.get('categories') or 0),
        'oils': int(row.get('oils') or 0),
        'risky': int(row.get('risky') or 0),
        'usage_count': int(row.get('usage_count') or 0),
        'category_options': cats,
        'hazard_options': hazards,
        'last_sync': latest_thermal_sync(),
    }

# ── Safety Scan Engine ──────────────────────────────────────────────

SAFETY_RULES = [
    {
        'key': 'max_temp_high',
        'label': '实测高温 > 330℃',
        'level': 'high',
        'score': 100,
        'sql_condition': 'max_pot_temp > 330',
        'description': '锅体最高温度超过330℃，接近油脂自燃风险区',
    },
    {
        'key': 'max_temp_elevated',
        'label': '实测高温 280-330℃',
        'level': 'medium',
        'score': 60,
        'sql_condition': 'max_pot_temp BETWEEN 280 AND 330',
        'description': '锅体温度偏高，需关注投料和功率控制',
    },
    {
        'key': 'no_power_samples',
        'label': '缺功率采样',
        'level': 'medium',
        'score': 50,
        'sql_condition': None,
        'description': '作业有温度数据但缺少功率采样，测温或通讯可能异常',
    },
    {
        'key': 'no_temp_samples',
        'label': '缺温度样本',
        'level': 'medium',
        'score': 55,
        'sql_condition': '(sample_count = 0 OR max_pot_temp IS NULL)',
        'description': '作业窗口内无有效温度样本，传感器可能故障',
    },
    {
        'key': 'recipe_high_freq',
        'label': '高频高温菜谱',
        'level': 'info',
        'score': 30,
        'sql_condition': None,
        'description': '该菜谱在全局统计中高温比例偏高',
    },
]


def run_safety_scan(scan_type='quick'):
    """Run safety rules against all cook_jobs in local DB. Returns scan run id."""
    ensure_analytics_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_id = None
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO safety_scan_runs(scan_type, status, started_at) VALUES (%s, %s, %s)",
                (scan_type, 'RUNNING', now),
            )
            run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    total_alerts = 0
    stats = {'total_jobs': 0, 'high_temp_jobs': 0, 'delay_risk_jobs': 0, 'sensor_gap_jobs': 0}

    try:
        # ── Rule 1 & 2: Temperature-based rules ──
        for rule in SAFETY_RULES:
            if not rule.get('sql_condition'):
                continue
            sql = f"""
                SELECT cj.id AS cook_job_id, cj.sn, cj.source_file_id, cj.recipe_name,
                       cj.cook_start_time, cj.cook_end_time, cj.duration_seconds,
                       cj.max_pot_temp, cj.avg_pot_temp, cj.sample_count,
                       cj.step_count, cj.android_action_count
                FROM cook_jobs cj
                WHERE {rule['sql_condition']}
                ORDER BY cj.max_pot_temp DESC
                LIMIT 5000
            """
            rows = fetch_all(sql)
            alerts = []
            for row in rows:
                alerts.append((
                    run_id, row.get('sn'), row.get('source_file_id'), row.get('cook_job_id'),
                    row.get('recipe_name'), row.get('cook_start_time'), row.get('cook_end_time'),
                    row.get('duration_seconds'), rule['key'], rule['label'], rule['level'],
                    rule['score'], row.get('max_pot_temp'), row.get('avg_pot_temp'),
                    None, None,
                    json.dumps({'rule': rule['key'], 'condition': rule['sql_condition']}, ensure_ascii=False),
                    now,
                ))
            if alerts:
                placeholders = ','.join(['(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'] * len(alerts))
                flat = [v for tup in alerts for v in tup]
                execute_local(
                    f"INSERT INTO safety_scan_alerts(scan_run_id, sn, source_file_id, cook_job_id, recipe_name, cook_start_time, cook_end_time, duration_seconds, rule_key, rule_label, risk_level, severity_score, max_pot_temp, avg_pot_temp, actual_energy_kwh, oil_to_food_interval, detail_json, created_at) VALUES {placeholders}",
                    flat,
                )
                total_alerts += len(alerts)
                if rule['key'] == 'max_temp_high':
                    stats['high_temp_jobs'] = len(alerts)
                if rule['key'] in ('no_power_samples', 'no_temp_samples'):
                    stats['sensor_gap_jobs'] += len(alerts)

            stats['total_jobs'] = (fetch_one("SELECT COUNT(*) AS cnt FROM cook_jobs") or {}).get('cnt', 0)

        # ── Rule "no_power_samples": jobs with temperature but no power events ──
        no_power_sql = """
            SELECT cj.id AS cook_job_id, cj.sn, cj.source_file_id, cj.recipe_name,
                   cj.cook_start_time, cj.cook_end_time, cj.duration_seconds,
                   cj.max_pot_temp, cj.avg_pot_temp, cj.sample_count
            FROM cook_jobs cj
            LEFT JOIN cook_power_events cpe ON cpe.cook_job_id = cj.id
            WHERE cj.sample_count > 0 AND cpe.id IS NULL
            ORDER BY cj.max_pot_temp DESC
            LIMIT 5000
        """
        no_power_rows = fetch_all(no_power_sql)
        no_power_alerts = []
        for row in (no_power_rows or []):
            no_power_alerts.append((
                run_id, row.get('sn'), row.get('source_file_id'), row.get('cook_job_id'),
                row.get('recipe_name'), row.get('cook_start_time'), row.get('cook_end_time'),
                row.get('duration_seconds'), 'no_power_samples', '缺功率采样',
                'medium', 50, row.get('max_pot_temp'), row.get('avg_pot_temp'),
                None, None,
                json.dumps({'sample_count': row.get('sample_count')}, ensure_ascii=False),
                now,
            ))
        if no_power_alerts:
            placeholders = ','.join(['(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'] * len(no_power_alerts))
            flat = [v for tup in no_power_alerts for v in tup]
            execute_local(
                f"INSERT INTO safety_scan_alerts(scan_run_id, sn, source_file_id, cook_job_id, recipe_name, cook_start_time, cook_end_time, duration_seconds, rule_key, rule_label, risk_level, severity_score, max_pot_temp, avg_pot_temp, actual_energy_kwh, oil_to_food_interval, detail_json, created_at) VALUES {placeholders}",
                flat,
            )
            total_alerts += len(no_power_alerts)
            stats['sensor_gap_jobs'] += len(no_power_alerts)

        # ── Rule: Oil-to-food delay (join action_events) ──
        delay_sql = """
            SELECT cj.id AS cook_job_id, cj.sn, cj.source_file_id, cj.recipe_name,
                   cj.cook_start_time, cj.cook_end_time, cj.duration_seconds,
                   cj.max_pot_temp, cj.avg_pot_temp,
                   oil.offset_seconds AS oil_offset,
                   food.offset_seconds AS food_offset,
                   (food.offset_seconds - oil.offset_seconds) AS oil_to_food_interval
            FROM cook_jobs cj
            JOIN cook_action_events oil ON oil.cook_job_id = cj.id AND oil.event_type = 'add_oil'
            JOIN cook_action_events food ON food.cook_job_id = cj.id
                AND food.event_type IN ('ingredient', 'stir', 'roll_move', 'pump')
                AND food.offset_seconds > oil.offset_seconds
            WHERE (food.offset_seconds - oil.offset_seconds) > 60
            GROUP BY cj.id, cj.sn, cj.source_file_id, cj.recipe_name,
                     cj.cook_start_time, cj.cook_end_time, cj.duration_seconds,
                     cj.max_pot_temp, cj.avg_pot_temp, oil.offset_seconds, food.offset_seconds
            ORDER BY (food.offset_seconds - oil.offset_seconds) DESC
            LIMIT 5000
        """
        delay_rows = fetch_all(delay_sql)
        delay_alerts = []
        for row in delay_rows:
            interval = int(row.get('oil_to_food_interval') or 0)
            score = min(100, 40 + interval // 2)
            level = 'high' if interval > 120 else 'medium'
            delay_alerts.append((
                run_id, row.get('sn'), row.get('source_file_id'), row.get('cook_job_id'),
                row.get('recipe_name'), row.get('cook_start_time'), row.get('cook_end_time'),
                row.get('duration_seconds'), 'oil_food_delay', '投油到主料间隔过长',
                level, score, row.get('max_pot_temp'), row.get('avg_pot_temp'),
                None, interval,
                json.dumps({'oil_offset': row.get('oil_offset'), 'food_offset': row.get('food_offset'), 'interval_s': interval}, ensure_ascii=False),
                now,
            ))
        if delay_alerts:
            placeholders = ','.join(['(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'] * len(delay_alerts))
            flat = [v for tup in delay_alerts for v in tup]
            execute_local(
                f"INSERT INTO safety_scan_alerts(scan_run_id, sn, source_file_id, cook_job_id, recipe_name, cook_start_time, cook_end_time, duration_seconds, rule_key, rule_label, risk_level, severity_score, max_pot_temp, avg_pot_temp, actual_energy_kwh, oil_to_food_interval, detail_json, created_at) VALUES {placeholders}",
                flat,
            )
            total_alerts += len(delay_alerts)
            stats['delay_risk_jobs'] = len(delay_alerts)

        # ── Update daily stats ──
        stats_sql = """
            INSERT INTO safety_daily_stats(stat_date, sn, total_jobs, high_temp_300c, high_temp_330c,
                oil_delay_60s, sensor_gap, max_temp_reached, avg_temp_across_jobs, total_energy_kwh, created_at)
            SELECT
                DATE(cook_start_time) AS stat_date,
                sn,
                COUNT(*) AS total_jobs,
                SUM(CASE WHEN max_pot_temp >= 300 THEN 1 ELSE 0 END),
                SUM(CASE WHEN max_pot_temp >= 330 THEN 1 ELSE 0 END),
                0,
                SUM(CASE WHEN sample_count = 0 OR max_pot_temp IS NULL THEN 1 ELSE 0 END),
                MAX(max_pot_temp),
                AVG(avg_pot_temp),
                NULL,
                NOW()
            FROM cook_jobs
            WHERE cook_start_time IS NOT NULL
            GROUP BY stat_date, sn
            ON DUPLICATE KEY UPDATE
                total_jobs = VALUES(total_jobs),
                high_temp_300c = VALUES(high_temp_300c),
                high_temp_330c = VALUES(high_temp_330c),
                max_temp_reached = VALUES(max_temp_reached),
                avg_temp_across_jobs = VALUES(avg_temp_across_jobs),
                created_at = NOW()
        """
        execute_local(stats_sql)

        finished = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        execute_local(
            "UPDATE safety_scan_runs SET status=%s, total_jobs=%s, high_temp_jobs=%s, delay_risk_jobs=%s, sensor_gap_jobs=%s, total_alerts=%s, finished_at=%s WHERE id=%s",
            ('COMPLETED', stats['total_jobs'], stats['high_temp_jobs'], stats['delay_risk_jobs'], stats['sensor_gap_jobs'], total_alerts, finished, run_id),
        )
        return {'run_id': run_id, 'alerts': total_alerts, 'stats': stats, 'status': 'COMPLETED'}

    except Exception as exc:
        finished = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        execute_local(
            "UPDATE safety_scan_runs SET status=%s, error_message=%s, finished_at=%s WHERE id=%s",
            ('FAILED', str(exc), finished, run_id),
        )
        raise


def safety_overview():
    """Aggregate safety stats for dashboard overview."""
    ensure_analytics_db()
    # High-level counts
    summary = fetch_one("""
        SELECT
            COUNT(*) AS total_jobs,
            SUM(CASE WHEN max_pot_temp >= 330 THEN 1 ELSE 0 END) AS critical_jobs,
            SUM(CASE WHEN max_pot_temp >= 300 AND max_pot_temp < 330 THEN 1 ELSE 0 END) AS warning_jobs,
            SUM(CASE WHEN max_pot_temp >= 250 AND max_pot_temp < 300 THEN 1 ELSE 0 END) AS caution_jobs,
            MAX(max_pot_temp) AS all_time_max_temp,
            AVG(max_pot_temp) AS all_time_avg_max_temp,
            COUNT(DISTINCT sn) AS device_count
        FROM cook_jobs
        WHERE max_pot_temp IS NOT NULL
    """) or {}

    # Top risky devices
    top_devices = fetch_all("""
        SELECT sn, COUNT(*) AS high_count, MAX(max_pot_temp) AS worst_temp,
               AVG(max_pot_temp) AS avg_high_temp
        FROM cook_jobs
        WHERE max_pot_temp >= 280
        GROUP BY sn
        HAVING high_count >= 2
        ORDER BY high_count DESC, worst_temp DESC
        LIMIT 20
    """)

    # Top risky recipes
    top_recipes = fetch_all("""
        SELECT recipe_name, COUNT(*) AS high_count, MAX(max_pot_temp) AS worst_temp,
               AVG(max_pot_temp) AS avg_high_temp
        FROM cook_jobs
        WHERE max_pot_temp >= 280
        GROUP BY recipe_name
        ORDER BY high_count DESC
        LIMIT 20
    """)

    # Temperature distribution
    temp_dist = fetch_all("""
        SELECT
            CASE
                WHEN max_pot_temp >= 350 THEN '>=350℃'
                WHEN max_pot_temp >= 300 THEN '300-350℃'
                WHEN max_pot_temp >= 250 THEN '250-300℃'
                WHEN max_pot_temp >= 200 THEN '200-250℃'
                WHEN max_pot_temp >= 150 THEN '150-200℃'
                WHEN max_pot_temp >= 100 THEN '100-150℃'
                ELSE '<100℃'
            END AS band,
            COUNT(*) AS cnt
        FROM cook_jobs
        WHERE max_pot_temp IS NOT NULL
        GROUP BY band
        ORDER BY MIN(max_pot_temp) DESC
    """)

    # Last scan
    last_scan = fetch_one("SELECT * FROM safety_scan_runs WHERE status='COMPLETED' ORDER BY finished_at DESC LIMIT 1")

    # Active alert counts
    alert_counts = fetch_one("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN risk_level='high' AND dismissed=0 THEN 1 ELSE 0 END) AS high_active,
            SUM(CASE WHEN risk_level='medium' AND dismissed=0 THEN 1 ELSE 0 END) AS medium_active,
            SUM(CASE WHEN dismissed=1 THEN 1 ELSE 0 END) AS dismissed
        FROM safety_scan_alerts
    """) or {}

    return {
        'summary': {
            'total_jobs': int(summary.get('total_jobs') or 0),
            'critical_jobs': int(summary.get('critical_jobs') or 0),
            'warning_jobs': int(summary.get('warning_jobs') or 0),
            'caution_jobs': int(summary.get('caution_jobs') or 0),
            'all_time_max_temp': float(summary.get('all_time_max_temp') or 0),
            'all_time_avg_max_temp': round(float(summary.get('all_time_avg_max_temp') or 0), 1),
            'device_count': int(summary.get('device_count') or 0),
        },
        'top_devices': [{
            'sn': r.get('sn'),
            'high_count': int(r.get('high_count') or 0),
            'worst_temp': float(r.get('worst_temp') or 0),
            'avg_high_temp': round(float(r.get('avg_high_temp') or 0), 1),
        } for r in (top_devices or [])],
        'top_recipes': [{
            'recipe_name': r.get('recipe_name'),
            'high_count': int(r.get('high_count') or 0),
            'worst_temp': float(r.get('worst_temp') or 0),
            'avg_high_temp': round(float(r.get('avg_high_temp') or 0), 1),
        } for r in (top_recipes or [])],
        'temp_distribution': [{
            'band': r.get('band'),
            'cnt': int(r.get('cnt') or 0),
        } for r in (temp_dist or [])],
        'last_scan': {
            'scan_type': last_scan.get('scan_type') if last_scan else None,
            'status': last_scan.get('status') if last_scan else None,
            'total_alerts': int(last_scan.get('total_alerts') or 0) if last_scan else 0,
            'finished_at': last_scan.get('finished_at').strftime("%Y-%m-%d %H:%M:%S") if last_scan and last_scan.get('finished_at') else None,
        } if last_scan else None,
        'alert_counts': {
            'total': int(alert_counts.get('total') or 0),
            'high_active': int(alert_counts.get('high_active') or 0),
            'medium_active': int(alert_counts.get('medium_active') or 0),
            'dismissed': int(alert_counts.get('dismissed') or 0),
        },
    }


def safety_alerts_query(risk_level=None, rule_key=None, sn=None, dismissed=None, limit=200, offset=0):
    """Paginated safety alerts query."""
    ensure_analytics_db()
    conditions = []
    args = []
    if risk_level:
        conditions.append("risk_level = %s")
        args.append(risk_level)
    if rule_key:
        conditions.append("rule_key = %s")
        args.append(rule_key)
    if sn:
        conditions.append("sn = %s")
        args.append(sn)
    if dismissed is not None:
        conditions.append("dismissed = %s")
        args.append(int(dismissed))

    where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    total_row = fetch_one(f"SELECT COUNT(*) AS total FROM safety_scan_alerts {where_sql}", tuple(args))
    rows = fetch_all(
        f"SELECT * FROM safety_scan_alerts {where_sql} ORDER BY severity_score DESC, created_at DESC LIMIT %s OFFSET %s",
        tuple(args + [limit, offset]),
    )

    items = []
    for row in (rows or []):
        item = dict(row)
        for dt_field in ('cook_start_time', 'cook_end_time', 'created_at', 'dismissed_at'):
            if item.get(dt_field) and hasattr(item[dt_field], 'strftime'):
                item[dt_field] = item[dt_field].strftime("%Y-%m-%d %H:%M:%S")
        if item.get('detail_json') and isinstance(item.get('detail_json'), str):
            try:
                item['detail'] = json.loads(item['detail_json'])
            except Exception:
                item['detail'] = {}
        else:
            item['detail'] = item.get('detail_json') or {}
        items.append(item)

    rule_counts = fetch_all("""
        SELECT rule_key, rule_label, risk_level, COUNT(*) AS cnt
        FROM safety_scan_alerts WHERE dismissed=0
        GROUP BY rule_key, rule_label, risk_level ORDER BY cnt DESC
    """)

    return {
        'total': int(total_row.get('total') or 0),
        'limit': limit,
        'offset': offset,
        'items': items,
        'rule_counts': [dict(r) for r in (rule_counts or [])],
    }

def step_type_label(value):
    labels = {
        '1': '投料',
        '2': '等待/执行',
        '3': '机器控制',
        '4': '出菜/提示',
    }
    return labels.get(str(value), str(value) if value not in (None, '') else '')

def auto_label(value):
    if str(value) == '1':
        return '自动'
    if str(value) == '0':
        return '手动'
    return value

def build_machine_execution_text(step):
    parts = []
    text = command_text(step)
    if text:
        parts.append(text)
    power = first_present(step, 'power')
    speed = first_present(step, 'speed', 'stirSpeed')
    position = first_present(step, 'position')
    if power not in (None, '', '0', '0.0'):
        parts.append(f"功率:{power}W")
    if speed not in (None, '', '0', '0.0'):
        parts.append(f"搅拌:{speed}")
    if position not in (None, '', '0', '0.0'):
        parts.append(f"锅位:{position}")
    if str(first_present(step, 'movepot') or '0') not in {'', '0', '0.0'}:
        parts.append(f"翻锅/移锅:{first_present(step, 'movepot')}")
    direction = first_present(step, 'direction')
    if direction not in (None, '', '0', '0.0'):
        parts.append(f"方向:{direction}")
    duration = first_present(step, 'time')
    if duration not in (None, '', '0', '0.0'):
        parts.append(f"时间:{duration}s")
    return '；'.join(parts)

def note_rows(recipe_id, section, value, recipe=None, usage=None, base_builder=None):
    parsed = parse_json_value(value)
    if parsed in (None, ''):
        return []
    items = parsed if isinstance(parsed, list) else [parsed]
    rows = []
    for index, item in enumerate(items, start=1):
        if isinstance(item, dict):
            content = first_present(item, 'content', 'note', 'desc', 'description', 'commands', 'name', 'text')
            raw = item
        else:
            content = str(item)
            raw = {'value': item}
        row = base_builder(recipe_id, recipe, usage) if base_builder else {'recipe_id': recipe_id}
        row.update({
            'section': section,
            'note_index': index,
            'content': content,
            'raw_json': json_cell(raw),
        })
        rows.append(row)
    return rows

def fetch_recipe_process_export_rows(recipe_rows):
    recipe_ids = sorted({
        int(row['recipe_id'])
        for row in recipe_rows
        if row.get('recipe_id')
    })
    if not recipe_ids:
        return {
            'cook_steps': [],
            'wash_steps': [],
            'moisten_steps': [],
            'ingredients': [],
            'prep_notes': [],
            'serve_notes': [],
            'temperature_curve': [],
            'raw_process': [],
        }

    placeholders = ','.join(['%s'] * len(recipe_ids))
    raw_recipes = fetch_all(
        f"SELECT id, name, group_name, type, description, steps_describe, ingredients_total_dosage, "
        f"weight, portion_size, cooking_time, pot_type, power_type, max_power, apply_moisten_pot "
        f"FROM main_recipe WHERE id IN ({placeholders})",
        tuple(recipe_ids),
        source=True,
        database='manage_backend',
    )
    raw_details = fetch_all(
        f"SELECT recipe_id, cook_time, cook_steps, wash_steps, moisten_steps, cooking_ingredient, "
        f"ingredient_note, serve_note, temperature_curve, initial_temperature, initial_temperature_array "
        f"FROM recipe_detail WHERE recipe_id IN ({placeholders})",
        tuple(recipe_ids),
        source=True,
        database='manage_backend',
    )
    recipe_by_id = {int(row['id']): row for row in raw_recipes}
    detail_by_recipe_id = {int(row['recipe_id']): row for row in raw_details}
    usage_by_recipe_id = {
        int(row['recipe_id']): row
        for row in recipe_rows
        if row.get('recipe_id')
    }
    ingredient_ids = []
    for detail in raw_details:
        for item in parse_json_array(detail.get('cooking_ingredient')):
            if isinstance(item, dict):
                ingredient_ids.append(first_present(item, 'ingredientsId', 'Ingredients_id', 'ingredientId', 'id'))
    ingredient_name_map = fetch_ingredient_name_map(ingredient_ids)

    def base(recipe_id, recipe=None, usage=None):
        recipe = recipe or {}
        usage = usage or {}
        return {
            'recipe_id': recipe_id,
            'recipe_name': recipe.get('name') or usage.get('recipe_name'),
            'category': recipe.get('group_name') or usage.get('category') or '未分类',
            'recipe_type': recipe.get('type') or usage.get('recipe_type'),
            'recipe_description': recipe.get('description'),
            'recipe_weight': recipe.get('weight'),
            'portion_size': recipe.get('portion_size'),
            'recipe_cooking_time': recipe.get('cooking_time'),
            'pot_type': recipe.get('pot_type'),
            'power_type': recipe.get('power_type'),
            'max_power': recipe.get('max_power'),
            'execution_count': usage.get('execution_count'),
            'device_count': usage.get('device_count'),
            'customer_count': usage.get('customer_count'),
        }

    def step_row(recipe_id, section, index, step, recipe=None, usage=None):
        if not isinstance(step, dict):
            step = {'commands': str(step)}
        row = base(recipe_id, recipe, usage)
        row.update({
            'section': section,
            'step_index': index,
            'time': step.get('time'),
            'cook_time': step.get('cook_time'),
            'type': step.get('type'),
            'type_label': step_type_label(step.get('type')),
            'automatic': step.get('automatic'),
            'automatic_label': auto_label(step.get('automatic')),
            'power': step.get('power'),
            'speed': step.get('speed') or step.get('stirSpeed'),
            'stir': step.get('stir'),
            'stir_mode': step.get('stirMode'),
            'mode': step.get('mode'),
            'position': step.get('position'),
            'movepot': step.get('movepot'),
            'direction': step.get('direction'),
            'type_operation': step.get('typeOperation'),
            'thedof_time': step.get('thedofTime'),
            'ingredients_time': step.get('ingredientsTime'),
            'initial_temperature': step.get('initialTemperature'),
            'initial_temperature_array': step.get('initialTemperatureArray'),
            'commands': command_text(step),
            'execution_content': build_machine_execution_text(step),
            'raw_json': json_cell(step),
        })
        return row

    cook_rows = []
    wash_rows = []
    moisten_rows = []
    ingredient_rows = []
    prep_note_rows = []
    serve_note_rows = []
    temperature_rows = []
    raw_rows = []

    for recipe_id in recipe_ids:
        recipe = recipe_by_id.get(recipe_id, {})
        detail = detail_by_recipe_id.get(recipe_id, {})
        usage = usage_by_recipe_id.get(recipe_id, {})
        cook_steps = parse_json_array(detail.get('cook_steps'))
        wash_steps = parse_json_array(detail.get('wash_steps'))
        moisten_steps = parse_json_array(detail.get('moisten_steps'))
        ingredients = parse_json_array(detail.get('cooking_ingredient'))
        temperature_curve = parse_json_value(detail.get('temperature_curve'))

        for index, step in enumerate(cook_steps, start=1):
            cook_rows.append(step_row(recipe_id, '烹饪', index, step, recipe, usage))
        for index, step in enumerate(wash_steps, start=1):
            wash_rows.append(step_row(recipe_id, '洗锅', index, step, recipe, usage))
        for index, step in enumerate(moisten_steps, start=1):
            moisten_rows.append(step_row(recipe_id, '润锅', index, step, recipe, usage))
        for index, item in enumerate(ingredients, start=1):
            if not isinstance(item, dict):
                item = {'name': str(item)}
            ingredient_id = first_present(item, 'ingredientsId', 'Ingredients_id', 'ingredientId', 'id')
            ingredient_meta = ingredient_name_map.get(str(ingredient_id)) or {}
            row = base(recipe_id, recipe, usage)
            row.update({
                'ingredient_index': index,
                'cooking_step_id': first_present(item, 'cookingId', 'cooking_id'),
                'ingredient_id': ingredient_id,
                'ingredient_name': first_present(item, 'name', 'materialName', 'food_name', 'ingredient_name') or ingredient_meta.get('ingredients_name'),
                'dosage': first_present(item, 'ingredientsDosage', 'Ingredients_dosage', 'dosage', 'weight', 'num', 'amount'),
                'unit': first_present(item, 'ingredientsUnit', 'Ingredients_unit', 'unit', 'dosageUnit', 'unit_name'),
                'preprocess': first_present(item, 'ingredientsTodyw', 'Ingredients_todyw', 'preprocess'),
                'feeding_mode': first_present(item, 'feedingMode', 'feeding_mode'),
                'insideand': first_present(item, 'insideand', 'Insideand'),
                'position': item.get('position'),
                'error_dosage': first_present(item, 'errorDosage', 'error_dosage'),
                'automatic': first_present(item, 'automatic') if first_present(item, 'automatic') is not None else ingredient_meta.get('automatic'),
                'ingredient_type': ingredient_meta.get('ingredinent_type'),
                'category_1': ingredient_meta.get('categories_1'),
                'category_2': ingredient_meta.get('categories_2'),
                'raw_json': json_cell(item),
            })
            ingredient_rows.append(row)

        prep_note_rows.extend(note_rows(recipe_id, '备菜须知', detail.get('ingredient_note'), recipe, usage, base))
        serve_note_rows.extend(note_rows(recipe_id, '出菜须知', detail.get('serve_note'), recipe, usage, base))
        if isinstance(temperature_curve, list):
            for index, item in enumerate(temperature_curve, start=1):
                data = item if isinstance(item, dict) else {'value': item}
                row = base(recipe_id, recipe, usage)
                row.update({
                    'point_index': index,
                    'time': first_present(data, 'time', 'x', 'second', 'seconds'),
                    'temperature': first_present(data, 'temperature', 'temp', 'y'),
                    'raw_json': json_cell(data),
                })
                temperature_rows.append(row)

        raw = base(recipe_id, recipe, usage)
        raw.update({
            'cook_time': detail.get('cook_time'),
            'cook_steps_count': len(cook_steps),
            'wash_steps_count': len(wash_steps),
            'moisten_steps_count': len(moisten_steps),
            'ingredient_count': len(ingredients),
            'description': recipe.get('description'),
            'steps_describe': recipe.get('steps_describe'),
            'ingredients_total_dosage': recipe.get('ingredients_total_dosage'),
            'ingredient_note_json': detail.get('ingredient_note'),
            'serve_note_json': detail.get('serve_note'),
            'temperature_curve_json': detail.get('temperature_curve'),
            'initial_temperature': detail.get('initial_temperature'),
            'initial_temperature_array': detail.get('initial_temperature_array'),
            'cook_steps_json': detail.get('cook_steps'),
            'wash_steps_json': detail.get('wash_steps'),
            'moisten_steps_json': detail.get('moisten_steps'),
            'cooking_ingredient_json': detail.get('cooking_ingredient'),
            'detail_missing': not bool(detail),
        })
        raw_rows.append(raw)

    return {
        'cook_steps': cook_rows,
        'wash_steps': wash_rows,
        'moisten_steps': moisten_rows,
        'ingredients': ingredient_rows,
        'prep_notes': prep_note_rows,
        'serve_notes': serve_note_rows,
        'temperature_curve': temperature_rows,
        'raw_process': raw_rows,
    }

def empty_recipe_search_result(params, reason='未匹配到设备'):
    return {
        'params': {
            'scope': params['scope'],
            'scope_label': params['scope_label'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'span_days': params['span_days'],
            'limit': params['limit'],
            'matched_device_count': int(params.get('matched_device_count') or 0),
        },
        'summary': {
            'total_logs': 0,
            'device_count': 0,
            'recipe_count': 0,
            'customer_count': 0,
            'first_time': None,
            'last_time': None,
            'total_duration_seconds': 0,
            'total_duration_label': seconds_label(0),
            'empty_reason': reason,
        },
        'recipes': [],
        'customers': [],
        'devices': [],
        'recent_logs': [],
    }

def build_recipe_search_result(params):
    sn_filter = resolve_recipe_search_sns(params)
    params['sn_filter'] = sn_filter
    params['matched_device_count'] = len(sn_filter)
    if not sn_filter:
        return empty_recipe_search_result(params)

    from_sql, where_sql, args = recipe_search_sql_parts(params)
    limit = params['limit']

    summary = fetch_one(
        f"""
        SELECT COUNT(*) AS total_logs,
               COUNT(DISTINCT l.sn) AS device_count,
               COUNT(DISTINCT CASE WHEN l.recipe_id IS NULL OR l.recipe_id = 0 THEN l.recipe_name ELSE l.recipe_id END) AS recipe_count,
               COUNT(DISTINCT r.company_id) AS customer_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               SUM(CAST(l.time AS SIGNED)) AS total_duration_seconds
        {from_sql}
        WHERE {where_sql}
        """,
        tuple(args),
        source=True,
        database='btyc',
    ) or {}

    recipe_rows = fetch_all(
        f"""
        SELECT COALESCE(NULLIF(CAST(l.recipe_id AS CHAR), '0'), CONCAT('name:', COALESCE(l.recipe_name, ''))) AS recipe_key,
               MAX(NULLIF(l.recipe_id, 0)) AS recipe_id,
               COALESCE(MAX(mr.name), MAX(l.recipe_name), '未知菜谱') AS recipe_name,
               COALESCE(MAX(mr.group_name), '未分类') AS category,
               MAX(mr.type) AS recipe_type,
               COUNT(*) AS execution_count,
               COUNT(DISTINCT l.sn) AS device_count,
               COUNT(DISTINCT r.company_id) AS customer_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               SUM(CAST(l.time AS SIGNED)) AS total_duration_seconds,
               ROUND(AVG(CAST(l.time AS SIGNED)), 1) AS avg_duration_seconds
        {from_sql}
        WHERE {where_sql}
        GROUP BY recipe_key
        ORDER BY execution_count DESC, last_time DESC
        LIMIT %s
        """,
        tuple(args + [limit]),
        source=True,
        database='btyc',
    )
    recipe_rows = hydrate_recipe_search_resources(recipe_rows)

    customer_rows = fetch_all(
        f"""
        SELECT COALESCE(r.company_id, 0) AS company_id,
               COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
               COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region,
               COUNT(*) AS production_count,
               COUNT(DISTINCT l.sn) AS device_count,
               COUNT(DISTINCT CASE WHEN l.recipe_id IS NULL OR l.recipe_id = 0 THEN l.recipe_name ELSE l.recipe_id END) AS recipe_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time
        {from_sql}
        WHERE {where_sql}
        GROUP BY COALESCE(r.company_id, 0)
        ORDER BY production_count DESC, last_time DESC
        LIMIT 80
        """,
        tuple(args),
        source=True,
        database='btyc',
    )

    device_rows = fetch_all(
        f"""
        SELECT l.sn,
               MAX(r.name) AS device_name,
               MAX(r.robot_type) AS robot_type,
               MAX(r.latest_update_package) AS version,
               COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
               COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region,
               COUNT(*) AS production_count,
               COUNT(DISTINCT CASE WHEN l.recipe_id IS NULL OR l.recipe_id = 0 THEN l.recipe_name ELSE l.recipe_id END) AS recipe_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time
        {from_sql}
        WHERE {where_sql}
        GROUP BY l.sn
        ORDER BY production_count DESC, last_time DESC
        LIMIT 120
        """,
        tuple(args),
        source=True,
        database='btyc',
    )

    recent_logs = fetch_all(
        f"""
        SELECT l.id, l.sn, l.recipe_id, l.recipe_name, l.time AS duration_seconds,
               l.create_time, l.end_time,
               MAX(r.name) AS device_name,
               COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
               COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region
        {from_sql}
        WHERE {where_sql}
        GROUP BY l.id, l.sn, l.recipe_id, l.recipe_name, l.time, l.create_time, l.end_time
        ORDER BY l.create_time DESC
        LIMIT %s
        """,
        tuple(args + [min(limit, 200)]),
        source=True,
        database='btyc',
    )

    total_duration = int(summary.get('total_duration_seconds') or 0)
    return {
        'params': {
            'scope': params['scope'],
            'scope_label': params['scope_label'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'span_days': params['span_days'],
            'limit': limit,
            'matched_device_count': len(sn_filter),
        },
        'summary': {
            'total_logs': int(summary.get('total_logs') or 0),
            'device_count': int(summary.get('device_count') or 0),
            'recipe_count': int(summary.get('recipe_count') or 0),
            'customer_count': int(summary.get('customer_count') or 0),
            'first_time': summary.get('first_time'),
            'last_time': summary.get('last_time'),
            'total_duration_seconds': total_duration,
            'total_duration_label': seconds_label(total_duration),
        },
        'recipes': recipe_rows,
        'customers': customer_rows,
        'devices': device_rows,
        'recent_logs': recent_logs,
    }

def get_cached_recipe_search(params, force_refresh=False):
    now = time.time()
    if not force_refresh:
        cached = load_cached_recipe_search(params, now)
        if cached:
            return cached
    result = build_recipe_search_result(params)
    key = save_cached_recipe_search(params, result, now)
    result['cache'] = {
        'hit': False,
        'created_at': now,
        'ttl_seconds': RECIPE_SEARCH_CACHE_TTL_SECONDS,
        'cache_key': key,
    }
    return result

def recipe_analysis_step_group(row):
    section = row.get('section') or ''
    text = ' '.join(str(row.get(k) or '') for k in ['commands', 'execution_content', 'raw_json'])
    step_type = str(row.get('type') or '')
    automatic = str(row.get('automatic') or '')
    power = number_value(row.get('power')) or 0
    speed = number_value(row.get('speed')) or 0
    position = str(row.get('position') or '').strip()
    groups = []
    if section == '洗锅':
        groups.append('洗锅')
    if section == '润锅':
        groups.append('润锅')
    if step_type == '3' or power:
        groups.append('加热/功率')
    if speed:
        groups.append('搅拌')
    if position and position not in {'0', '0.0'}:
        groups.append('锅位')
    if step_type == '2' or automatic == '1' or '自动' in str(row.get('automatic_label') or ''):
        groups.append('自动投料')
    if step_type == '1' and automatic != '1':
        groups.append('人工/预制投料')
    if any(k in text for k in ['猪油', '菜籽油', '色拉油', '橄榄油', '油桶', '注油', 'add oil', 'oil']):
        groups.append('油/注油')
    if any(k in text for k in ['水淀粉', '清水', '饮用水', '开水', '冷水', '热水', '高汤', '汤汁', '汤底', '加水', 'add water']):
        groups.append('水/水淀粉')
    if any(k in text for k in ['盐', '鸡精', '味精', '生抽', '老抽', '蚝油', '醋', '料酒', '酱', '粉']):
        groups.append('调料')
    return groups or ['其他动作']

def build_recipe_analysis_classification(process_rows):
    all_steps = (
        process_rows.get('cook_steps', []) +
        process_rows.get('moisten_steps', []) +
        process_rows.get('wash_steps', [])
    )
    group_counter = Counter()
    classified_steps = []
    power_counter = Counter()
    position_counter = Counter()
    for row in all_steps:
        groups = recipe_analysis_step_group(row)
        for group in groups:
            group_counter[group] += 1
        power = row.get('power')
        speed = row.get('speed')
        position = row.get('position')
        if power not in (None, '', '0', '0.0'):
            power_counter[f"{power}W / 速度{speed or '-'} / 锅位{position or '-'}"] += 1
        if position not in (None, '', '0', '0.0'):
            position_counter[str(position)] += 1
        item = dict(row)
        item['resource_groups'] = groups
        item['resource_group_label'] = ' / '.join(groups)
        classified_steps.append(item)

    ingredient_rows = []
    feed_summary = {
        'total_count': 0,
        'liquid_count': 0,
        'liquid_amount': 0,
        'water_count': 0,
        'water_amount': 0,
        'oil_count': 0,
        'oil_amount': 0,
        'seasoning_count': 0,
        'seasoning_amount': 0,
    }
    for row in process_rows.get('ingredients', []):
        classes = ingredient_classes(row)
        amount = normalized_amount(row.get('dosage'), row.get('unit'))
        labels = []
        if classes['liquid']:
            labels.append('液料')
            feed_summary['liquid_count'] += 1
            feed_summary['liquid_amount'] += amount or 0
        if classes['water']:
            labels.append('水/汤汁')
            feed_summary['water_count'] += 1
            feed_summary['water_amount'] += amount or 0
        if classes['oil']:
            labels.append('油')
            feed_summary['oil_count'] += 1
            feed_summary['oil_amount'] += amount or 0
        if classes['seasoning']:
            labels.append('调料')
            feed_summary['seasoning_count'] += 1
            feed_summary['seasoning_amount'] += amount or 0
        if not labels:
            labels.append('食材')
        feed_summary['total_count'] += 1
        item = dict(row)
        item.update({
            'class_label': ' / '.join(labels),
            'normalized_amount': amount,
        })
        ingredient_rows.append(item)

    return {
        'step_groups': [{'group': group, 'count': count} for group, count in group_counter.most_common()],
        'power_profile': [{'profile': profile, 'count': count} for profile, count in power_counter.most_common()],
        'position_profile': [{'position': position, 'count': count} for position, count in position_counter.most_common()],
        'feed_summary': feed_summary,
        'classified_steps': classified_steps,
        'classified_ingredients': ingredient_rows,
    }

def enrich_recipe_logs_with_temperature(log_rows, span_days):
    if not log_rows:
        return {'logs': log_rows, 'summary': {'enabled': False, 'reason': '无生产记录'}}
    if span_days > 2:
        return {'logs': log_rows, 'summary': {'enabled': False, 'reason': '温度匹配仅在指定日期/两日内查询启用，避免批量下载日志拖慢页面'}}

    series_cache = {}
    matched = 0
    missing = 0
    for row in log_rows:
        sn = row.get('sn')
        cook_end = row.get('create_time')
        duration = int(row.get('duration_seconds') or 0)
        if not sn or not cook_end or not duration:
            row['temperature'] = {'covered': False, 'status': '缺少作业时间'}
            missing += 1
            continue
        cook_start = cook_end - timedelta(seconds=duration)
        day_key = (sn, cook_start.date().isoformat())
        if day_key not in series_cache or series_cache.get(day_key) is None:
            series_cache[day_key] = find_temperature_series_for_window(sn, cook_start - timedelta(minutes=5), cook_end + timedelta(minutes=5))
        payload = series_cache.get(day_key)
        if not payload:
            row['temperature'] = {'covered': False, 'status': '未找到覆盖该日期的日志温度'}
            missing += 1
            continue
        temp = cooking_temperature_summary_from_series(payload.get('series') or [], cook_start, cook_end)
        temp['file_id'] = payload.get('file', {}).get('id')
        temp['file_name'] = payload.get('file', {}).get('file_name')
        temp['coverage_start'] = payload.get('coverage', {}).get('start')
        temp['coverage_end'] = payload.get('coverage', {}).get('end')
        row['temperature'] = temp
        if temp.get('covered'):
            matched += 1
        else:
            missing += 1
    return {
        'logs': log_rows,
        'summary': {
            'enabled': True,
            'matched_count': matched,
            'missing_count': missing,
            'unit': '锅体温度 ℃，优先取 android 日志 温度:_x_y_z 第3位',
        },
    }

def build_recipe_analysis_result(params, recipe_id=None, recipe_name=''):
    sn_filter = resolve_recipe_search_sns(params)
    params['sn_filter'] = sn_filter
    params['matched_device_count'] = len(sn_filter)
    if not sn_filter:
        return {
            **empty_recipe_search_result(params),
            'summary': {'empty_reason': '未匹配到当前客户/设备范围'},
            'devices': [],
            'recent_logs': [],
            'daily_stats': [],
            'temperature_summary': {'enabled': False, 'reason': '无生产记录'},
            'process': {},
            'classification': {},
        }

    from_sql, where_sql, args = recipe_search_sql_parts(params)
    extra_conditions = []
    extra_args = []
    clean_recipe_name = (recipe_name or '').strip()
    if recipe_id:
        extra_conditions.append("l.recipe_id = %s")
        extra_args.append(int(recipe_id))
    elif clean_recipe_name:
        extra_conditions.append("(l.recipe_name = %s OR mr.name = %s)")
        extra_args.extend([clean_recipe_name, clean_recipe_name])
    else:
        raise HTTPException(status_code=400, detail="需要选择一个菜谱后再分析")
    if extra_conditions:
        where_sql = f"{where_sql} AND {' AND '.join(extra_conditions)}"
        args = args + extra_args

    summary = fetch_one(
        f"""
        SELECT MAX(NULLIF(l.recipe_id, 0)) AS recipe_id,
               COALESCE(MAX(mr.name), MAX(l.recipe_name), %s) AS recipe_name,
               COALESCE(MAX(mr.group_name), '未分类') AS category,
               COUNT(*) AS execution_count,
               COUNT(DISTINCT l.sn) AS device_count,
               COUNT(DISTINCT r.company_id) AS customer_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               SUM(CAST(l.time AS SIGNED)) AS total_duration_seconds,
               ROUND(AVG(CAST(l.time AS SIGNED)), 1) AS avg_duration_seconds
        {from_sql}
        WHERE {where_sql}
        """,
        tuple([clean_recipe_name] + args),
        source=True,
        database='btyc',
    ) or {}
    if not int(summary.get('execution_count') or 0):
        return {
            'params': {
                'scope': params['scope'],
                'scope_label': params['scope_label'],
                'keyword': params['keyword'],
                'recipe_keyword': params['recipe_keyword'],
                'start_date': params['start_date'],
                'end_date': params['end_date'],
                'matched_device_count': len(sn_filter),
            },
            'summary': {'empty_reason': '当前范围内没有这道菜的生产记录'},
            'devices': [],
            'recent_logs': [],
            'daily_stats': [],
            'temperature_summary': {'enabled': False, 'reason': '无生产记录'},
            'process': {},
            'classification': {},
        }

    devices = fetch_all(
        f"""
        SELECT l.sn,
               MAX(r.name) AS device_name,
               MAX(r.robot_type) AS robot_type,
               MAX(r.latest_update_package) AS version,
               COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
               COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region,
               COUNT(*) AS execution_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               ROUND(AVG(CAST(l.time AS SIGNED)), 1) AS avg_duration_seconds
        {from_sql}
        WHERE {where_sql}
        GROUP BY l.sn
        ORDER BY execution_count DESC, last_time DESC
        LIMIT 120
        """,
        tuple(args),
        source=True,
        database='btyc',
    )
    recent_logs = fetch_all(
        f"""
        SELECT l.id, l.sn, l.recipe_id, l.recipe_name, l.time AS duration_seconds,
               l.create_time, l.end_time, l.whether, l.manual,
               COALESCE(MAX(c.common_name), MAX(c.company_name), '未知客户') AS customer_name,
               COALESCE(MAX(c.geo_cityname), MAX(c.geo_pname), MAX(c.area_code), '') AS region
        {from_sql}
        WHERE {where_sql}
        GROUP BY l.id, l.sn, l.recipe_id, l.recipe_name, l.time, l.create_time, l.end_time, l.whether, l.manual
        ORDER BY l.create_time DESC
        LIMIT 100
        """,
        tuple(args),
        source=True,
        database='btyc',
    )
    daily_stats = fetch_all(
        f"""
        SELECT DATE(l.create_time) AS day,
               COUNT(*) AS execution_count,
               COUNT(DISTINCT l.sn) AS device_count,
               MIN(l.create_time) AS first_time,
               MAX(l.create_time) AS last_time,
               ROUND(AVG(CAST(l.time AS SIGNED)), 1) AS avg_duration_seconds
        {from_sql}
        WHERE {where_sql}
        GROUP BY DATE(l.create_time)
        ORDER BY day DESC
        LIMIT 180
        """,
        tuple(args),
        source=True,
        database='btyc',
    )

    total_duration = int(summary.get('total_duration_seconds') or 0)
    recipe_row = {
        'recipe_id': summary.get('recipe_id') or recipe_id,
        'recipe_name': summary.get('recipe_name') or clean_recipe_name,
        'category': summary.get('category') or '未分类',
        'execution_count': int(summary.get('execution_count') or 0),
        'device_count': int(summary.get('device_count') or 0),
        'customer_count': int(summary.get('customer_count') or 0),
        'first_time': summary.get('first_time'),
        'last_time': summary.get('last_time'),
        'total_duration_seconds': total_duration,
        'avg_duration_seconds': summary.get('avg_duration_seconds'),
    }
    process_rows = fetch_recipe_process_export_rows([recipe_row])
    classification = build_recipe_analysis_classification(process_rows)
    temperature_payload = enrich_recipe_logs_with_temperature(recent_logs, params.get('span_days') or 0)
    recent_logs = temperature_payload['logs']
    return {
        'params': {
            'scope': params['scope'],
            'scope_label': params['scope_label'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'matched_device_count': len(sn_filter),
        },
        'summary': {
            **recipe_row,
            'total_duration_label': seconds_label(total_duration),
        },
        'devices': devices,
        'recent_logs': recent_logs,
        'daily_stats': daily_stats,
        'temperature_summary': temperature_payload['summary'],
        'process': process_rows,
        'classification': classification,
    }

@app.get("/api/me")
def me(authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    return {"username": username, "role": ZHIKU_USERS[username].get('role', 'user'), "profile": get_profile(username)}

@app.patch("/api/profile")
def profile_update(payload: ProfileUpdate, request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    profile = update_profile(username, payload.display_name)
    log_event(username, 'profile_update', request, detail={'display_name': profile.get('display_name')})
    return {"profile": profile}

@app.get("/api/devices/search")
def device_lookup(
    request: Request,
    keyword: str = Query(''),
    limit: int = Query(50),
    refresh: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    result = search_devices_by_keyword(keyword, limit=limit, force_refresh=bool(refresh))
    log_event(
        username,
        'device_lookup_refresh' if refresh else 'device_lookup',
        request,
        detail={
            'keyword': keyword,
            'result_count': result.get('total'),
            'cache_hit': bool(result.get('cache', {}).get('hit')),
        },
    )
    return result

@app.get("/api/device-versions")
def device_versions(
    request: Request,
    version: str = Query(''),
    keyword: str = Query(''),
    limit: int = Query(500),
    refresh: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    result = build_device_version_stats(version=version, keyword=keyword, limit=limit, force_refresh=bool(refresh))
    log_event(
        username,
        'device_versions_refresh' if refresh else 'device_versions',
        request,
        detail={
            'version': version,
            'keyword': keyword,
            'limit': limit,
            'cache_hit': bool(result.get('cache', {}).get('hit')),
            'total_devices': result.get('summary', {}).get('total_devices'),
            'version_count': result.get('summary', {}).get('version_count'),
        },
    )
    return result

@app.get("/api/search/{sn}")
def search_device(sn: str, request: Request, refresh: int = 0, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    report = get_cached_report(sn, force_refresh=bool(refresh))
    try:
        mark_device_watched(report['stats']['sn'], username=username, priority=30)
        queued = queue_recent_log_packages_for_device(report['stats']['sn'], limit=2)
        if queued:
            kick_auto_parse_once()
            report['structured_summary'] = device_structured_summary(report['stats']['sn'])
    except Exception as exc:
        print(f"watch device failed: {exc}")
    log_event(
        username,
        'search_refresh' if refresh else 'search',
        request,
        sn=report['stats']['sn'],
        detail={
            'input_sn': sn,
            'cache_hit': bool(report.get('cache', {}).get('hit')),
            'total_logs': report.get('stats', {}).get('total_logs'),
            'recipe_count': len(report.get('recipe_archive', [])),
        },
    )
    return report

@app.get("/api/search/{sn}/month/{month}")
def search_device_month(sn: str, month: str, request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    detail = get_device_month_detail(sn, month)
    log_event(
        username,
        'search_month',
        request,
        sn=detail['sn'],
        detail={
            'input_sn': sn,
            'month': month,
            'total_logs': detail.get('summary', {}).get('total_logs'),
            'recipe_count': detail.get('summary', {}).get('recipe_count'),
        },
    )
    return detail

@app.get("/api/cook-temperature/{sn}")
def cook_temperature(sn: str, request: Request, file_id: int = Query(None), refresh: int = Query(0), authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    mark_device_watched(real_sn, username=username, priority=20)
    result = build_cook_temperature_analysis(sn, file_id=file_id, force_refresh=bool(refresh) and is_admin(username))
    gc.collect()
    log_event(
        username,
        'cook_temperature_refresh' if refresh else 'cook_temperature',
        request,
        sn=result.get('sn'),
        detail={
            'file_id': result.get('file', {}).get('id'),
            'cook_id': result.get('cook', {}).get('id'),
            'recipe_id': result.get('cook', {}).get('recipe_id'),
            'step_count': len(result.get('steps', [])),
            'cache_hit': bool(result.get('cache', {}).get('hit')),
        },
    )
    return result

@app.get("/api/cook-temperature-structured/{sn}")
def cook_temperature_structured(sn: str, request: Request, day: str = Query(None), limit: int = Query(300), authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    mark_device_watched(real_sn, username=username, priority=25)
    target_day = day or datetime.now().strftime("%Y-%m-%d")
    result = structured_cook_temperature_by_day(real_sn, target_day, limit=limit)
    result['structured_summary'] = device_structured_summary(real_sn)
    log_event(
        username,
        'cook_temperature_structured',
        request,
        sn=real_sn,
        detail={
            'day': target_day,
            'cook_count': result.get('cook_count'),
            'package_count': result.get('summary_day', {}).get('package_count'),
        },
    )
    return result

@app.post("/api/structured/watch/{sn}")
def watch_structured_device(sn: str, request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    mark_device_watched(real_sn, username=username, priority=10)
    logs = get_device_log_files(real_sn)
    queued = queue_recent_log_packages_for_device(real_sn, limit=3)
    if queued:
        kick_auto_parse_once()
    summary = device_structured_summary(real_sn)
    log_event(username, 'structured_watch', request, sn=real_sn, detail={'queued': queued, 'log_count': len(logs)})
    return {'sn': real_sn, 'queued': queued, 'summary': summary}

@app.get("/api/temperature-calibrations/{sn}")
def temperature_calibrations(
    sn: str,
    request: Request,
    start_date: str = Query(None),
    end_date: str = Query(None),
    auto_queue: int = Query(1),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    end_text = end_date or datetime.now().strftime("%Y-%m-%d")
    start_text = start_date or (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    start, end = parse_date_range(start_text, end_text)
    mark_device_watched(real_sn, username=username, priority=15)
    rows = get_device_log_files_in_range(real_sn, start, end)
    queued = queue_temperature_calibration_scans(real_sn, rows, retry_failed=False) if auto_queue else 0
    if queued:
        kick_temperature_calibration_scan(real_sn)
    result = temperature_calibration_payload(real_sn, start, end, rows)
    result['newly_queued'] = queued
    log_event(
        username,
        'temperature_calibration_search',
        request,
        sn=real_sn,
        detail={
            'start_date': start_text,
            'end_date': end_text,
            'event_count': result.get('summary', {}).get('event_count'),
            'package_count': len(rows),
            'newly_queued': queued,
        },
    )
    return result

@app.post("/api/temperature-calibrations/{sn}/sync")
def sync_temperature_calibrations(
    sn: str,
    request: Request,
    start_date: str = Query(None),
    end_date: str = Query(None),
    retry_failed: int = Query(1),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    end_text = end_date or datetime.now().strftime("%Y-%m-%d")
    start_text = start_date or (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    start, end = parse_date_range(start_text, end_text)
    mark_device_watched(real_sn, username=username, priority=5)
    rows = get_device_log_files_in_range(real_sn, start, end)
    queued = queue_temperature_calibration_scans(real_sn, rows, retry_failed=bool(retry_failed))
    started = kick_temperature_calibration_scan(real_sn) if queued else False
    result = temperature_calibration_payload(real_sn, start, end, rows)
    result.update({'newly_queued': queued, 'worker_started': started})
    log_event(
        username,
        'temperature_calibration_sync',
        request,
        sn=real_sn,
        detail={
            'start_date': start_text,
            'end_date': end_text,
            'package_count': len(rows),
            'newly_queued': queued,
            'retry_failed': bool(retry_failed),
        },
    )
    return result

@app.get("/api/logs/detail")
def log_evidence_detail(
    request: Request,
    sn: str = Query(...),
    date: str = Query(''),
    source_file_id: int = Query(None),
    source_zip: str = Query(''),
    source_file: str = Query(...),
    timestamp: str = Query(''),
    line_hash: str = Query(''),
    line_no: int = Query(None),
    context_before: int = Query(20),
    context_after: int = Query(20),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    real_sn, robot = resolve_sn(sn)
    package, actual_source_file, lines = load_log_evidence_file(
        real_sn,
        source_file_id=source_file_id,
        source_zip=source_zip,
        source_file=source_file,
    )
    event = None
    if line_hash:
        event = fetch_one(
            """
            SELECT id, event_type, event_label, event_time, matched_keyword, source_log_name,
                   line_no, line_hash, raw_log_excerpt, evidence_level, event_hash
            FROM device_log_events
            WHERE sn=%s AND source_file_id=%s AND (line_hash=%s OR event_hash=%s)
            ORDER BY id DESC LIMIT 1
            """,
            (real_sn, int(package['id']), line_hash, line_hash),
        )
    effective_line_no = line_no or ((event or {}).get('line_no'))
    effective_timestamp = timestamp or str((event or {}).get('event_time') or '')
    raw_line = (event or {}).get('raw_log_excerpt') or ''
    target_idx = locate_log_target_line(
        real_sn,
        int(package['id']),
        actual_source_file,
        lines,
        line_hash=line_hash,
        line_no=effective_line_no,
        timestamp=effective_timestamp,
        raw_line=raw_line,
    )
    if target_idx is None:
        raise HTTPException(
            status_code=404,
            detail="日志文件已找到，但未能定位目标原始行；请重新扫描日志包后再试",
        )
    before = max(0, min(int(context_before or 20), 200))
    after = max(0, min(int(context_after or 20), 200))
    start_idx = max(0, target_idx - before)
    end_idx = min(len(lines), target_idx + after + 1)
    context_lines = []
    for idx in range(start_idx, end_idx):
        raw = lines[idx]
        context_lines.append({
            'line_no': idx + 1,
            'timestamp': parse_log_ts(raw),
            'line_hash': build_log_line_hash(real_sn, package['id'], actual_source_file, idx + 1, raw),
            'raw_line': raw,
            'is_target': idx == target_idx,
        })
    target_line = context_lines[target_idx - start_idx]
    device_meta = get_log_evidence_device_meta(real_sn, robot=robot)
    event_type = (event or {}).get('event_type') or 'log_evidence'
    matched_keyword = (event or {}).get('matched_keyword') or ''
    matched_job_id = find_matched_job_id(real_sn, package['id'], effective_timestamp)
    payload = {
        'meta': {
            'sn': real_sn,
            'customer': device_meta['customer'],
            'region': device_meta['region'],
            'date': date or str(effective_timestamp)[:10] or None,
            'timestamp': effective_timestamp or target_line.get('timestamp'),
            'source_zip': package.get('file_name'),
            'source_file_id': package.get('id'),
            'source_file': actual_source_file,
            'event_type': event_type,
            'event_label': (event or {}).get('event_label') or '原始日志证据',
            'matched_job_id': matched_job_id,
            'matched_job': bool(matched_job_id),
        },
        'target_line': {
            **target_line,
            'event_type': event_type,
            'matched_keyword': matched_keyword,
            'evidence_level': (event or {}).get('evidence_level') or ('高' if effective_line_no else '中'),
        },
        'context_lines': context_lines,
        'source_status': 'available',
        'matched_keyword': matched_keyword,
        'event_type': event_type,
        'evidence_level': (event or {}).get('evidence_level') or ('高' if effective_line_no else '中'),
    }
    log_event(
        username,
        'log_evidence_detail',
        request,
        sn=real_sn,
        detail={'source_file_id': package.get('id'), 'source_file': actual_source_file, 'line_no': target_idx + 1},
    )
    return payload

@app.get("/api/logs/search-in-file")
def search_log_evidence_file(
    request: Request,
    sn: str = Query(...),
    source_file_id: int = Query(None),
    source_zip: str = Query(''),
    source_file: str = Query(...),
    keyword: str = Query(...),
    limit: int = Query(200),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    real_sn, _ = resolve_sn(sn)
    clean_keyword = str(keyword or '').strip()
    if not clean_keyword:
        raise HTTPException(status_code=400, detail="请输入当前日志文件内的搜索关键词")
    package, actual_source_file, lines = load_log_evidence_file(
        real_sn,
        source_file_id=source_file_id,
        source_zip=source_zip,
        source_file=source_file,
    )
    safe_limit = max(1, min(int(limit or 200), 500))
    needle = clean_keyword.lower()
    matched_lines = []
    total = 0
    for idx, raw in enumerate(lines):
        if needle not in raw.lower():
            continue
        total += 1
        if len(matched_lines) < safe_limit:
            matched_lines.append({
                'line_no': idx + 1,
                'timestamp': parse_log_ts(raw),
                'line_hash': build_log_line_hash(real_sn, package['id'], actual_source_file, idx + 1, raw),
                'raw_line': raw,
            })
    log_event(
        username,
        'log_evidence_search',
        request,
        sn=real_sn,
        detail={'source_file_id': package.get('id'), 'source_file': actual_source_file, 'keyword': clean_keyword, 'count': total},
    )
    return {
        'matched_lines': matched_lines,
        'count': total,
        'returned_count': len(matched_lines),
        'source_status': 'available',
        'source_zip': package.get('file_name'),
        'source_file': actual_source_file,
    }

@app.get("/api/thermal-knowledge")
def thermal_knowledge(
    request: Request,
    keyword: str = Query(''),
    category: str = Query(''),
    hazard: str = Query(''),
    limit: int = Query(300),
    offset: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    ensure_analytics_db()
    limit = max(20, min(int(limit or 300), 1000))
    offset = max(0, int(offset or 0))
    where = []
    args = []
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        where.append("(canonical_name LIKE %s OR aliases_json LIKE %s OR source_category_1 LIKE %s OR source_category_2 LIKE %s OR hazard_class LIKE %s)")
        args.extend([like, like, like, like, like])
    if category.strip():
        where.append("category = %s")
        args.append(category.strip())
    if hazard.strip():
        where.append("hazard_class = %s")
        args.append(hazard.strip())
    where_sql = f"WHERE {' AND '.join(where)}" if where else ''
    total_row = fetch_one(f"SELECT COUNT(*) AS total FROM ingredient_thermal_properties {where_sql}", tuple(args)) or {}
    rows = fetch_all(
        f"""
        SELECT id, source_ingredient_id, canonical_name, aliases_json, category, source_category_1, source_category_2,
               ingredient_type, automatic, specific_heat_kj_kg_c, water_fraction, oil_fraction, boiling_c,
               smoke_point_c, flash_point_c, autoignition_c, hazard_class, confidence, source_note,
               recipe_usage_count, recipe_count, total_amount_g, total_amount_ml, last_seen_recipe_id, updated_at
        FROM ingredient_thermal_properties
        {where_sql}
        ORDER BY recipe_usage_count DESC, recipe_count DESC, canonical_name
        LIMIT %s OFFSET %s
        """,
        tuple(args + [limit, offset]),
    )
    items = []
    for row in rows:
        item = dict(row)
        try:
            item['aliases'] = json.loads(item.get('aliases_json') or '[]')
        except Exception:
            item['aliases'] = []
        items.append(item)
    summary = thermal_knowledge_summary()
    log_event(username, 'thermal_knowledge_query', request, detail={'keyword': keyword, 'category': category, 'hazard': hazard, 'limit': limit, 'offset': offset, 'total': total_row.get('total')})
    return {
        'summary': summary,
        'total': int(total_row.get('total') or 0),
        'limit': limit,
        'offset': offset,
        'items': items,
    }

@app.post("/api/thermal-knowledge/sync")
def sync_thermal_knowledge(request: Request, recipe_limit: int = Query(20000), authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    ensure_analytics_db()
    started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_id = None
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO ingredient_thermal_sync_runs(sync_type, status, created_by, started_at) VALUES (%s, %s, %s, %s)",
                ('manual_source_sync', 'RUNNING', username, started),
            )
            run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    try:
        base = sync_base_ingredients_to_thermal()
        recipe = sync_recipe_ingredients_to_thermal(limit=recipe_limit)
        finished = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        execute_local(
            "UPDATE ingredient_thermal_sync_runs SET status=%s, base_rows=%s, recipe_rows=%s, ingredient_rows=%s, finished_at=%s WHERE id=%s",
            ('COMPLETED', base.get('source_rows', 0), recipe.get('recipe_rows', 0), base.get('ingredient_rows', 0) + recipe.get('ingredient_rows', 0), finished, run_id),
        )
        result = {'run_id': run_id, 'base': base, 'recipe': recipe, 'summary': thermal_knowledge_summary()}
        log_event(username, 'thermal_knowledge_sync', request, detail=result)
        return result
    except Exception as exc:
        finished = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        execute_local(
            "UPDATE ingredient_thermal_sync_runs SET status=%s, error_message=%s, finished_at=%s WHERE id=%s",
            ('FAILED', str(exc), finished, run_id),
        )
        raise HTTPException(status_code=500, detail=f"热物性库同步失败：{exc}")

# ── Safety Scan APIs ──────────────────────────────────────────────

@app.get("/api/safety/overview")
def safety_overview_api(request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    ensure_analytics_db()
    try:
        result = safety_overview()
        log_event(username, 'safety_overview', request, detail={'summary': result.get('summary')})
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"安全总览查询失败：{exc}")


@app.get("/api/safety/alerts")
def safety_alerts_api(
    request: Request,
    risk_level: str = Query(None),
    rule_key: str = Query(None),
    sn: str = Query(None),
    dismissed: int = Query(None),
    limit: int = Query(200),
    offset: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    ensure_analytics_db()
    try:
        result = safety_alerts_query(
            risk_level=risk_level, rule_key=rule_key, sn=sn,
            dismissed=dismissed, limit=limit, offset=offset,
        )
        log_event(username, 'safety_alerts_query', request, detail={
            'risk_level': risk_level, 'rule_key': rule_key, 'sn': sn, 'total': result.get('total'),
        })
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"安全告警查询失败：{exc}")


@app.post("/api/safety/scan")
def trigger_safety_scan(request: Request, scan_type: str = Query('quick'), authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    ensure_analytics_db()
    try:
        result = run_safety_scan(scan_type=scan_type)
        log_event(username, 'safety_scan', request, detail=result)
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"安全扫描失败：{exc}")


@app.post("/api/safety/alerts/{alert_id}/dismiss")
def dismiss_safety_alert(alert_id: int, request: Request, authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    ensure_analytics_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    affected = execute_local(
        "UPDATE safety_scan_alerts SET dismissed=1, dismissed_by=%s, dismissed_at=%s WHERE id=%s AND dismissed=0",
        (username, now, alert_id),
    )
    if not affected:
        raise HTTPException(status_code=404, detail="告警不存在或已处理")
    log_event(username, 'safety_alert_dismiss', request, detail={'alert_id': alert_id})
    return {'dismissed': True, 'alert_id': alert_id}


@app.get("/api/safety/trends")
def safety_trends_api(
    request: Request,
    days: int = Query(7),
    sn: str = Query(None),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    ensure_analytics_db()
    try:
        sn_filter = "WHERE sn = %s" if sn else ""
        args = (sn,) if sn else ()
        trends = fetch_all(
            f"""
            SELECT stat_date, SUM(total_jobs) AS total_jobs,
                   SUM(high_temp_300c) AS high_temp_300c,
                   SUM(high_temp_330c) AS high_temp_330c,
                   MAX(max_temp_reached) AS max_temp_reached,
                   AVG(avg_temp_across_jobs) AS avg_temp_across_jobs
            FROM safety_daily_stats
            {sn_filter} {'AND' if sn else 'WHERE'} stat_date >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
            GROUP BY stat_date ORDER BY stat_date ASC
            """,
            args + (days,),
        )
        log_event(username, 'safety_trends', request, detail={'days': days, 'sn': sn, 'rows': len(trends or [])})
        return {
            'days': days,
            'sn': sn,
            'trends': [{**dict(r), 'stat_date': r['stat_date'].strftime('%Y-%m-%d') if hasattr(r['stat_date'], 'strftime') else str(r['stat_date'])} for r in (trends or [])],
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"安全趋势查询失败：{exc}")


@app.get("/api/recipe-search")
def recipe_search(
    request: Request,
    scope: str = Query('customer'),
    keyword: str = Query(''),
    recipe_keyword: str = Query(''),
    start_date: str = Query(None),
    end_date: str = Query(None),
    limit: int = Query(200),
    refresh: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    params = normalize_recipe_search_params(scope, keyword, recipe_keyword, start_date, end_date, limit)
    result = get_cached_recipe_search(params, force_refresh=bool(refresh))
    log_event(
        username,
        'recipe_search_refresh' if refresh else 'recipe_search',
        request,
        detail={
            'scope': params['scope'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'cache_hit': bool(result.get('cache', {}).get('hit')),
            'total_logs': result.get('summary', {}).get('total_logs'),
            'recipe_count': result.get('summary', {}).get('recipe_count'),
        },
    )
    return result

@app.get("/api/recipe-analysis")
def recipe_analysis(
    request: Request,
    scope: str = Query('customer'),
    keyword: str = Query(''),
    recipe_keyword: str = Query(''),
    start_date: str = Query(None),
    end_date: str = Query(None),
    limit: int = Query(200),
    recipe_id: int = Query(None),
    recipe_name: str = Query(''),
    refresh: int = Query(0),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    params = normalize_recipe_search_params(scope, keyword, recipe_keyword, start_date, end_date, limit)
    result = get_cached_recipe_analysis(params, recipe_id=recipe_id, recipe_name=recipe_name, force_refresh=bool(refresh))
    log_event(
        username,
        'recipe_analysis',
        request,
        detail={
            'scope': params['scope'],
            'keyword': params['keyword'],
            'recipe_id': recipe_id,
            'recipe_name': recipe_name,
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'cache_hit': bool(result.get('cache', {}).get('hit')),
            'execution_count': result.get('summary', {}).get('execution_count'),
        },
    )
    return result

@app.post("/api/analytics/recipe-top/jobs")
def create_recipe_top_job(payload: RecipeTopJobRequest, request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    params = normalize_recipe_top_params(payload)
    params_hash = analytics_params_hash(params)
    now = int(time.time())
    existing = None if payload.refresh and is_admin(username) else find_reusable_analytics_job('recipe_top', params_hash, now)
    if existing:
        job = get_analytics_job(existing['job_id'])
        log_event(username, 'recipe_top_query', request, detail={'job_id': job['job_id'], 'cache_hit': True, 'params': job.get('params')})
        return {**job, 'cache_hit': True}
    if payload.refresh and not is_admin(username):
        raise HTTPException(status_code=403, detail="只有 admin 可以强制刷新缓存")
    job_id = insert_analytics_job('recipe_top', params_hash, params, username)
    start_analytics_thread(job_id, params, username)
    job = get_analytics_job(job_id)
    log_event(username, 'recipe_top_query', request, detail={'job_id': job_id, 'cache_hit': False, 'params': {k: v for k, v in params.items() if k not in {'start_time', 'end_time'}}})
    return {**job, 'cache_hit': False}

@app.get("/api/analytics/jobs/{job_id}")
def analytics_job_status(job_id: str, authorization: str = Header(None)):
    require_auth(authorization=authorization)
    job = get_analytics_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/api/analytics/recipe-top/jobs/{job_id}/result")
def recipe_top_job_result(job_id: str, request: Request, authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    job = get_analytics_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job['status'] != 'COMPLETED':
        return {'job': job, 'result': None}
    path = Path(job.get('result_path') or '')
    if not path.exists():
        raise HTTPException(status_code=404, detail="Result file not found")
    result = json.loads(path.read_text())
    log_event(username, 'recipe_top_result', request, detail={'job_id': job_id, 'recipe_count': result.get('summary', {}).get('recipe_count')})
    return {'job': job, 'result': result}

@app.get("/api/analytics/recipe-top/jobs/{job_id}/export")
def recipe_top_job_export(job_id: str, request: Request, token: str = Query(None), authorization: str = Header(None)):
    username = require_auth(authorization=authorization, token=token)
    job = get_analytics_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job['status'] != 'COMPLETED':
        raise HTTPException(status_code=409, detail="Job not completed")
    path = Path(job.get('xlsx_path') or '')
    if not path.exists():
        result_path = Path(job.get('result_path') or '')
        if not result_path.exists():
            raise HTTPException(status_code=404, detail="Export file not found")
        result = json.loads(result_path.read_text())
        wb = build_recipe_top_workbook(result, job)
        path = analytics_xlsx_path(job_id)
        wb.save(path)
        update_analytics_job(job_id, xlsx_path=str(path))
    log_event(username, 'recipe_top_export', request, detail={'job_id': job_id, 'file_name': path.name})
    data = BytesIO(path.read_bytes())
    filename = f"zhiku_recipe_top_{job_id}.xlsx"
    return StreamingResponse(
        data,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

@app.get("/api/recipe-search/export")
def export_recipe_search(
    request: Request,
    scope: str = Query('customer'),
    keyword: str = Query(''),
    recipe_keyword: str = Query(''),
    start_date: str = Query(None),
    end_date: str = Query(None),
    limit: int = Query(200),
    token: str = Query(None),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization, token=token)
    params = normalize_recipe_search_params(scope, keyword, recipe_keyword, start_date, end_date, limit)
    result = get_cached_recipe_search(params)
    process_rows = fetch_recipe_process_export_rows(result.get('recipes', []))
    log_event(
        username,
        'recipe_search_export',
        request,
        detail={
            'scope': params['scope'],
            'keyword': params['keyword'],
            'recipe_keyword': params['recipe_keyword'],
            'start_date': params['start_date'],
            'end_date': params['end_date'],
            'total_logs': result.get('summary', {}).get('total_logs'),
        },
    )

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, '检索条件', [
        {'field': '检索维度', 'value': result['params'].get('scope_label')},
        {'field': '范围关键词', 'value': result['params'].get('keyword')},
        {'field': '菜谱关键词', 'value': result['params'].get('recipe_keyword')},
        {'field': '开始日期', 'value': result['params'].get('start_date')},
        {'field': '结束日期', 'value': result['params'].get('end_date')},
        {'field': '生产次数', 'value': result['summary'].get('total_logs')},
        {'field': '设备数', 'value': result['summary'].get('device_count')},
        {'field': '客户数', 'value': result['summary'].get('customer_count')},
        {'field': '菜谱数', 'value': result['summary'].get('recipe_count')},
    ], [('field', '字段'), ('value', '值')])
    write_sheet(wb, '菜谱汇总', result.get('recipes', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'),
        ('execution_count', '生产次数'), ('device_count', '设备数'), ('customer_count', '客户数'),
        ('first_time', '首次生产'), ('last_time', '末次生产'),
        ('total_duration_seconds', '累计耗时秒'), ('avg_duration_seconds', '平均耗时秒'),
        ('resources', '资源覆盖'), ('has_lard', '是否含猪油')
    ])
    write_sheet(wb, '客户汇总', result.get('customers', []), [
        ('company_id', '客户ID'), ('customer_name', '客户'), ('region', '地区'),
        ('production_count', '生产次数'), ('device_count', '设备数'), ('recipe_count', '菜谱数'),
        ('first_time', '首次生产'), ('last_time', '末次生产')
    ])
    write_sheet(wb, '设备汇总', result.get('devices', []), [
        ('sn', '设备SN'), ('device_name', '设备名'), ('customer_name', '客户'), ('region', '地区'),
        ('robot_type', '设备类型'), ('version', '版本'),
        ('production_count', '生产次数'), ('recipe_count', '菜谱数'),
        ('first_time', '首次生产'), ('last_time', '末次生产')
    ])
    write_sheet(wb, '明细样本', result.get('recent_logs', []), [
        ('id', '日志ID'), ('sn', '设备SN'), ('customer_name', '客户'), ('region', '地区'),
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'),
        ('duration_seconds', '耗时秒'), ('create_time', '生产时间'), ('end_time', '结束时间')
    ])
    write_sheet(wb, '菜谱烹饪步骤', process_rows.get('cook_steps', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('recipe_description', '备菜/预制说明'), ('recipe_weight', '菜谱总重量g'), ('portion_size', '份量'),
        ('recipe_cooking_time', '菜谱总时长秒'), ('pot_type', '锅类型'), ('power_type', '功率类型'), ('max_power', '最大功率'),
        ('execution_count', '生产次数'), ('device_count', '设备数'), ('customer_count', '客户数'),
        ('step_index', '步骤序号'), ('time', '时间'), ('cook_time', '烹饪时长'),
        ('type', '步骤类型编码'), ('type_label', '步骤类型'), ('automatic', '自动编码'), ('automatic_label', '自动/手动'),
        ('power', '功率'), ('speed', '搅拌速度'), ('stir', '搅拌'), ('stir_mode', '搅拌模式'), ('mode', '模式'),
        ('position', '锅位'), ('movepot', '翻锅/移锅'), ('direction', '方向'), ('type_operation', '操作类型'),
        ('thedof_time', '录菜总时长'), ('ingredients_time', '投料时间'), ('initial_temperature', '初始温度'),
        ('initial_temperature_array', '初始温度曲线'), ('commands', '原始指令内容'),
        ('execution_content', '机器执行内容'), ('raw_json', '原始步骤JSON')
    ])
    write_sheet(wb, '菜谱洗锅步骤', process_rows.get('wash_steps', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('step_index', '步骤序号'), ('time', '时间'), ('type', '步骤类型编码'), ('type_label', '步骤类型'),
        ('automatic', '自动编码'), ('automatic_label', '自动/手动'), ('power', '功率'), ('speed', '搅拌速度'),
        ('position', '锅位'), ('commands', '洗锅内容'), ('execution_content', '机器执行内容'), ('raw_json', '原始步骤JSON')
    ])
    write_sheet(wb, '菜谱润锅步骤', process_rows.get('moisten_steps', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('step_index', '步骤序号'), ('time', '时间'), ('type', '步骤类型编码'), ('type_label', '步骤类型'),
        ('automatic', '自动编码'), ('automatic_label', '自动/手动'), ('power', '功率'), ('speed', '搅拌速度'),
        ('position', '锅位'), ('commands', '润锅内容'), ('execution_content', '机器执行内容'), ('raw_json', '原始步骤JSON')
    ])
    write_sheet(wb, '菜谱配料明细', process_rows.get('ingredients', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('recipe_description', '备菜/预制说明'), ('ingredient_index', '配料序号'), ('cooking_step_id', '关联烹饪步骤ID'),
        ('ingredient_id', '食材ID'), ('ingredient_name', '食材名称'), ('dosage', '用量'), ('unit', '单位'),
        ('preprocess', '备菜处理方式'), ('feeding_mode', '投料模式'), ('insideand', '详情展示'),
        ('position', '位置'), ('error_dosage', '误差用量'), ('automatic', '是否自动'),
        ('ingredient_type', '食材类型'), ('category_1', '一级分类'), ('category_2', '二级分类'), ('raw_json', '原始配料JSON')
    ])
    write_sheet(wb, '备菜须知', process_rows.get('prep_notes', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('recipe_description', '备菜/预制说明'), ('note_index', '序号'), ('content', '备菜内容'), ('raw_json', '原始JSON')
    ])
    write_sheet(wb, '出菜须知', process_rows.get('serve_notes', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('note_index', '序号'), ('content', '出菜内容'), ('raw_json', '原始JSON')
    ])
    write_sheet(wb, '温度曲线', process_rows.get('temperature_curve', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('point_index', '点位序号'), ('time', '时间'), ('temperature', '温度'), ('raw_json', '原始JSON')
    ])
    write_sheet(wb, '菜谱原始过程', process_rows.get('raw_process', []), [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('category', '分类'), ('recipe_type', '类型'),
        ('execution_count', '生产次数'), ('device_count', '设备数'), ('customer_count', '客户数'),
        ('cook_time', '菜谱烹饪时长'), ('cook_steps_count', '烹饪步骤数'),
        ('wash_steps_count', '洗锅步骤数'), ('moisten_steps_count', '润锅步骤数'), ('ingredient_count', '配料数'),
        ('description', '备菜/预制说明'), ('steps_describe', '步骤描述'), ('ingredients_total_dosage', '配料总量'),
        ('ingredient_note_json', '备菜须知JSON'), ('serve_note_json', '出菜须知JSON'),
        ('temperature_curve_json', '温度曲线JSON'), ('initial_temperature', '初始温度'),
        ('initial_temperature_array', '初始温度曲线'),
        ('cook_steps_json', '烹饪步骤JSON'), ('wash_steps_json', '洗锅步骤JSON'),
        ('moisten_steps_json', '润锅步骤JSON'), ('cooking_ingredient_json', '配料JSON'),
        ('detail_missing', '详情是否缺失')
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"zhiku_recipe_search_{params['scope']}_{params['start_date']}_{params['end_date']}.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

def write_sheet(wb, title, rows, columns):
    ws = wb.create_sheet(title=title[:31])
    ws.append([label for _, label in columns])
    header_fill = PatternFill('solid', fgColor='E9EEF7')
    for cell in ws[1]:
        cell.font = Font(name='Arial', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append([normalize_cell(row.get(key)) for key, _ in columns])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 8
        for cell in col:
            max_len = max(max_len, min(len(str(cell.value or '')), 40))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

def normalize_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value

@app.get("/api/export/{sn}")
def export_device_report(sn: str, request: Request, token: str = Query(None), authorization: str = Header(None)):
    username = require_auth(authorization=authorization, token=token)
    report = get_cached_report(sn)
    log_event(
        username,
        'export',
        request,
        sn=report['stats']['sn'],
        detail={'recipe_count': len(report.get('recipe_archive', [])), 'log_count': len(report.get('logs', []))},
    )
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(wb, '设备档案', [{'field': k, 'value': v} for k, v in report['info'].items()], [('field', '字段'), ('value', '值')])
    write_sheet(wb, '客户信息', [{'field': k, 'value': v} for k, v in report.get('customer', {}).items()], [('field', '字段'), ('value', '值')])
    write_sheet(wb, '月度生产', report.get('monthly_summary', []), [
        ('month', '月份'), ('total_logs', '生产次数'), ('recipe_count', '菜谱数'),
        ('first_time', '首次生产'), ('last_time', '末次生产'),
        ('total_duration_seconds', '累计耗时秒'), ('avg_duration_seconds', '平均耗时秒')
    ])
    write_sheet(wb, '菜谱分类汇总', report.get('recipe_category_summary', []), [
        ('category', '分类'), ('recipe_count', '菜谱数'), ('execution_count', '执行次数'),
        ('lard_recipe_count', '含猪油菜谱数'), ('resources', '资源覆盖')
    ])
    write_sheet(wb, '菜谱归档', report.get('recipe_archive', []), [
        ('id', '菜谱ID'), ('name', '菜谱名称'), ('category', '分类'), ('execution_count', '执行次数'),
        ('total_duration_seconds', '累计耗时秒'), ('first_time', '首次执行'), ('last_time', '末次执行'),
        ('resources', '资源覆盖'), ('has_lard', '是否含猪油'), ('wash_steps_count', '洗锅步骤数'),
        ('moisten_steps_count', '润锅步骤数'), ('ingredient_count', '配料项数')
    ])

    step_rows = []
    for recipe in report.get('recipe_archive', []):
        for step in recipe.get('steps', []):
            step_rows.append({'recipe_id': recipe.get('id'), 'recipe_name': recipe.get('name'), **step})
    write_sheet(wb, '菜谱步骤', step_rows, [
        ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'), ('time', '时间点秒'),
        ('type', '步骤类型'), ('automatic', '自动'), ('power', '功率'), ('speed', '速度'),
        ('position', '锅位'), ('movepot', '移锅'), ('commands', '动作/投料')
    ])

    write_sheet(wb, '生产日志', report.get('logs', []), [
        ('id', '日志ID'), ('recipe_id', '菜谱ID'), ('recipe_name', '菜谱名称'),
        ('duration_seconds', '耗时秒'), ('create_time', '执行时间'), ('end_time', '结束时间'),
        ('expected_duration_seconds', '同菜谱平均耗时秒'), ('duration_ratio', '耗时倍率'),
        ('whether_label', '执行状态'), ('manual_label', '控制方式'), ('behavior_tags', '行为标签'),
        ('username', '账号/门店'), ('data_time', '数据时间'), ('comment', '备注')
    ])
    write_sheet(wb, '间隔分析', report.get('intervals', []), [
        ('id', '日志ID'), ('create_time', '执行时间'), ('recipe_name', '菜谱名称'), ('gap', '距上次间隔')
    ])
    write_sheet(wb, '故障日志', report.get('faults', []), [
        ('id', '故障ID'), ('create_time', '时间'), ('module', '模块'), ('details', '详情'), ('deal_state', '处理状态')
    ])
    write_sheet(wb, '维护记录', report.get('maintenance', []), [
        ('id', '记录ID'), ('create_time', '时间'), ('status', '状态'), ('mode', '模式'), ('duration', '时长')
    ])
    write_sheet(wb, '设备日志文件', report.get('device_logs', []), [
        ('id', '文件ID'), ('sn', '设备SN'), ('file_length', '文件大小'), ('file_name', '文件名'),
        ('create_time', '创建时间'), ('update_time', '修改时间'), ('type', '类型'),
        ('cos_deleted', '是否删除'), ('url', '下载URL')
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"zhiku_device_{report['stats']['sn']}.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

@app.get("/api/device-log-download/{file_id}")
def download_device_log(file_id: int, request: Request, token: str = Query(None), authorization: str = Header(None)):
    username = require_auth(authorization=authorization, token=token)
    row = fetch_one(
        "SELECT sn, file_name, file_length, pic AS url FROM machine_ftp WHERE id = %s LIMIT 1",
        (file_id,),
        source=True,
        database='btyc',
    )
    if not row or not row.get('url'):
        raise HTTPException(status_code=404, detail="Log file not found")
    log_event(
        username,
        'log_download',
        request,
        sn=row.get('sn'),
        detail={'file_id': file_id, 'file_name': row.get('file_name'), 'file_length': row.get('file_length')},
    )
    return RedirectResponse(row['url'])

def get_log_analysis_payload(file_id, request, username, refresh=False, allow_refresh=False, action='log_analysis_view'):
    row = fetch_one(
        "SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted "
        "FROM machine_ftp WHERE id = %s LIMIT 1",
        (file_id,),
        source=True,
        database='btyc',
    )
    if not row or not row.get('url'):
        raise HTTPException(status_code=404, detail="Log file not found")
    if row.get('cos_deleted'):
        raise HTTPException(status_code=410, detail="Log file has been deleted from COS")
    if refresh and not allow_refresh:
        raise HTTPException(status_code=403, detail="Only admin can refresh log analysis")

    if not refresh:
        cached = read_cached_log_analysis(file_id)
        if cached:
            cached['cache'] = {'hit': True}
            log_event(
                username,
                action,
                request,
                sn=row.get('sn'),
                detail={'file_id': file_id, 'cache_hit': True},
            )
            return cached

    zip_bytes = download_log_zip(row['url'])
    payload = build_log_analysis(row, zip_bytes)
    payload['cache'] = {'hit': False}
    save_cached_log_analysis(file_id, payload)
    log_event(
        username,
        action,
        request,
        sn=row.get('sn'),
        detail={
            'file_id': file_id,
            'cache_hit': False,
            'zip_size_mb': payload.get('zip_size_mb'),
            'file_name': row.get('file_name'),
        },
    )
    return payload

@app.get("/api/log-analysis/{file_id}")
def log_analysis(file_id: int, request: Request, refresh: int = Query(0), authorization: str = Header(None)):
    username = require_auth(authorization=authorization)
    return get_log_analysis_payload(
        file_id,
        request,
        username,
        refresh=bool(refresh),
        allow_refresh=is_admin(username),
        action='log_analysis_view',
    )

@app.get("/api/log-package-diagnostics/{file_id}")
def log_package_diagnostics(
    file_id: int,
    request: Request,
    refresh: int = Query(0),
    target_job_time: str = Query(None),
    authorization: str = Header(None),
):
    username = require_auth(authorization=authorization)
    if refresh and not is_admin(username):
        raise HTTPException(status_code=403, detail="Only admin can refresh log package diagnostics")
    row = fetch_one(
        "SELECT id, sn, file_name, file_length, pic AS url, type, create_time, update_time, cos_deleted "
        "FROM machine_ftp WHERE id=%s LIMIT 1",
        (file_id,),
        source=True,
        database='btyc',
    )
    if not row:
        raise HTTPException(status_code=404, detail={
            'failure_reason': 'machine_ftp_missing',
            'message': 'machine_ftp 中未找到日志包索引。',
            'suggested_next_action': '重新查询设备日志列表，确认 file_id 是否仍然有效。',
        })
    cached = None if refresh or target_job_time else read_cached_log_package_diagnostics(file_id)
    if cached:
        log_event(username, 'log_package_diagnostics', request, sn=row.get('sn'), detail={'file_id': file_id, 'cache_hit': True})
        return cached
    try:
        payload = build_log_package_diagnostics(row, target_job_time=target_job_time)
    except Exception as exc:
        payload = {
            'diagnostics_version': LOG_PACKAGE_DIAGNOSTICS_VERSION,
            'file_id': file_id,
            'sn': row.get('sn'),
            'file_name': row.get('file_name'),
            'diagnosis_level': 'error',
            'diagnosis_code': 'parser_error',
            'diagnosis_message': f'日志包诊断异常：{str(exc)[:300]}',
            'suggested_action': '保留该包并提交解析器排查，不要据此判断设备没有生产。',
            'internal_sessions': [],
        }
    if not target_job_time:
        save_cached_log_package_diagnostics(file_id, payload)
    payload['cache'] = {'hit': False}
    log_event(
        username,
        'log_package_diagnostics',
        request,
        sn=row.get('sn'),
        detail={
            'file_id': file_id,
            'cache_hit': False,
            'diagnosis_code': payload.get('diagnosis_code'),
            'internal_session_count': len(payload.get('internal_sessions') or []),
        },
    )
    return payload

@app.get("/api/admin/log-analysis/{file_id}")
def admin_log_analysis(file_id: int, request: Request, refresh: int = Query(0), authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    return get_log_analysis_payload(
        file_id,
        request,
        username,
        refresh=bool(refresh),
        allow_refresh=True,
        action='admin_log_analysis_view',
    )

@app.get("/api/admin/status")
def admin_status(authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    return {"status": "ok", "user": username, "cache_entries": len(REPORT_CACHE), **audit_summary()}

@app.get("/api/admin/emergency-devices")
def admin_emergency_devices(request: Request, authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    sns = ['0105222512180008', '0105222512180015']
    data = emergency_device_summary(sns)
    log_event(username, 'admin_emergency_view', request, detail={'sns': sns})
    return {'devices': data, 'sns': sns, 'generated_at': int(time.time())}

@app.get("/api/admin/oil-thermal/{sn}")
def admin_oil_thermal(sn: str, request: Request, authorization: str = Header(None)):
    username = require_admin(authorization=authorization)
    safe_sn = ''.join(ch for ch in sn if ch.isdigit())
    path = OIL_THERMAL_DATA_DIR / f'oil_thermal_{safe_sn}.json'
    if not path.exists():
        raise HTTPException(status_code=404, detail="Oil thermal log data not found")
    payload = json.loads(path.read_text())
    log_event(username, 'admin_oil_thermal_view', request, sn=safe_sn, detail={'source': str(path)})
    return payload

def auto_parse_cycle():
    if not ensure_analytics_db():
        return 0
    recover_stale_parsing_packages()
    rows = fetch_all(
        """
        SELECT p.sn, p.source_file_id, p.file_name
        FROM device_log_packages p
        JOIN watched_devices w ON w.sn = p.sn AND w.status = 'active'
        WHERE p.parse_status = 'queued'
          AND p.storage_status != 'stored'
          AND p.download_status = 'remote_available'
          AND COALESCE(p.file_size_mb, 0) > 0
          AND COALESCE(p.file_size_mb, 0) <= 50
        ORDER BY w.priority ASC, w.last_seen_at DESC, COALESCE(p.remote_create_time, p.log_time_hint) DESC
        LIMIT %s
        """,
        (max(1, AUTO_PARSE_MAX_PACKAGES_PER_CYCLE),),
    )
    handled = 0
    for row in rows:
        sn = row.get('sn')
        file_id = int(row.get('source_file_id'))
        try:
            mark_log_package_status(sn, file_id, parse_status='parsing')
            build_cook_temperature_analysis(sn, file_id=file_id, force_refresh=False)
            execute_local(
                "UPDATE watched_devices SET last_parse_at=%s, last_error=NULL, updated_at=%s WHERE sn=%s",
                (datetime.now().replace(microsecond=0), datetime.now().replace(microsecond=0), sn),
            )
        except HTTPException as exc:
            reason = str(exc.detail)
            parse_status = 'no_production_match' if exc.status_code == 404 and ('没有匹配到生产记录' in reason or 'DB 未匹配到生产记录' in reason) else 'parse_failed'
            mark_log_package_status(sn, file_id, parse_status=parse_status, error_message=reason)
            execute_local(
                "UPDATE watched_devices SET last_error=%s, updated_at=%s WHERE sn=%s",
                (reason[:500], datetime.now().replace(microsecond=0), sn),
            )
        except Exception as exc:
            reason = f"自动解析失败：{exc}"
            mark_log_package_status(sn, file_id, parse_status='parse_failed', error_message=reason)
            execute_local(
                "UPDATE watched_devices SET last_error=%s, updated_at=%s WHERE sn=%s",
                (reason[:500], datetime.now().replace(microsecond=0), sn),
            )
        handled += 1
        gc.collect()
    return handled

def auto_parse_worker():
    time.sleep(8)
    while AUTO_PARSE_ENABLED:
        try:
            auto_parse_cycle()
        except Exception as exc:
            print(f"auto parse worker cycle failed: {exc}")
        time.sleep(max(60, AUTO_PARSE_INTERVAL_SECONDS))

def kick_auto_parse_once():
    if not AUTO_PARSE_ENABLED:
        return
    def run_once():
        try:
            auto_parse_cycle()
        except Exception as exc:
            print(f"auto parse kick failed: {exc}")
    threading.Thread(target=run_once, name="zhiku-auto-parse-kick", daemon=True).start()

@app.on_event("startup")
def startup_tasks():
    global AUTO_PARSE_WORKER_STARTED
    ensure_analytics_db()
    if AUTO_PARSE_ENABLED:
        with AUTO_PARSE_WORKER_LOCK:
            if not AUTO_PARSE_WORKER_STARTED:
                thread = threading.Thread(target=auto_parse_worker, name="zhiku-auto-parse", daemon=True)
                thread.start()
                AUTO_PARSE_WORKER_STARTED = True

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
