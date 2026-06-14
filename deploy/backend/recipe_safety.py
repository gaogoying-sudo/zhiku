import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


RECIPE_SAFETY_RULE_VERSION = "recipe_safety_v1.1"
RECIPE_SAFETY_SHEET = "菜谱原始过程"

REQUIRED_COLUMNS = [
    "菜谱ID", "菜谱名称", "菜谱分类", "菜谱类型", "时间范围内生产次数", "设备数", "客户数",
    "详情烹饪时间", "烹饪步骤数", "洗锅步骤数", "润锅步骤数", "配料项数", "描述", "步骤描述",
    "配料总量", "备菜须知JSON", "出菜须知JSON", "温度曲线JSON", "初始温度", "初始温度数组",
    "烹饪步骤JSON", "洗锅步骤JSON", "润锅步骤JSON", "配料JSON", "详情是否缺失",
]

DEFAULT_RULE_CONFIG = {
    "full_power_w": 15000,
    "high_power_ratio": 0.8,
    "high_power_w": 12000,
    "high_power_seconds": 10,
    "short_high_power_seconds": 5,
    "long_high_power_seconds": 30,
    "all_cook_power_w": 10000,
    "all_cook_power_ratio": 0.95,
    "initial_high_temp_c": 150,
    "hot_temp_c": 300,
    "hot_temp_ratio": 0.5,
    "oil_window_seconds": 30,
    "frequent_production_count": 30,
    "frequent_device_count": 10,
    "frequent_customer_count": 5,
}

RULE_DEFS = {
    "R001": ("前两步高功率候选", "中", "烹饪前两步存在高功率持续输出。"),
    "R002": ("前置投料前高功率", "高", "第一次投料前存在高功率持续输出，疑似空锅加热候选。"),
    "R003": ("投油后高功率/高温候选", "高", "投油后短时间内存在高功率输出。"),
    "R004": ("高初始温度候选", "中", "初始温度或初始温度数组达到高温阈值。"),
    "R005": ("高功率长持续", "中", "单个步骤高功率持续时间较长。"),
    "R006": ("15kW候选", "高", "存在 15kW 或以上功率持续输出。"),
    "R007": ("润锅后高功率", "中", "存在润锅步骤且烹饪开始早期高功率输出。"),
    "R008": ("高频使用风险", "低", "生产次数、设备数或客户数达到高频阈值。"),
    "R009": ("重点品类关键词", "低", "菜谱名称或分类命中重点品类关键词。"),
    "R010": ("详情缺失", "低", "菜谱详情、烹饪步骤或配料数据缺失，需人工复核。"),
    "R011": ("润锅后未投油高功率", "高", "润锅结束后至投油前，80%至满功率运行 5 秒以上。"),
    "R012": ("投油后未投食材高功率", "高", "投油结束后至投放任意非油食材前，80%至满功率运行 5 秒以上。"),
    "R013": ("全程10kW以上烹饪", "高", "烹饪加热段几乎全程在 10kW 以上。"),
    "R014": ("300℃高温占比超50%", "高", "温度曲线中 300℃ 以上时间占总时间超过 50%。"),
}

OIL_KEYWORDS = [
    "油", "猪油", "牛油", "色拉油", "食用油", "菜籽油", "花生油", "Butter",
    "butter", "Oil", "oil", "油桶", "注油",
]
NON_FOOD_KEYWORDS = ["锅设置", "Set pot", "功率", "炒菜", "洗锅", "润锅"]
FOCUS_CATEGORY_KEYWORDS = ["炒饭", "辣子鸡", "牛肉粒", "干煸", "锅气", "锁汁", "煎", "炸", "Steak", "Fried Rice"]


def file_sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def as_int(value, default=0):
    text = as_text(value)
    if not text:
        return default
    try:
        return int(float(text))
    except Exception:
        return default


def as_float(value, default=None):
    text = as_text(value)
    if not text:
        return default
    try:
        return float(text)
    except Exception:
        return default


def as_bool(value):
    text = as_text(value).lower()
    return text in ("1", "true", "yes", "y", "是", "缺失")


def json_loads_safe(value, default=None):
    if default is None:
        default = []
    text = as_text(value)
    if not text:
        return default
    try:
        parsed = json.loads(text)
        return parsed if parsed is not None else default
    except Exception:
        return default


def json_dumps(value):
    return json.dumps(value if value is not None else [], ensure_ascii=False, separators=(",", ":"))


def has_oil(text):
    low = as_text(text).lower()
    return any(k.lower() in low for k in OIL_KEYWORDS)


def is_feed_step(step):
    commands = as_text(step.get("commands"))
    step_type = as_int(step.get("type"), -1)
    if step_type in (1, 2):
        return True
    if re.search(r"\d+(\.\d+)?\s*(g|克|kg|千克|ml|毫升|l|升)", commands, re.I):
        return not any(k.lower() in commands.lower() for k in NON_FOOD_KEYWORDS)
    return False


def is_food_feed_step(step):
    commands = as_text(step.get("commands"))
    return is_feed_step(step) and not has_oil(commands)


def extract_duration(step, steps, index, cook_time=0):
    current = as_float(step.get("time"), 0) or 0
    next_time = None
    for later in steps[index + 1:]:
        t = as_float(later.get("time"), None)
        if t is not None and t > current:
            next_time = t
            break
    if next_time is None and cook_time and cook_time > current:
        next_time = cook_time
    duration = (next_time - current) if next_time is not None else 0
    return max(0, int(round(duration)))


def normalize_steps(raw_steps, cook_time=0):
    if not isinstance(raw_steps, list):
        raw_steps = []
    sorted_steps = sorted(raw_steps, key=lambda row: (as_float(row.get("time"), 0) or 0, as_int(row.get("id"), 0)))
    rows = []
    for idx, step in enumerate(sorted_steps):
        power = as_float(step.get("power"), None)
        step_time = as_int(step.get("time"), 0)
        duration = extract_duration(step, sorted_steps, idx, cook_time)
        rows.append({
            "step_index": idx,
            "step_time": step_time,
            "inferred_duration_seconds": duration,
            "step_type": as_text(step.get("type")),
            "power_w": power,
            "power_kw": round(power / 1000, 3) if power is not None else None,
            "speed": as_text(step.get("speed")),
            "position": as_text(step.get("position")),
            "automatic": as_text(step.get("automatic")),
            "commands": as_text(step.get("commands")),
            "raw": step,
        })
    return rows


def parse_ingredient_names_by_cooking_step(steps):
    names = defaultdict(list)
    for step in steps:
        if not is_feed_step(step):
            continue
        commands = as_text(step.get("commands"))
        cooking_id = as_int(step.get("id"), None)
        if cooking_id is None:
            continue
        parts = [p.strip() for p in re.split(r"[,，、]", commands) if p.strip()]
        for part in parts:
            name = re.sub(r"\s*\d+(\.\d+)?\s*(g|克|kg|千克|ml|毫升|l|升).*", "", part, flags=re.I).strip()
            if name:
                names[cooking_id].append(name)
    return names


def normalize_ingredients(raw_ingredients, cook_steps):
    if not isinstance(raw_ingredients, list):
        raw_ingredients = []
    name_map = parse_ingredient_names_by_cooking_step([s.get("raw") or {} for s in cook_steps])
    rows = []
    for idx, item in enumerate(raw_ingredients):
        cooking_id = as_int(item.get("cookingId"), None)
        inferred_name = ""
        if cooking_id in name_map and name_map[cooking_id]:
            inferred_name = name_map[cooking_id].pop(0)
        ingredient_id = as_text(item.get("ingredientsId"))
        name = inferred_name or ingredient_id
        rows.append({
            "ingredient_index": idx,
            "ingredient_name": name,
            "ingredient_id": ingredient_id,
            "dosage": as_float(item.get("ingredientsDosage"), None),
            "unit": as_text(item.get("ingredientsUnit")),
            "automatic": as_text(item.get("insideand")),
            "feeding_mode": as_text(item.get("feedingMode")),
            "position": as_text(item.get("position")),
            "cooking_id": cooking_id,
            "raw": item,
            "matched_keywords": [k for k in OIL_KEYWORDS if k.lower() in name.lower()],
        })
    return rows


def normalize_temperature_points(raw_curve):
    points = []
    if not isinstance(raw_curve, list):
        return points
    for idx, item in enumerate(raw_curve):
        if isinstance(item, dict):
            t = as_float(item.get("time") or item.get("x") or item.get("offset") or idx, idx)
            temp = as_float(item.get("temperature") or item.get("temp") or item.get("y") or item.get("value"), None)
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            t = as_float(item[0], idx)
            temp = as_float(item[1], None)
        else:
            t = idx
            temp = as_float(item, None)
        if temp is not None:
            points.append({"time": float(t or 0), "temperature": float(temp)})
    return sorted(points, key=lambda row: row["time"])


def high_power_threshold(config):
    return float(config.get("high_power_w") or (float(config.get("full_power_w", 15000)) * float(config.get("high_power_ratio", 0.8))))


def evidence_step(recipe, step, explanation, evidence_type="step"):
    return {
        "recipe_id": recipe["recipe_id"],
        "recipe_name": recipe["recipe_name"],
        "evidence_type": evidence_type,
        "evidence_step_index": step.get("step_index") if step else None,
        "evidence_time": step.get("step_time") if step else None,
        "evidence_power_w": step.get("power_w") if step else None,
        "evidence_duration_seconds": step.get("inferred_duration_seconds") if step else None,
        "evidence_commands": step.get("commands") if step else "",
        "evidence_json": step.get("raw") if step else {},
        "explanation": explanation,
    }


def add_hit(hits, recipe, code, evidence=None):
    name, severity, desc = RULE_DEFS[code]
    item = {
        "rule_code": code,
        "rule_name": name,
        "severity": severity,
        "risk_level": severity,
        "explanation": desc,
    }
    item.update(evidence or {
        "recipe_id": recipe["recipe_id"],
        "recipe_name": recipe["recipe_name"],
        "evidence_type": "recipe",
        "evidence_step_index": None,
        "evidence_time": None,
        "evidence_power_w": None,
        "evidence_duration_seconds": None,
        "evidence_commands": "",
        "evidence_json": {},
    })
    hits.append(item)


def temperature_hot_ratio(points, threshold):
    if len(points) < 2:
        return 0.0
    hot = 0.0
    total = 0.0
    for idx, point in enumerate(points[:-1]):
        span = max(0, points[idx + 1]["time"] - point["time"])
        total += span
        if point["temperature"] >= threshold:
            hot += span
    return (hot / total) if total else 0.0


def analyze_recipe(recipe, config):
    steps = recipe["cook_steps"]
    ingredients = recipe["ingredients"]
    hits = []
    high_w = high_power_threshold(config)
    high_seconds = int(config["high_power_seconds"])
    short_seconds = int(config["short_high_power_seconds"])
    full_w = float(config["full_power_w"])
    first_feed_time = None
    first_food_time = None
    first_oil_time = None
    first_high_power_time = None
    max_power_w = 0
    matched_step_rules = defaultdict(list)

    for step in steps:
        p = step.get("power_w")
        if p is not None:
            max_power_w = max(max_power_w, p)
            if p >= high_w and first_high_power_time is None:
                first_high_power_time = step["step_time"]
        raw = step.get("raw") or {}
        if is_feed_step(raw) and first_feed_time is None:
            first_feed_time = step["step_time"]
        if has_oil(step.get("commands")) and first_oil_time is None:
            first_oil_time = step["step_time"]
        if is_food_feed_step(raw) and first_food_time is None:
            first_food_time = step["step_time"]

    for step in steps:
        p = step.get("power_w") or 0
        dur = step.get("inferred_duration_seconds") or 0
        if step["step_index"] <= 1 and p >= high_w and dur >= high_seconds:
            add_hit(hits, recipe, "R001", evidence_step(recipe, step, f"前两步第 {step['step_index']} 步 {p:.0f}W 持续约 {dur}s。"))
            matched_step_rules[step["step_index"]].append("R001")
        if first_feed_time is None or step["step_time"] < first_feed_time:
            if p >= high_w and dur >= high_seconds:
                add_hit(hits, recipe, "R002", evidence_step(recipe, step, f"首次投料前 {p:.0f}W 持续约 {dur}s。"))
                matched_step_rules[step["step_index"]].append("R002")
        if first_oil_time is not None and first_oil_time <= step["step_time"] <= first_oil_time + int(config["oil_window_seconds"]):
            if p >= high_w:
                add_hit(hits, recipe, "R003", evidence_step(recipe, step, f"投油后 {int(config['oil_window_seconds'])}s 内 {p:.0f}W 输出。"))
                matched_step_rules[step["step_index"]].append("R003")
        if p >= high_w and dur >= int(config["long_high_power_seconds"]):
            add_hit(hits, recipe, "R005", evidence_step(recipe, step, f"高功率 {p:.0f}W 持续约 {dur}s。"))
            matched_step_rules[step["step_index"]].append("R005")
        if p >= full_w and dur >= high_seconds:
            add_hit(hits, recipe, "R006", evidence_step(recipe, step, f"满功率候选 {p:.0f}W 持续约 {dur}s。"))
            matched_step_rules[step["step_index"]].append("R006")

    init_temp = recipe.get("initial_temperature")
    init_array = recipe.get("initial_temperature_array") or []
    init_values = [as_float(init_temp, None)] + [as_float(v, None) for v in init_array if as_float(v, None) is not None]
    if any(v is not None and v >= float(config["initial_high_temp_c"]) for v in init_values):
        add_hit(hits, recipe, "R004", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "temperature",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": None, "evidence_commands": "",
            "evidence_json": {"initial_temperature": init_temp, "initial_temperature_array": init_array},
            "explanation": f"初始温度达到 {config['initial_high_temp_c']}℃ 阈值。",
        })

    if recipe["moisten_steps"] and any((s.get("power_w") or 0) >= high_w for s in steps[:3]):
        add_hit(hits, recipe, "R007", evidence_step(recipe, next((s for s in steps[:3] if (s.get("power_w") or 0) >= high_w), None), "润锅后烹饪早期存在高功率输出。"))

    if (
        (recipe.get("production_count") or 0) >= int(config["frequent_production_count"])
        or (recipe.get("device_count") or 0) >= int(config["frequent_device_count"])
        or (recipe.get("customer_count") or 0) >= int(config["frequent_customer_count"])
    ):
        add_hit(hits, recipe, "R008", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "usage",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": None, "evidence_commands": "",
            "evidence_json": {"production_count": recipe.get("production_count"), "device_count": recipe.get("device_count"), "customer_count": recipe.get("customer_count")},
            "explanation": "生产、设备或客户覆盖达到高频阈值。",
        })

    recipe_text = f"{recipe['recipe_name']} {recipe.get('category') or ''}"
    if any(k.lower() in recipe_text.lower() for k in FOCUS_CATEGORY_KEYWORDS):
        add_hit(hits, recipe, "R009", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "keyword",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": None, "evidence_commands": recipe_text,
            "evidence_json": {"keywords": [k for k in FOCUS_CATEGORY_KEYWORDS if k.lower() in recipe_text.lower()]},
            "explanation": "菜谱名称或分类命中重点品类关键词。",
        })

    if recipe.get("detail_missing") or not recipe["cook_steps"] or not recipe["ingredients"]:
        add_hit(hits, recipe, "R010", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "missing_detail",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": None, "evidence_commands": "",
            "evidence_json": {"detail_missing": recipe.get("detail_missing"), "cook_steps": len(recipe["cook_steps"]), "ingredients": len(recipe["ingredients"])},
            "explanation": "详情缺失或关键 JSON 为空，需人工复核。",
        })

    moisten_end = 0
    for item in recipe["moisten_steps"]:
        moisten_end = max(moisten_end, as_int(item.get("rotatePotTime") or item.get("endTime") or item.get("time"), 0))
    if moisten_end:
        window_end = first_oil_time if first_oil_time is not None else recipe.get("cook_time") or 999999
        for step in steps:
            if moisten_end <= step["step_time"] < window_end and (step.get("power_w") or 0) >= high_w and (step.get("inferred_duration_seconds") or 0) >= short_seconds:
                add_hit(hits, recipe, "R011", evidence_step(recipe, step, f"润锅结束约 {moisten_end}s 后、投油前 {step.get('power_w'):.0f}W 持续约 {step.get('inferred_duration_seconds')}s。"))
                matched_step_rules[step["step_index"]].append("R011")
                break

    if first_oil_time is not None:
        window_end = first_food_time if first_food_time is not None and first_food_time > first_oil_time else recipe.get("cook_time") or 999999
        for step in steps:
            if first_oil_time <= step["step_time"] < window_end and (step.get("power_w") or 0) >= high_w and (step.get("inferred_duration_seconds") or 0) >= short_seconds:
                add_hit(hits, recipe, "R012", evidence_step(recipe, step, f"投油后至投放非油食材前 {step.get('power_w'):.0f}W 持续约 {step.get('inferred_duration_seconds')}s。"))
                matched_step_rules[step["step_index"]].append("R012")
                break

    powered = [s for s in steps if (s.get("power_w") or 0) > 0 and (s.get("inferred_duration_seconds") or 0) > 0]
    powered_seconds = sum(s.get("inferred_duration_seconds") or 0 for s in powered)
    power_10kw_seconds = sum(s.get("inferred_duration_seconds") or 0 for s in powered if (s.get("power_w") or 0) >= float(config["all_cook_power_w"]))
    all_high_ratio = (power_10kw_seconds / powered_seconds) if powered_seconds else 0
    if powered_seconds and all_high_ratio >= float(config["all_cook_power_ratio"]):
        add_hit(hits, recipe, "R013", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "power_ratio",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": int(powered_seconds), "evidence_commands": "",
            "evidence_json": {"powered_seconds": powered_seconds, "power_10kw_seconds": power_10kw_seconds, "ratio": all_high_ratio},
            "explanation": f"10kW 以上加热时长占加热总时长 {all_high_ratio:.1%}。",
        })

    hot_ratio = temperature_hot_ratio(recipe["temperature_points"], float(config["hot_temp_c"]))
    if hot_ratio >= float(config["hot_temp_ratio"]):
        add_hit(hits, recipe, "R014", {
            "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "evidence_type": "temperature_ratio",
            "evidence_step_index": None, "evidence_time": None, "evidence_power_w": None,
            "evidence_duration_seconds": None, "evidence_commands": "",
            "evidence_json": {"hot_temp_c": config["hot_temp_c"], "hot_ratio": hot_ratio, "sample_count": len(recipe["temperature_points"])},
            "explanation": f"{config['hot_temp_c']}℃ 以上温度占比 {hot_ratio:.1%}。",
        })

    weights = {"高": 35, "中": 20, "低": 8}
    seen_codes = []
    score = 0
    for hit in hits:
        if hit["rule_code"] not in seen_codes:
            seen_codes.append(hit["rule_code"])
            score += weights.get(hit["severity"], 5)
    score = min(100, score)
    high_codes = {"R002", "R003", "R006", "R011", "R012", "R013", "R014"}
    medium_codes = {"R001", "R004", "R005", "R007"}
    code_set = set(seen_codes)
    if code_set & high_codes or score >= 80:
        level = "高风险"
    elif code_set & medium_codes or score >= 50:
        level = "中风险"
    elif score >= 20 or code_set:
        level = "低风险"
    else:
        level = "无明显候选"

    tags = [RULE_DEFS[c][0] for c in seen_codes]
    return {
        "hits": hits,
        "matched_step_rules": matched_step_rules,
        "summary": {
            "recipe_id": recipe["recipe_id"],
            "recipe_name": recipe["recipe_name"],
            "category": recipe.get("category"),
            "production_count": recipe.get("production_count") or 0,
            "device_count": recipe.get("device_count") or 0,
            "customer_count": recipe.get("customer_count") or 0,
            "max_power_w": max_power_w or None,
            "max_power_kw": round(max_power_w / 1000, 3) if max_power_w else None,
            "first_feed_time": first_feed_time,
            "first_high_power_time": first_high_power_time,
            "initial_temperature": init_temp,
            "risk_score": score,
            "risk_level": level,
            "risk_tags": tags,
            "hit_rule_codes": seen_codes,
            "review_status": "待复核" if level in ("高风险", "中风险") or "R010" in code_set else "未复核",
            "review_note": "",
        },
    }


def parse_workbook(path, batch_code="", imported_by="script", config=None):
    config = {**DEFAULT_RULE_CONFIG, **(config or {})}
    path = Path(path)
    source_hash = file_sha256(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if RECIPE_SAFETY_SHEET not in wb.sheetnames:
        raise ValueError(f"缺少 sheet：{RECIPE_SAFETY_SHEET}")
    ws = wb[RECIPE_SAFETY_SHEET]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing:
        raise ValueError(f"缺少必要字段：{', '.join(missing)}")
    idx = {name: header.index(name) for name in header}
    raw_sheet_rows = [header]
    recipes, all_steps, all_ingredients, all_hits, all_evidence_steps, oil_hits = [], [], [], [], [], []
    review_rows = []
    failed = 0
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not values or not as_text(values[idx["菜谱ID"]]):
            continue
        raw_sheet_rows.append(list(values))
        try:
            cook_json = values[idx["烹饪步骤JSON"]]
            wash_json = values[idx["洗锅步骤JSON"]]
            moisten_json = values[idx["润锅步骤JSON"]]
            ingredient_json = values[idx["配料JSON"]]
            temp_curve_json = values[idx["温度曲线JSON"]]
            cook_time = as_int(values[idx["详情烹饪时间"]], 0)
            raw_cook_steps = json_loads_safe(cook_json, [])
            cook_steps = normalize_steps(raw_cook_steps, cook_time)
            recipe = {
                "recipe_id": as_text(values[idx["菜谱ID"]]),
                "recipe_name": as_text(values[idx["菜谱名称"]]),
                "category": as_text(values[idx["菜谱分类"]]),
                "recipe_type": as_text(values[idx["菜谱类型"]]),
                "production_count": as_int(values[idx["时间范围内生产次数"]], 0),
                "device_count": as_int(values[idx["设备数"]], 0),
                "customer_count": as_int(values[idx["客户数"]], 0),
                "cook_time": cook_time,
                "cook_step_count": as_int(values[idx["烹饪步骤数"]], len(cook_steps)),
                "wash_step_count": as_int(values[idx["洗锅步骤数"]], 0),
                "moisten_step_count": as_int(values[idx["润锅步骤数"]], 0),
                "ingredient_count": as_int(values[idx["配料项数"]], 0),
                "description": as_text(values[idx["描述"]]),
                "steps_description": as_text(values[idx["步骤描述"]]),
                "ingredient_total_weight": as_float(values[idx["配料总量"]], None),
                "prep_notes_json": json_loads_safe(values[idx["备菜须知JSON"]], []),
                "serve_notes_json": json_loads_safe(values[idx["出菜须知JSON"]], []),
                "temperature_curve_json": json_loads_safe(temp_curve_json, []),
                "initial_temperature": as_float(values[idx["初始温度"]], None),
                "initial_temperature_array": json_loads_safe(values[idx["初始温度数组"]], []),
                "cook_steps_json": raw_cook_steps,
                "wash_steps_json": json_loads_safe(wash_json, []),
                "moisten_steps_json": json_loads_safe(moisten_json, []),
                "ingredients_json": json_loads_safe(ingredient_json, []),
                "detail_missing": as_bool(values[idx["详情是否缺失"]]),
                "source_row_number": row_number,
            }
            recipe["cook_steps"] = cook_steps
            recipe["moisten_steps"] = recipe["moisten_steps_json"] if isinstance(recipe["moisten_steps_json"], list) else []
            recipe["wash_steps"] = recipe["wash_steps_json"] if isinstance(recipe["wash_steps_json"], list) else []
            recipe["ingredients"] = normalize_ingredients(recipe["ingredients_json"], cook_steps)
            recipe["temperature_points"] = normalize_temperature_points(recipe["temperature_curve_json"])
            result = analyze_recipe(recipe, config)
            recipe["risk_summary"] = result["summary"]
            recipes.append(recipe)
            for step in cook_steps:
                rules = result["matched_step_rules"].get(step["step_index"], [])
                step_row = {**step, "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"], "matched_rules": rules}
                all_steps.append(step_row)
                if rules:
                    all_evidence_steps.append(step_row)
            for ing in recipe["ingredients"]:
                ing_row = {**ing, "recipe_id": recipe["recipe_id"], "recipe_name": recipe["recipe_name"]}
                all_ingredients.append(ing_row)
                if ing["matched_keywords"] or has_oil(ing.get("ingredient_name")):
                    oil_hits.append(ing_row)
            all_hits.extend(result["hits"])
            if result["summary"]["review_status"] == "待复核":
                review_rows.append({
                    "recipe_id": recipe["recipe_id"],
                    "recipe_name": recipe["recipe_name"],
                    "risk_level": result["summary"]["risk_level"],
                    "risk_tags": "、".join(result["summary"]["risk_tags"]),
                    "reason": "、".join(result["summary"]["hit_rule_codes"]),
                    "suggested_action": "人工复核菜谱步骤、投油/投料窗口和高功率持续时间。",
                    "review_owner": "",
                    "review_status": result["summary"]["review_status"],
                    "review_note": "",
                })
        except Exception:
            failed += 1
    summaries = [r["risk_summary"] for r in recipes]
    overview = build_overview(summaries)
    return {
        "batch": {
            "batch_code": batch_code or f"recipe_safety_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            "source_file_name": path.name,
            "source_file_hash": source_hash,
            "source_sheet_name": RECIPE_SAFETY_SHEET,
            "row_count": len(recipes) + failed,
            "success_count": len(recipes),
            "failed_count": failed,
            "rule_version": RECIPE_SAFETY_RULE_VERSION,
            "imported_by": imported_by,
            "imported_at": datetime.now().replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S"),
            "status": "DRY_RUN",
            "error_message": "",
            "thresholds": config,
        },
        "raw_sheet_rows": raw_sheet_rows,
        "recipes": recipes,
        "steps": all_steps,
        "ingredients": all_ingredients,
        "hits": all_hits,
        "evidence_steps": all_evidence_steps,
        "oil_hits": oil_hits,
        "review_rows": review_rows,
        "summaries": summaries,
        "overview": overview,
        "rules": rule_rows(config),
    }


def build_overview(summaries):
    by_level = Counter(row["risk_level"] for row in summaries)
    rule_counts = Counter(code for row in summaries for code in row.get("hit_rule_codes", []))
    return {
        "total_recipes": len(summaries),
        "risk_recipes": sum(1 for row in summaries if row["risk_level"] != "无明显候选"),
        "high_risk": by_level["高风险"],
        "medium_risk": by_level["中风险"],
        "low_risk": by_level["低风险"],
        "missing_detail": rule_counts["R010"],
        "frequent_usage": rule_counts["R008"],
        "power_15kw": rule_counts["R006"],
        "after_oil_window_high_power": rule_counts["R003"],
        "after_oil_no_food_high_power": rule_counts["R012"],
        "after_oil_high_power": rule_counts["R003"] + rule_counts["R012"],
        "before_feed_high_power": rule_counts["R002"],
        "after_moisten_no_oil_high_power": rule_counts["R011"],
        "all_10kw": rule_counts["R013"],
        "hot_300_ratio": rule_counts["R014"],
    }


def rule_rows(config):
    rows = []
    for code, (name, severity, desc) in RULE_DEFS.items():
        rows.append({
            "rule_code": code,
            "rule_name": name,
            "rule_description": desc,
            "enabled": 1,
            "severity": severity,
            "config_json": json_dumps(config),
        })
    return rows


def write_cleaned_workbook(source_path, analysis, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = RECIPE_SAFETY_SHEET
    for row in analysis["raw_sheet_rows"]:
        ws.append([cell_value(v) for v in row])
    style_sheet(ws)

    write_sheet(wb, "清洗总览", [analysis["overview"]], [
        ("total_recipes", "总菜谱数"), ("risk_recipes", "命中风险菜谱数"), ("high_risk", "高风险数"),
        ("medium_risk", "中风险数"), ("low_risk", "低风险数"), ("missing_detail", "详情缺失数"),
        ("frequent_usage", "高频使用数"), ("power_15kw", "15kW候选数"),
        ("after_oil_window_high_power", "投油窗口高功率数"),
        ("after_oil_no_food_high_power", "投油后未投食材高功率数"),
        ("after_oil_high_power", "投油后高功率合计数"), ("before_feed_high_power", "投料前高功率数"),
        ("after_moisten_no_oil_high_power", "润锅后未投油高功率数"), ("all_10kw", "全程10kW以上数"),
        ("hot_300_ratio", "300℃高温占比数"),
    ])
    write_sheet(wb, "风险菜谱汇总", analysis["summaries"], [
        ("import_batch_id", "import_batch_id"), ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"),
        ("category", "分类"), ("production_count", "生产次数"), ("device_count", "设备数"),
        ("customer_count", "客户数"), ("max_power_w", "最大功率W"), ("max_power_kw", "最大功率kW"),
        ("first_feed_time", "首次投料时间"), ("first_high_power_time", "首次高功率时间"),
        ("initial_temperature", "初始温度"), ("risk_score", "risk_score"), ("risk_level", "risk_level"),
        ("risk_tags", "risk_tags"), ("hit_rule_codes", "hit_rule_codes"), ("review_status", "review_status"),
        ("review_note", "review_note"),
    ])
    write_sheet(wb, "规则命中明细", analysis["hits"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("rule_code", "规则编码"),
        ("rule_name", "规则名称"), ("risk_level", "风险等级"), ("evidence_type", "证据类型"),
        ("evidence_step_index", "步骤序号"), ("evidence_time", "证据时间"),
        ("evidence_power_w", "证据功率W"), ("evidence_duration_seconds", "持续秒"),
        ("evidence_commands", "动作/投料"), ("evidence_json", "证据JSON"), ("explanation", "说明"),
    ])
    write_sheet(wb, "命中步骤证据", analysis["evidence_steps"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("step_index", "步骤序号"),
        ("step_time", "时间点"), ("step_type", "类型"), ("power_w", "功率W"), ("power_kw", "功率kW"),
        ("speed", "速度"), ("position", "锅位"), ("automatic", "自动"), ("commands", "动作/投料"),
        ("matched_rules", "命中规则"),
    ])
    write_sheet(wb, "配料与油脂命中", analysis["oil_hits"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("ingredient_index", "配料序号"),
        ("ingredient_name", "配料名"), ("dosage", "用量"), ("unit", "单位"), ("automatic", "自动"),
        ("feeding_mode", "投料模式"), ("position", "位置"), ("raw", "原始JSON"), ("matched_keywords", "命中关键词"),
    ])
    write_sheet(wb, "待人工复核", analysis["review_rows"], [
        ("recipe_id", "菜谱ID"), ("recipe_name", "菜谱名称"), ("risk_level", "风险等级"),
        ("risk_tags", "风险标签"), ("reason", "原因"), ("suggested_action", "建议动作"),
        ("review_owner", "复核人"), ("review_status", "复核状态"), ("review_note", "备注"),
    ])
    write_sheet(wb, "清洗参数", [
        {"field": "文件名", "value": analysis["batch"]["source_file_name"]},
        {"field": "文件hash", "value": analysis["batch"]["source_file_hash"]},
        {"field": "导入时间", "value": analysis["batch"]["imported_at"]},
        {"field": "规则版本", "value": analysis["batch"]["rule_version"]},
        {"field": "阈值配置", "value": analysis["batch"]["thresholds"]},
        {"field": "总行数", "value": analysis["batch"]["row_count"]},
        {"field": "成功解析数", "value": analysis["batch"]["success_count"]},
        {"field": "解析失败数", "value": analysis["batch"]["failed_count"]},
    ], [("field", "字段"), ("value", "值")])
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)


def cell_value(value):
    if isinstance(value, (dict, list)):
        return json_dumps(value)
    return value


def write_sheet(wb, title, rows, columns):
    ws = wb.create_sheet(title=title[:31])
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([cell_value(row.get(key)) for key, _ in columns])
    style_sheet(ws)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="E9EEF7")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = 10
        for cell in col[:80]:
            width = max(width, min(len(str(cell.value or "")), 42))
        ws.column_dimensions[col[0].column_letter].width = width + 2
